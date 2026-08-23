from __future__ import annotations

import json
import logging
import os
import base64
import re
import time
import math
import csv
import io
import sqlite3
import subprocess
import zlib
import uuid
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from html import unescape
from urllib.parse import quote
from collections import Counter
from difflib import SequenceMatcher
from typing import Any
from xml.etree import ElementTree as ET

import httpx

from app.services.openai_http import (
    check_openai_connectivity_with_retries,
    post_chat_completions_with_retries,
)
from app.services.provider_config import (
    OPENROUTER_PROVIDER,
    normalize_generation_provider,
    openrouter_recovery_model,
    prepare_chat_payload,
    provider_candidate_concurrency,
    provider_model,
    provider_timeout_sec,
)
from app.services.openai_quality_config import (
    DEFAULT_QUALITY_MODEL,
    apply_quality_reasoning,
    quality_candidate_variants,
    quality_completion_budget,
)
from app.services.question_candidate_selection import select_question_candidates
from app.services.question_generation import (
    _editorial_realism_prompt_contract,
    _extract_json_text,
    _generate_questions_with_openai_from_ncs,
    _neutral_attitude_prompt_contract,
    _slice_balanced_json,
    _untrusted_context_prompt_contract,
    _unverified_material_precision_prompt_contract,
)
from app.services.question_intent import (
    FOCUS_SCOPED_GENERAL_QUESTION_INTENTS,
    GENERAL_QUESTION_INTENTS,
    classify_question_intent,
)
from app.services.question_surface import (
    build_question_task_frame,
    stable_ksa_evidence_id,
)
from app.services.ncs_mcp_client import (
    NcsMcpError,
    get_ksa_by_units,
    suggest_units_by_text,
)
from app.settings import settings

logger = logging.getLogger("ncscope.jd_strategy")


MOJIBAKE_ALIAS: dict[str, str] = {
    "珥앸Т": "총무",
    "珥앸Т?": "총무",
    "?먯궛愿由?": "자산관리",
    "?먯궛愿由": "자산관리",
    "?먯궛": "자산",
    "?뚭퀎쨌媛먯궗": "회계감사",
    "?뚭퀎": "회계",
    "?щТ?됱젙": "사무행정",
    "?щТ": "사무",
    "遺꾨쪟泥닿퀎": "분류체계",
    "?몃텇瑜?": "세분류",
    "?뚮텇瑜?": "소분류",
    "吏곷Т?섑뻾": "직무수행",
    "?λ젰?⑥쐞": "능력단위",
    "?꾩슂吏??": "필요지식",
    "?꾩슂湲곗닠": "필요기술",
    "臾몄꽌": "문서",
    "?됱젙": "행정",
}


def _count_hangul(text: str) -> int:
    return sum(1 for c in text if "\uac00" <= c <= "\ud7a3")


def _safe_tmp_root() -> str:
    root = os.path.join(os.getcwd(), ".tmp")
    os.makedirs(root, exist_ok=True)
    return root


def _safe_tmp_dir() -> str:
    path = os.path.join(_safe_tmp_root(), f"run_{uuid.uuid4().hex}")
    os.makedirs(path, exist_ok=True)
    return path


def _repair_mojibake(text: str) -> str:
    """Try to recover UTF-8 text that was decoded as latin-1/cp1252."""
    if not text:
        return text
    candidates = [text]
    for enc in ("latin-1", "cp1252"):
        try:
            repaired = text.encode(enc, errors="ignore").decode("utf-8", errors="ignore")
            if repaired:
                candidates.append(repaired)
        except Exception:
            pass
    best = max(candidates, key=_count_hangul)
    for broken, fixed in MOJIBAKE_ALIAS.items():
        best = best.replace(broken, fixed)
    return best


def extract_pdf_text(file_bytes: bytes) -> str:
    # 1) Preferred extractor: Python313 + pdfminer.
    py313 = r"C:\Python313\python.exe"
    if os.path.exists(py313):
        helper = (
            "from pdfminer.high_level import extract_text\n"
            "import sys\n"
            "t = extract_text(sys.argv[1]) or ''\n"
            "sys.stdout.buffer.write(t.encode('utf-8', 'ignore'))\n"
        )
        try:
            td = _safe_tmp_dir()
            try:
                pdf_path = os.path.join(td, "in.pdf")
                script_path = os.path.join(td, "extract_pdfminer.py")
                with open(pdf_path, "wb") as f:
                    f.write(file_bytes)
                with open(script_path, "w", encoding="utf-8") as f:
                    f.write(helper)
                p = subprocess.run(
                    [py313, script_path, pdf_path],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="ignore",
                    timeout=40,
                    check=False,
                )
                if p.returncode == 0:
                    text = _repair_mojibake((p.stdout or "").strip())
                    # Even if Korean glyph mapping is partially broken, this output is
                    # usually far better than raw binary stream fallback.
                    if len(text) >= 120:
                        return text
            finally:
                shutil.rmtree(td, ignore_errors=True)
        except Exception:
            pass

    # Vercel/Linux does not have the developer workstation's Python313
    # pdfminer helper.  pypdf is a declared runtime dependency, so use it
    # before the byte-stream/OCR fallbacks; this keeps a text-based PDF usable
    # even when the optional Node/Kordoc bridge is unavailable in a serverless
    # build.
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(file_bytes))
        pages: list[str] = []
        for page in reader.pages:
            try:
                text = str(page.extract_text() or "").strip()
            except Exception:
                text = ""
            if text:
                pages.append(text)
        extracted = _repair_mojibake("\n".join(pages).strip())
        if len(extracted) >= 120:
            return extracted
    except Exception:
        pass

    # 2) Best-effort standard-library fallback.
    content = file_bytes
    chunks: list[str] = []
    for match in re.finditer(rb"stream[\r\n]+(.*?)[\r\n]+endstream", content, flags=re.S):
        raw = match.group(1)
        stream_data = raw
        for _ in range(2):
            try:
                stream_data = zlib.decompress(stream_data)
                break
            except Exception:
                pass
        text = stream_data.decode("latin-1", errors="ignore")
        for token in re.findall(r"\(([^()]*)\)\s*T[Jj]", text):
            token = token.replace(r"\n", " ").replace(r"\r", " ").replace(r"\t", " ")
            token = token.replace(r"\(", "(").replace(r"\)", ")")
            if token.strip():
                chunks.append(token.strip())
    merged = _repair_mojibake("\n".join(chunks).strip())
    if merged:
        return merged
    # If stream parsing yields nothing, try offline OCR fallback (Windows OCR).
    if os.getenv("ENABLE_WINDOWS_OCR", "true").strip().lower() in {"1", "true", "yes", "y"}:
        try:
            ocr_pages = int(str(os.getenv("WINDOWS_OCR_MAX_PAGES", "2")).strip())
        except Exception:
            ocr_pages = 2
        ocr_text = _extract_pdf_text_via_windows_ocr(file_bytes=file_bytes, max_pages=max(1, min(3, ocr_pages)))
        if len(str(ocr_text or "").strip()) >= 10:
            return _repair_mojibake(ocr_text)

    # No readable text (image-only or unsupported encoding).
    return ""


def _parse_items(content_type: str, body: str) -> list[dict[str, Any]]:
    """
    Parse items from NCS API response (JSON or XML).

    Handles both formats:
    - JSON: response.body.items.item (single dict or list)
    - XML: <item> elements within <response>
    """
    if not body:
        return []

    # Determine format from content-type
    is_json = "json" in content_type.lower()

    if is_json:
        try:
            data = json.loads(body)
        except json.JSONDecodeError:
            return []

        # Navigate JSON structure: response.body.items.item
        items = data.get("response", {}).get("body", {}).get("items", {}).get("item")

        if items is None:
            return []

        # Convert single dict to list
        if isinstance(items, dict):
            return [items]
        elif isinstance(items, list):
            return items
        else:
            return []
    else:
        # XML parsing
        try:
            root = ET.fromstring(body)
        except ET.ParseError:
            return []

        # Extract all <item> elements
        items = []
        for item_elem in root.findall(".//item"):
            item_dict: dict[str, Any] = {}
            for child in item_elem:
                tag = child.tag
                text = (child.text or "").strip()
                item_dict[tag] = text
            if item_dict:
                items.append(item_dict)

        return items


def _render_pdf_pages_png_py313(file_bytes: bytes, max_pages: int = 2) -> list[bytes]:
    """PDF 페이지를 PNG로 렌더링. 현재 환경의 fitz 우선, 없으면 Python313 서브프로세스로 폴백."""
    # 1) 현재 Python 환경에 fitz(PyMuPDF)가 있으면 직접 사용 (가장 빠름)
    try:
        import fitz  # type: ignore
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        out: list[bytes] = []
        for i, page in enumerate(doc):
            if i >= max_pages:
                break
            pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=False)
            out.append(pix.tobytes("png"))
        doc.close()
        return out
    except ImportError:
        pass
    except Exception:
        # Current interpreter may not have working fitz; fall through to py313.
        pass

    # 2) fitz 없으면 Python313 서브프로세스로 폴백
    py313 = r"C:\Python313\python.exe"
    if not os.path.exists(py313):
        return []
    try:
        td = _safe_tmp_dir()
        try:
            pdf_path = os.path.join(td, "in.pdf")
            out_dir = os.path.join(td, "out")
            os.makedirs(out_dir, exist_ok=True)
            with open(pdf_path, "wb") as f:
                f.write(file_bytes)
            script = os.path.join(td, "render_pdf.py")
            code = (
                "import fitz, os, sys\n"
                "pdf_path, out_dir, max_pages = sys.argv[1], sys.argv[2], int(sys.argv[3])\n"
                "doc = fitz.open(pdf_path)\n"
                "for i, page in enumerate(doc):\n"
                "    if i >= max_pages:\n"
                "        break\n"
                "    pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=False)\n"
                "    pix.save(os.path.join(out_dir, f'page_{i+1}.png'))\n"
            )
            with open(script, "w", encoding="utf-8") as f:
                f.write(code)
            p = subprocess.run(
                [py313, script, pdf_path, out_dir, str(max_pages)],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="ignore",
                timeout=45,
                check=False,
            )
            if p.returncode != 0:
                return []
            out2: list[bytes] = []
            for name in sorted(os.listdir(out_dir)):
                if not name.lower().endswith(".png"):
                    continue
                with open(os.path.join(out_dir, name), "rb") as f:
                    out2.append(f.read())
            return out2
        finally:
            shutil.rmtree(td, ignore_errors=True)
    except Exception:
        return []


def extract_focus_terms_from_pdf_vision(
    file_bytes: bytes,
    max_pages: int = 2,
    api_key_override: str = "",
    generation_provider: str = "openai_api",
) -> list[str]:
    """
    Use OpenAI vision to extract role keywords when PDF text layer is broken.
    Returns canonical Korean terms suitable for NCS matching.
    """
    generation_provider = normalize_generation_provider(generation_provider)
    api_key = (
        settings.resolve_openrouter_key(api_key_override)
        if generation_provider == OPENROUTER_PROVIDER
        else settings.resolve_openai_key(api_key_override)
    )
    if not api_key:
        return []
    images = _render_pdf_pages_png_py313(file_bytes=file_bytes, max_pages=max_pages)
    if not images:
        return []

    content: list[dict[str, Any]] = [
        {
            "type": "text",
            "text": (
                "업로드된 직무기술서 이미지에서 직무 분류와 능력단위를 읽고, "
                "NCS 매핑용 핵심 키워드만 JSON으로 추출하세요. "
                "반드시 한국어 명사 키워드만 반환하세요. "
                "예: 총무, 자산관리, 사무행정, 회계감사, 회계처리, 문서관리, 계약관리, 구매관리.\n"
                "형식: {\"focus_terms\":[\"...\"]}"
            ),
        }
    ]
    for img in images:
        data_url = "data:image/png;base64," + base64.b64encode(img).decode("ascii")
        content.append({"type": "image_url", "image_url": {"url": data_url}})

    vision_model = provider_model(
        generation_provider,
        os.getenv("OPENAI_VISION_MODEL", "gpt-4o-mini").strip() or "gpt-4o-mini",
    )
    payload = {
        "model": vision_model,
        "messages": [
            {"role": "system", "content": "너는 직무기술서 분석기다. JSON만 출력한다."},
            {"role": "user", "content": content},
        ],
        "temperature": 0.0,
        "response_format": {"type": "json_object"},
    }
    try:
        data = post_chat_completions_with_retries(
            payload=prepare_chat_payload(payload, generation_provider),
            api_key=api_key,
            timeout_sec=provider_timeout_sec(generation_provider, 60.0),
            max_attempts=1,
            provider=generation_provider,
        )
        obj = json.loads(data["choices"][0]["message"]["content"])
        terms = obj.get("focus_terms", [])
        if not isinstance(terms, list):
            return []
        clean = []
        seen = set()
        for t in terms:
            t = str(t).strip()
            if len(t) < 2:
                continue
            if t not in seen:
                seen.add(t)
                clean.append(t)
        return clean[:20]
    except Exception:
        return []


def _tokenize(text: str) -> list[str]:
    text = _repair_mojibake(text)
    words = re.findall(r"[\uac00-\ud7a3A-Za-z0-9]{2,}", text)
    stop = {
        "및",
        "관련",
        "업무",
        "직무",
        "공공기관",
        "수행",
        "경험",
        "기술",
        "기반",
        "활용",
        "가능",
        "이해",
        "등",
    }
    return [w.lower() for w in words if w.lower() not in stop]


def _extract_focus_terms(jd_text: str) -> list[str]:
    raw_text = jd_text
    jd_text = _repair_mojibake(jd_text)
    lines = [ln.strip() for ln in jd_text.splitlines() if ln.strip()]

    terms: list[str] = []
    focus_labels = ["세분류", "소분류", "능력단위", "직무수행 내용", "필요지식", "필요기술"]
    for ln in lines:
        if any(label in ln for label in focus_labels):
            terms.extend(re.findall(r"[\uac00-\ud7a3]{2,}", ln))

    strong_seeds = [
        "총무",
        "자산관리",
        "사무행정",
        "회계",
        "회계감사",
        "문서관리",
        "행정지원",
        "재무회계",
        "구매",
        "비품",
        "재물조사",
        "전표",
        "결산",
    ]
    low = jd_text.lower()
    for s in strong_seeds:
        if s.lower() in low:
            terms.append(s)

    # Handle broken-text PDFs by direct alias detection.
    for broken, fixed in MOJIBAKE_ALIAS.items():
        if broken in jd_text and len(fixed) >= 2:
            terms.append(fixed)

    # Pattern-level rescue for broken Korean glyph mappings from HWP-origin PDFs.
    rescue_rules = [
        (["珥앸Т"], "총무"),
        (["먯궛", "鍮꾪뭹", "援щℓ", "臾쇳뭹", "재물조사"], "자산관리"),
        (["됱젙", "행정", "?щТ?됱젙"], "사무행정"),
        (["뚭퀎", "회계", "?꾪몴", "결산"], "회계처리"),
        (["媛먯궗", "감사"], "회계감사"),
        (["臾몄꽌", "문서"], "문서관리"),
    ]
    combined = f"{raw_text}\n{jd_text}"
    for needles, fixed in rescue_rules:
        if any(n in combined for n in needles):
            terms.append(fixed)

    dedup: list[str] = []
    seen = set()
    for t in terms:
        t = t.strip()
        if len(t) < 2:
            continue
        if t not in seen:
            dedup.append(t)
            seen.add(t)
    return dedup[:25]


def _dedup_keep_order(values: list[str]) -> list[str]:
    out: list[str] = []
    seen = set()
    for v in values:
        t = str(v or "").strip()
        if not t or t in seen:
            continue
        seen.add(t)
        out.append(t)
    return out


def _compact_line(text: str) -> str:
    return re.sub(r"\s+", "", str(text or ""))


def _sclass_norm_key(v: str) -> str:
    """Normalize sclass-like labels to a strict matching key."""
    n = _norm_text(v or "")
    if not n:
        return ""
    # absorb spacing/punctuation variants: "정보기술 운영" == "정보기술운영"
    return re.sub(r"[·‧･ㆍ•∙⋅\-\_/|(),.\[\]{}]", "", n)


def _collect_classification_lines(jd_text: str, max_lines: int = 90) -> list[str]:
    lines = [ln.strip() for ln in (jd_text or "").splitlines() if ln.strip()]
    if not lines:
        return []

    header_terms = ("분류체계", "대분류", "중분류", "소분류", "세분류")
    stop_terms = (
        "직무수행",
        "능력단위",
        "필요지식",
        "필요기술",
        "담당업무",
    )

    header_idx = [i for i, ln in enumerate(lines) if any(t in _compact_line(ln) for t in header_terms)]
    start = max(0, min(header_idx) - 1) if header_idx else 0
    # Some PDFs place "기관/주요사업" above 분류 헤더. Do not early-stop before header block ends.
    header_floor = max(header_idx) if header_idx else start
    end = len(lines)
    for i in range(start + 1, len(lines)):
        if i <= header_floor:
            continue
        if any(t in _compact_line(lines[i]) for t in stop_terms):
            end = i
            break
    return lines[start : min(end, start + max_lines)]


_CODE_NAME_PAIR_RE = re.compile(r"(?<!\d)(\d{1,2})\s*[.)]\s*([^\d]+?)(?=(?<!\d)\d{1,2}\s*[.)]|$)")


def _clean_category_value(text: str) -> str:
    t = str(text or "").strip()
    t = re.sub(r"\([^)]*\)", "", t)
    t = re.sub(r"\s+", " ", t).strip(" ,:;|-")
    return t


def _extract_small_categories_from_html_table(
    jd_text: str,
    known_categories: set[str] | list[str],
) -> list[str]:
    """Recover the ``소분류`` row from Kordoc's HTML table markdown.

    Kordoc preserves many PDF tables as HTML.  In that representation the
    classification row is structurally unambiguous, but the generic line
    scanner sees closing tags (for example ``재무</td></tr>``) as part of a
    category and may return a malformed seed.  Read only cells following the
    explicit ``소분류`` header and canonicalize them against the local NCS
    catalogue.  This is intentionally catalogue-backed so arbitrary text from
    a duty row cannot become an NCS category.
    """
    text = str(jd_text or "")
    if not text or "<tr" not in text.lower():
        return []

    by_key = {
        _sclass_norm_key(name): str(name).strip()
        for name in (known_categories or [])
        if _sclass_norm_key(str(name).strip())
    }
    if not by_key:
        return []

    def _cell_text(raw_cell: str) -> str:
        value = re.sub(r"<br\s*/?>", " ", str(raw_cell or ""), flags=re.IGNORECASE)
        value = re.sub(r"<[^>]+>", " ", value)
        value = unescape(value)
        value = re.sub(r"\s+", " ", value).strip()
        # Codes in the row are presentation metadata, not part of the label.
        value = re.sub(r"^\s*\.?\d{1,2}\s*[.)]?\s*", "", value)
        return _clean_category_value(value)

    recovered: list[str] = []
    for raw_row in re.findall(r"<tr\b[^>]*>(.*?)</tr>", text, flags=re.IGNORECASE | re.DOTALL):
        raw_cells = re.findall(r"<t[dh]\b[^>]*>(.*?)</t[dh]>", raw_row, flags=re.IGNORECASE | re.DOTALL)
        if not raw_cells:
            continue
        plain_cells = [_cell_text(cell) for cell in raw_cells]
        label_index = next(
            (
                index
                for index, cell in enumerate(plain_cells)
                if re.sub(r"\s+", "", cell) == "소분류"
            ),
            -1,
        )
        if label_index < 0:
            continue
        for cell in plain_cells[label_index + 1 :]:
            key = _sclass_norm_key(cell)
            canonical = by_key.get(key)
            if canonical:
                recovered.append(canonical)
        if recovered:
            break
    return _dedup_keep_order(recovered)


def _infer_column_major_counts(total: int, levels: int = 4) -> list[int] | None:
    if total <= 0 or levels <= 0:
        return None
    if levels != 4:
        return None
    if total < 3:
        return None
    if total <= 4:
        # [대,중,소,세]가 1개씩 혹은 일부만 보이는 단순 케이스
        return [1, 1, 1, max(1, total - 3)]

    best: list[int] | None = None
    best_cost = 10**9
    target = [1, 2, 3, 4]
    for c1 in range(1, total - 2):
        for c2 in range(c1, total - c1 - 1):
            for c3 in range(c2, total - c1 - c2):
                c4 = total - (c1 + c2 + c3)
                if c4 < c3 or c4 < 1:
                    continue
                cand = [c1, c2, c3, c4]
                cost = sum((cand[i] - target[i]) ** 2 for i in range(4))
                if cost < best_cost:
                    best_cost = cost
                    best = cand
    return best


def _extract_sclass_by_header_position(lines: list[str]) -> list[str]:
    """소분류 추출 - 두 가지 레이아웃 처리.

    1. 가로형: 소분류 라벨 오른쪽에 값 존재
       예) "소분류  총무" / "소분류: 총무, 일반사무"

    2. 세로형(헤더 블록): 대분류/중분류/소분류/세분류가 각 줄에 하나씩 연속으로 나열되고
       이후 값들이 따라오는 구조. 특히 첫 데이터 줄에 2쌍(중분류+소분류)이 오는 경우 처리.
       예) [대분류] [중분류] [소분류] [세분류] ... [06. 산업안전 01. 산업안전관리] [01. 기계 02. 전기 ...]
    """
    HEADERS = ["대분류", "중분류", "소분류", "세분류"]
    blocked = set(HEADERS) | {"분류체계"}

    # --- 가로형: 소분류 라벨 오른쪽에 값 ---
    for ln in lines:
        compact = re.sub(r"\s+", " ", ln).strip()
        m = re.match(r"소\s*분\s*류\s*[：:　\s]+(.+)", compact)
        if m:
            rest = m.group(1).strip()
            names = [_clean_category_value(n) for _, n in _CODE_NAME_PAIR_RE.findall(rest)]
            if not names:
                names = [_clean_category_value(p) for p in re.split(r"[,/|·]", rest)]
            names = [n for n in names if n and n not in blocked]
            if names:
                return _dedup_keep_order(names)

    # --- 세로형: 4개 헤더가 연속으로 있는 블록 감지 ---
    header_seq: list[int] = []  # line indices of 대중소세 in order
    for i, ln in enumerate(lines):
        compact = re.sub(r"\s+", "", ln)
        if compact in HEADERS:
            expected_idx = len(header_seq)
            if compact == HEADERS[expected_idx]:
                header_seq.append(i)
                if len(header_seq) == 4:
                    break
            else:
                header_seq = []  # 순서 깨지면 리셋

    if len(header_seq) < 3:  # 소분류까지만 있어도 처리
        return []

    block_end = header_seq[-1]
    value_lines = lines[block_end + 1:]
    if not value_lines:
        return []

    # 값 라인에서 첫 번째 코드-이름 쌍 라인 탐색
    for vln in value_lines:
        pairs = [_clean_category_value(n) for _, n in _CODE_NAME_PAIR_RE.findall(vln)]
        pairs = [p for p in pairs if p and p not in blocked and not p.isdigit()]
        if not pairs:
            continue

        n = len(pairs)
        if n == 2:
            # 중분류 + 소분류가 한 줄에 → 두 번째가 소분류
            return [pairs[1]] if pairs[1] else []
        elif n >= 4:
            # 대/중/소/세 한 줄에 → 세 번째가 소분류 (row-major는 기존 로직에 맡김)
            return []
        elif n == 1:
            # 열 단위/혼합 레이아웃에서는 다음 라인에 소분류가 이어질 수 있다.
            continue
        elif n == 3:
            # 중/소/세 또는 대/중/소 한 줄 → 두 번째가 소분류일 가능성
            return [pairs[1]] if pairs[1] else []

    return []


def _extract_small_categories_by_code_pairs(lines: list[str]) -> list[str]:
    if not lines:
        return []

    pair_rows: list[list[str]] = []
    flat_names: list[str] = []
    max_pairs_in_line = 0

    for ln in lines:
        pairs = [_clean_category_value(name) for _, name in _CODE_NAME_PAIR_RE.findall(ln)]
        pairs = [p for p in pairs if p and not p.isdigit()]
        if not pairs:
            continue
        max_pairs_in_line = max(max_pairs_in_line, len(pairs))
        pair_rows.append(pairs)
        flat_names.extend(pairs)

    if not flat_names:
        return []

    out: list[str] = []
    # 1) 열(컬럼) 단위로 텍스트가 쏟아지는 문서: 각 라인에 코드-값 1개
    if max_pairs_in_line == 1:
        counts = _infer_column_major_counts(total=len(flat_names), levels=4)
        if counts:
            start = counts[0] + counts[1]
            length = counts[2]
            out.extend(flat_names[start : start + length])
        elif len(flat_names) >= 3:
            out.append(flat_names[2])
        # 1-b) 일부 PDF는 [대,중,소,세,소,세,...] 순으로 압축되어 추출된다.
        #      이 경우 소분류는 index 2부터 2칸 간격으로 등장한다.
        interleaved = [flat_names[i] for i in range(2, len(flat_names), 2)]
        if interleaved:
            alias_index = _build_sclass_exact_alias_index()
            if alias_index:
                interleaved_valid: list[str] = []
                for v in interleaved:
                    key = _sclass_norm_key(v)
                    if key and key in alias_index:
                        interleaved_valid.append(v)
                # 기존 결과보다 유효 소분류가 더 많으면 interleaved 결과를 우선.
                if len(_dedup_keep_order(interleaved_valid)) > len(_dedup_keep_order(out)):
                    out = interleaved_valid
    else:
        # 2) 행 단위 표 문서: 한 줄에 대/중/소/세가 동시에 존재하거나 줄바꿈으로 일부 분리
        row_acc: list[str] = []
        for row in pair_rows:
            row_acc.extend(row)
            if len(row_acc) >= 3:
                out.append(row_acc[2])
            if len(row_acc) >= 4:
                row_acc = []

    cleaned = []
    blocked = {"대분류", "중분류", "소분류", "세분류", "분류체계"}
    for c in out:
        v = _clean_category_value(c)
        if not v or v in blocked:
            continue
        cleaned.append(v)
    return _dedup_keep_order(cleaned)


def _decide_sclass_anchor_scan_mode(lines: list[str], small_idx: int) -> str:
    """Decide scan direction around 소분류 anchor.

    Rule requested by user:
    - if headers look horizontal around 소분류 -> scan downward
    - if headers look vertical (중분류 above, 세분류 below close) -> scan rightward
    """
    compact = [_compact_line(ln) for ln in lines]
    mids = [i for i, ln in enumerate(compact) if "중분류" in ln]
    details = [i for i, ln in enumerate(compact) if "세분류" in ln]

    near_mid = min(mids, key=lambda x: abs(x - small_idx)) if mids else None
    near_detail = min(details, key=lambda x: abs(x - small_idx)) if details else None

    if near_mid is not None and near_detail is not None:
        up_dist = small_idx - near_mid
        down_dist = near_detail - small_idx
        # Vertical stack around 소분류 (중분류/세분류 above/below) -> scan right.
        if up_dist > 0 and down_dist > 0 and up_dist <= 3 and down_dist <= 3:
            return "right"
    # Default: horizontal header table -> scan downward by rows.
    return "down"


def _extract_anchor_line_terms(line: str) -> list[str]:
    line = str(line or "").strip()
    if not line:
        return []
    pairs = [_clean_category_value(n) for _, n in _CODE_NAME_PAIR_RE.findall(line)]
    pairs = [p for p in pairs if p and not p.isdigit()]
    if pairs:
        return pairs
    # Handle plain forms like "01 법무" (without dot/paren).
    cleaned_src = re.sub(r"^[•·▪◦\-\*]+\s*", "", line)
    cleaned_src = re.sub(r"^\d{1,2}\s*[.)]?\s+", "", cleaned_src).strip()
    cleaned = _clean_category_value(cleaned_src)
    if not cleaned:
        return []
    return [cleaned]


def _build_sclass_exact_alias_index(cache_ttl_sec: int = 60 * 30) -> dict[str, dict[str, Any]]:
    cache_key = "_sclass_exact_alias_index_cache"
    now = time.time()
    cached = globals().get(cache_key)
    if isinstance(cached, dict) and cached.get("items"):
        if (now - float(cached.get("ts", 0.0))) < cache_ttl_sec:
            return dict(cached.get("items", {}))

    catalog = load_sclass_catalog_from_csv()
    if not catalog:
        return {}
    synonym_pack = load_sclass_synonym_dictionary()
    synonym_by_code = synonym_pack.get("by_code_no", {})
    synonym_by_name = synonym_pack.get("by_name", {})

    index: dict[str, dict[str, Any]] = {}
    for row in catalog:
        code_no = str(row.get("ncs_code_no", "")).strip()
        name = str(row.get("ncs_sclass_name", "")).strip()
        if not (code_no and name):
            continue
        official_key = _sclass_norm_key(name)
        if official_key:
            index[official_key] = {"row": row, "official": True}
        aliases = _build_sclass_aliases(
            sclass_name=name,
            code_no=code_no,
            synonym_by_code=synonym_by_code,
            synonym_by_name=synonym_by_name,
        )
        for alias in aliases:
            k = _sclass_norm_key(alias)
            if not k:
                continue
            if k == official_key:
                continue
            # Keep first alias mapping; official key always overrides.
            if k not in index:
                index[k] = {"row": row, "official": False}

    globals()[cache_key] = {"ts": now, "items": index}
    return dict(index)


def _build_mclass_to_sclass_keys_index(cache_ttl_sec: int = 60 * 30) -> dict[str, set[str]]:
    """Build normalized middle-class -> small-class key index from local catalog."""
    cache_key = "_mclass_to_sclass_keys_index_cache"
    now = time.time()
    cached = globals().get(cache_key)
    if isinstance(cached, dict) and cached.get("items"):
        if (now - float(cached.get("ts", 0.0))) < cache_ttl_sec:
            cached_items = cached.get("items", {})
            if isinstance(cached_items, dict):
                return {str(k): set(v or set()) for k, v in cached_items.items()}

    catalog = load_sclass_catalog_from_csv()
    out: dict[str, set[str]] = {}
    for row in catalog:
        m_name = str(row.get("ncs_mclass_name", "")).strip()
        s_name = str(row.get("ncs_sclass_name", "")).strip()
        m_key = _sclass_norm_key(m_name)
        s_key = _sclass_norm_key(s_name)
        if not (m_key and s_key):
            continue
        out.setdefault(m_key, set()).add(s_key)

    globals()[cache_key] = {"ts": now, "items": out}
    return {str(k): set(v or set()) for k, v in out.items()}


def _extract_small_categories_by_vertical_blocks(
    lines: list[str],
    max_items: int = 15,
) -> list[str]:
    """Extract small categories for vertical-broken tables around 소분류.

    Pattern:
    - 중분류/소분류/세분류 headers are stacked vertically
    - values are read as sequential lines where 중분류 block and 소분류 rows are mixed

    Strategy:
    - split rows into middle-class blocks
    - within each block, keep only increasing code sequence (01 -> 02 -> 03 ...)
      as 소분류 candidates, and stop at reset (detail section starts)
    """
    if not lines:
        return []

    compact = [_compact_line(ln) for ln in lines]
    anchor_idxs = [i for i, ln in enumerate(compact) if "소분류" in ln]
    if not anchor_idxs:
        return []

    alias_index = _build_sclass_exact_alias_index()
    if not alias_index:
        return []
    mclass_index = _build_mclass_to_sclass_keys_index()

    mids = [i for i, ln in enumerate(compact) if "중분류" in ln]
    details = [i for i, ln in enumerate(compact) if "세분류" in ln]
    if not mids or not details:
        return []

    stop_terms = (
        "직무수행",
        "능력단위",
        "필요지식",
        "필요기술",
        "전형방법",
        "일반요건",
        "교육요건",
        "기타요건",
        "직무수행내용",
        "내용",
    )

    def _row_pairs(src: str) -> list[tuple[int, str]]:
        vals: list[tuple[int, str]] = []
        for code, name in _CODE_NAME_PAIR_RE.findall(src):
            cleaned = _clean_category_value(name)
            if not cleaned:
                continue
            try:
                num = int(str(code).strip())
            except Exception:
                continue
            vals.append((num, cleaned))
        return vals

    def _is_marker_row(pairs: list[tuple[int, str]]) -> tuple[bool, str, bool]:
        if not pairs:
            return False, "", False
        first_name = pairs[0][1]
        first_key = _sclass_norm_key(first_name)
        if first_key and first_key in mclass_index:
            return True, first_key, True
        if not first_key or first_key not in alias_index:
            return True, first_key, False
        return False, first_key, False

    out: list[str] = []
    seen: set[str] = set()

    for idx in anchor_idxs[:2]:
        near_mid = min(mids, key=lambda x: abs(x - idx))
        near_detail = min(details, key=lambda x: abs(x - idx))
        up_dist = idx - near_mid
        down_dist = near_detail - idx
        # Vertical header condition only.
        if not (up_dist > 0 and down_dist > 0 and up_dist <= 3 and down_dist <= 3):
            continue

        row_data: list[tuple[int, list[tuple[int, str]]]] = []
        for j in range(near_detail + 1, len(lines)):
            cj = compact[j]
            if any(t in cj for t in stop_terms):
                break
            pairs = _row_pairs(lines[j])
            if pairs:
                row_data.append((j, pairs))
        if not row_data:
            continue

        r = 0
        while r < len(row_data):
            _, pairs = row_data[r]
            marker, marker_key, marker_known = _is_marker_row(pairs)
            if not marker:
                r += 1
                continue

            allowed_keys = mclass_index.get(marker_key, set()) if marker_known else set()
            prev_num = 0
            block_terms: list[str] = []

            def _consume_pair(num: int, name: str) -> bool:
                nonlocal prev_num
                key = _sclass_norm_key(name)
                if not key or key not in alias_index:
                    return False
                if allowed_keys and key not in allowed_keys:
                    if "서무" not in _compact_line(name):
                        return False
                if prev_num and num <= prev_num:
                    return True
                prev_num = num
                disp = re.sub(r"[·‧･ㆍ•∙⋅]", "", _clean_category_value(name))
                if disp:
                    block_terms.append(disp)
                return False

            reset = False
            for num, name in pairs[1:]:
                reset = _consume_pair(num, name)
                if reset:
                    break

            r += 1
            while r < len(row_data) and not reset:
                _, next_pairs = row_data[r]
                next_marker, _, _ = _is_marker_row(next_pairs)
                if next_marker:
                    break
                for num, name in next_pairs:
                    reset = _consume_pair(num, name)
                    if reset:
                        break
                r += 1

            block_terms = _dedup_keep_order(block_terms)
            accepted = block_terms if marker_known else (block_terms if len(block_terms) >= 2 else [])
            for term in accepted:
                if term in seen:
                    continue
                seen.add(term)
                out.append(term)
                if len(out) >= max_items:
                    return out[:max_items]

    return out[:max_items]


def _extract_small_categories_by_anchor_direction(
    lines: list[str],
    down_scan_lines: int = 12,
    right_scan_lines: int = 8,
    max_items: int = 15,
) -> list[str]:
    """Anchor-based small-category extraction using scan direction heuristics."""
    if not lines:
        return []

    compact = [_compact_line(ln) for ln in lines]
    anchor_idxs = [i for i, ln in enumerate(compact) if "소분류" in ln]
    if not anchor_idxs:
        return []

    stop_down = (
        "세분류",
        "직무수행",
        "능력단위",
        "필요지식",
        "필요기술",
        "전형방법",
        "일반요건",
        "교육요건",
        "기타요건",
        "직무수행내용",
        "내용",
    )
    stop_right = (
        "중분류",
        "세분류",
        "직무수행",
        "능력단위",
        "필요지식",
        "필요기술",
        "전형방법",
        "일반요건",
        "교육요건",
        "기타요건",
    )
    header_only = ("대분류", "중분류", "소분류", "세분류", "분류체계", "채용분야", "구분")
    blocked = {"대분류", "중분류", "소분류", "세분류", "분류체계", "구분", "직무", "내용"}

    raw_terms: list[tuple[str, bool]] = []
    for idx in anchor_idxs[:2]:
        mode = _decide_sclass_anchor_scan_mode(lines, idx)
        max_scan = down_scan_lines if mode == "down" else right_scan_lines
        stop_terms = stop_down if mode == "down" else stop_right

        line = lines[idx]
        m = re.search(r"소\s*분\s*류\s*[：:\s]+\s*(.+)$", re.sub(r"\s+", " ", line).strip())
        if m:
            raw_terms.extend([(x, True) for x in _extract_anchor_line_terms(m.group(1))])

        scanned = 0
        for j in range(idx + 1, len(lines)):
            c = compact[j]
            if scanned == 0 and any(h in c for h in header_only):
                continue
            if any(t in c for t in stop_terms):
                # In vertical-header layouts, immediate stop token can appear before data rows.
                if mode == "right" and scanned == 0:
                    continue
                break
            line_terms = _extract_anchor_line_terms(lines[j])
            line_explicit_small = False
            # Row pattern in some PDFs:
            #   "01. 소분류  02. 세분류" (same physical row, collapsed into one line)
            # In this case keep the first pair for 소분류 scanning.
            pair_terms = [_clean_category_value(n) for _, n in _CODE_NAME_PAIR_RE.findall(lines[j])]
            pair_terms = [p for p in pair_terms if p and not p.isdigit()]
            if mode == "down" and len(pair_terms) == 2 and scanned <= 1:
                line_terms = [pair_terms[0]]
                line_explicit_small = True
            raw_terms.extend([(x, line_explicit_small) for x in line_terms])
            # Recover split code-name rows:
            # "02.인사∙" + "조직" -> "02.인사∙조직"
            # "03.일반" + "사무"   -> "03.일반사무"
            if j + 1 < len(lines):
                cur_comp = _compact_line(lines[j])
                nxt_comp = _compact_line(lines[j + 1])
                if (
                    re.match(r"^\d{1,2}[.)]?[가-힣A-Za-z·‧∙ㆍ･]+$", cur_comp)
                    and re.match(r"^[가-힣A-Za-z]{1,8}$", nxt_comp)
                ):
                    merged_line = f"{cur_comp}{nxt_comp}"
                    merged_terms = _extract_anchor_line_terms(merged_line)
                    merged_explicit_small = False
                    merged_pairs = [_clean_category_value(n) for _, n in _CODE_NAME_PAIR_RE.findall(merged_line)]
                    merged_pairs = [p for p in merged_pairs if p and not p.isdigit()]
                    if mode == "down" and len(merged_pairs) == 2 and scanned <= 1:
                        merged_terms = [merged_pairs[0]]
                        merged_explicit_small = True
                    raw_terms.extend([(x, merged_explicit_small) for x in merged_terms])
            scanned += 1
            if scanned >= max_scan:
                break

    alias_index = _build_sclass_exact_alias_index()
    if not alias_index:
        return []

    out: list[str] = []
    seen = set()
    extras_unmapped: list[str] = []
    for term, is_explicit_small in raw_terms:
        cleaned = _clean_category_value(term)
        if not cleaned or cleaned in blocked:
            continue
        key = _sclass_norm_key(cleaned)
        if not key:
            continue
        entry = alias_index.get(key)
        if not entry:
            # Keep only labels that came from explicit "소분류 슬롯" rows.
            if not is_explicit_small:
                continue
            disp = re.sub(r"[·‧･ㆍ•∙⋅]", "", cleaned)
            if len(_compact_line(disp)) >= 2 and disp not in blocked:
                extras_unmapped.append(disp)
            continue
        # Composite labels like "재무·회계" are often 중분류 labels in tables.
        # Do not map those through alias-only paths (prevents 회계 과매칭).
        if not bool(entry.get("official")) and bool(re.search(r"[·･ㆍ•∙⋅/|]", cleaned)):
            continue
        row = entry.get("row") or {}
        name = str(row.get("ncs_sclass_name", "")).strip()
        if not name or name in seen:
            continue
        seen.add(name)
        out.append(name)
        if len(out) >= max_items:
            break
    merged = _dedup_keep_order(out + extras_unmapped)
    return merged[:max_items]


def extract_subcategory_text(jd_text: str) -> str:
    """
    Extract text around '소분류' (preferred) / '세분류' row from JD.
    Works with both normal and partially-broken glyph text.
    """
    src = _repair_mojibake(jd_text)
    lines = _collect_classification_lines(src, max_lines=70)
    if not lines:
        lines = [ln.strip() for ln in src.splitlines() if ln.strip()]
    if not lines:
        return ""

    # 소분류 후보를 상단에 붙여 후속 키워드 추론 품질을 높인다.
    smalls = _extract_small_categories_by_anchor_direction(lines, max_items=8)
    if not smalls:
        smalls = _extract_small_categories_by_code_pairs(lines)
    out = list(lines[:30])
    if smalls:
        out.insert(0, "소분류 후보: " + ", ".join(smalls))
    return "\n".join(out)[:1200]



def _load_ncs_small_categories() -> set[str]:
    """Load NCS small categories from CSV cache.

    Reads from ncs_sclass_codes_with_code_no.csv which contains all official
    NCS small category names (소분류). Uses caching for performance.
    """
    cache_key = "_ncs_small_categories_cache"
    if cache_key in globals():
        return globals()[cache_key]

    categories = set()
    csv_path = os.path.join(os.path.dirname(__file__), "..", "..", "ncs_sclass_codes_with_code_no.csv")

    try:
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row and "NCS_SCLAS_CDNM" in row:
                    sclass = row["NCS_SCLAS_CDNM"].strip()
                    if sclass:
                        categories.add(sclass)
    except Exception:
        # Fallback to hardcoded list if CSV not found
        pass

    # If empty, use fallback list
    if not categories:
        categories = {
            "총무", "자산관리", "사무행정", "회계감사", "회계처리",
            "문서관리", "계약관리", "구매관리", "물품관리", "재물조사",
            "비품관리", "행정지원", "일반사무", "예산관리", "세무회계",
            "인사관리", "기초회계", "경영", "경영기획", "사업계획",
            "간호", "임상병리", "방사선", "물리치료", "작업치료",
        }
    categories.update(
        {
            "교육",
            "정보처리",
            "건축",
            "자동차",
            "마케팅",
            "학사운영",
            "학교교육",
            "경비·경호",
        }
    )

    globals()[cache_key] = categories
    return categories


def lookup_ncs_codes_by_sclass(sclass_names: list[str]) -> list[dict]:
    """소분류 이름 목록으로 CSV에서 NCS 코드 정보를 직접 조회한다 (AI 불필요).

    Returns:
        list of {sclass_name, ncs_code_no, lclas_cd, lclas_nm, mclas_cd, mclas_nm, sclas_cd}
        매칭 안된 항목은 제외.
    """
    cache_key = "_ncs_sclass_rows_cache"
    if cache_key not in globals():
        rows: list[dict] = []
        csv_path = os.path.join(os.path.dirname(__file__), "..", "..", "ncs_sclass_codes_with_code_no.csv")
        for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
            try:
                with open(csv_path, "r", encoding=enc, newline="") as f:
                    reader = csv.DictReader(f)
                    rows = [row for row in reader if row]
                if rows:
                    break
            except Exception:
                rows = []
        globals()[cache_key] = rows

    all_rows: list[dict] = globals()[cache_key]
    # BOM 처리된 첫 번째 컬럼 키 정규화
    code_no_key = next((k for k in (all_rows[0].keys() if all_rows else []) if "NCS_CODE_NO" in k), "NCS_CODE_NO")

    # 소분류명 → 행 인덱스 (정규화 exact 우선, 없으면 정규화 포함 관계)
    results: list[dict] = []
    seen_query: set[str] = set()
    seen_codes: set[tuple[str, str, str, str]] = set()

    # Pre-normalize CSV rows for robust matching.
    prepared_rows: list[tuple[dict, str, str]] = []
    for r in all_rows:
        raw_nm = str(r.get("NCS_SCLAS_CDNM", "")).strip()
        if not raw_nm:
            continue
        prepared_rows.append((r, raw_nm, _sclass_norm_key(raw_nm)))

    synonym_pack = load_sclass_synonym_dictionary()
    synonym_by_code = synonym_pack.get("by_code_no", {})
    synonym_by_name = synonym_pack.get("by_name", {})

    alias_rows: dict[str, tuple[dict, str, str]] = {}
    for r, raw_nm, raw_norm in prepared_rows:
        code_no = str(r.get(code_no_key, "")).strip()
        aliases = _build_sclass_aliases(
            sclass_name=raw_nm,
            code_no=code_no,
            synonym_by_code=synonym_by_code,
            synonym_by_name=synonym_by_name,
        )
        aliases.add(raw_nm)
        for alias in aliases:
            ak = _sclass_norm_key(alias)
            if not ak:
                continue
            prev = alias_rows.get(ak)
            if prev is None or ak == raw_norm:
                alias_rows[ak] = (r, raw_nm, raw_norm)

    for name in sclass_names:
        name = str(name or "").strip()
        q_key = _sclass_norm_key(name)
        if not q_key or q_key in seen_query:
            continue
        seen_query.add(q_key)

        # 1) exact by normalized key (official name + alias dictionary)
        match_item = alias_rows.get(q_key)
        # 2) near-exact contain fallback with overlap guard (avoid 과매칭:
        #    "경영회계사무" -> "회계", "총무인사" -> "총무")
        if match_item is None:
            match_item = next(
                (
                    it for it in prepared_rows
                    if (
                        (q_key in it[2] or it[2] in q_key)
                        and min(len(q_key), len(it[2])) >= 4
                        and (min(len(q_key), len(it[2])) / max(len(q_key), len(it[2]))) >= 0.8
                    )
                ),
                None,
            )
        # 3) raw exact fallback
        if match_item is None:
            match_item = next((it for it in prepared_rows if it[1] == name), None)

        if match_item is None:
            continue

        match = match_item[0]
        canonical_name = match_item[1]
        code_tuple = (
            str(match.get(code_no_key, "")).strip(),
            str(match.get("NCS_LCLAS_CD", "")).strip(),
            str(match.get("NCS_MCLAS_CD", "")).strip(),
            str(match.get("NCS_SCLAS_CD", "")).strip(),
        )
        if code_tuple in seen_codes:
            continue
        seen_codes.add(code_tuple)

        exact_norm = (match_item[2] == q_key)
        results.append({
            # fetch_ncs_units_hrdk_by_verified_sclass 호환 스키마
            "sclass_name": canonical_name,
            "ncs_code_no": match.get(code_no_key, ""),
            "ncs_lclass_code": match.get("NCS_LCLAS_CD", ""),
            "ncs_lclass_name": match.get("NCS_LCLAS_CDNM", ""),
            "ncs_mclass_code": match.get("NCS_MCLAS_CD", ""),
            "ncs_mclass_name": match.get("NCS_MCLAS_CDNM", ""),
            "ncs_sclass_code": match.get("NCS_SCLAS_CD", ""),
            "confidence": 1.0 if exact_norm else 0.85,
            "evidence": "csv-direct-sclass-match",
        })
    return results


def infer_sclass_candidates_from_text_catalog(
    jd_text: str,
    max_items: int = 5,
    hint_terms: list[str] | None = None,
) -> list[dict[str, Any]]:
    """Infer sclass candidates by direct text matching against CSV catalog.

    Strategy:
    - exact/contain matching by normalized sclass name count in jd_text
    - optional hint alias mapping for domain terms (e.g., 일반행정 -> 일반사무)
    - returns schema compatible with fetch_ncs_units_hrdk_by_verified_sclass
    """
    txt = _repair_mojibake(str(jd_text or ""))
    if not txt.strip():
        return []
    norm_txt = _norm_text(txt)
    if not norm_txt:
        return []

    catalog = load_sclass_catalog_from_csv()
    if not catalog:
        return []

    scored: list[tuple[float, dict[str, str], str]] = []
    for row in catalog:
        s_nm = str(row.get("ncs_sclass_name", "")).strip()
        s_norm = _norm_text(s_nm)
        if not s_norm or len(s_norm) < 2:
            continue
        cnt = norm_txt.count(s_norm)
        if cnt <= 0:
            continue
        scored.append((float(cnt), row, f"text-catalog-count:{cnt}"))

    # Hint aliases for terms that may not exist verbatim in catalog labels.
    aliases = {
        "일반행정": "일반사무",
        "학사운영": "평생교육운영",
        "학사": "평생교육운영",
        "경비경호": "경비·경호",
        "경비": "경비·경호",
        "경호": "경비·경호",
    }
    for h in (hint_terms or []):
        hn = _norm_text(str(h or ""))
        if not hn:
            continue
        target = aliases.get(hn, "")
        if not target:
            continue
        tnorm = _norm_text(target)
        for row in catalog:
            s_nm = str(row.get("ncs_sclass_name", "")).strip()
            if _norm_text(s_nm) == tnorm:
                scored.append((0.95, row, f"hint-alias:{h}->{target}"))
                break

    if not scored:
        return []

    scored.sort(key=lambda x: (x[0], len(str(x[1].get("ncs_sclass_name", "")))), reverse=True)
    out: list[dict[str, Any]] = []
    seen = set()
    for score, row, ev in scored:
        key = (
            str(row.get("ncs_code_no", "")).strip(),
            str(row.get("ncs_lclass_code", "")).strip(),
            str(row.get("ncs_mclass_code", "")).strip(),
            str(row.get("ncs_sclass_code", "")).strip(),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(
            {
                "sclass_name": str(row.get("ncs_sclass_name", "")).strip(),
                "ncs_sclass_code": key[3],
                "ncs_lclass_code": key[1],
                "ncs_mclass_code": key[2],
                "ncs_code_no": key[0],
                "confidence": float(min(1.0, max(0.62, score / 3.0 if score > 1 else score))),
                "evidence": ev,
            }
        )
        if len(out) >= max_items:
            break
    return out


_DEFAULT_SCLASS_ALIASES_BY_CODE: dict[str, list[str]] = {
    "020203": ["일반행정", "일반서무", "행정직", "행정지원직", "행정지원", "사무행정", "일반사무", "행정사무"],
    "020201": ["총무", "총무업무", "총무관리", "총무행정"],
    "020302": ["회계", "회계처리", "회계실무", "회계업무"],
    "020301": ["재무", "재무관리", "재무기획"],
    "040202": ["학사운영", "학사", "교육운영", "평생교육운영"],
    "110101": ["경비", "경호", "경비경호", "경비·경호", "시설경비"],
    "230601": ["산업안전", "안전관리", "산업안전관리", "안전보건"],
}

_REVERSE_ANCHOR_TERMS: tuple[str, ...] = ("소분류", "세분류", "ncs소분류", "ncs세분류")
_REVERSE_SECTION_TERMS: tuple[str, ...] = ("분류체계", "ncs분류", "직무분류", "능력단위")


def _build_sclass_aliases(
    sclass_name: str,
    code_no: str,
    synonym_by_code: dict[str, list[str]],
    synonym_by_name: dict[str, list[str]],
) -> set[str]:
    aliases: set[str] = set()
    name = str(sclass_name or "").strip()
    if not name:
        return aliases

    aliases.add(name)
    aliases.add(name.replace("·", ""))
    aliases.add(name.replace("·", " ").strip())
    aliases.add(name.replace("/", " ").strip())
    aliases.update(_DEFAULT_SCLASS_ALIASES_BY_CODE.get(str(code_no or "").strip(), []))
    aliases.update(synonym_by_code.get(str(code_no or "").strip(), []))
    aliases.update(synonym_by_name.get(_norm_text(name), []))
    return {a.strip() for a in aliases if str(a or "").strip()}


def _build_reverse_line_context(
    text: str,
    near_anchor_window: int = 8,
    near_section_window: int = 2,
) -> tuple[list[str], set[int], set[int]]:
    raw_lines = [str(ln).strip() for ln in str(text or "").splitlines() if str(ln).strip()]
    norm_lines = [_norm_text(ln) for ln in raw_lines]

    anchor_indices = [i for i, ln in enumerate(norm_lines) if any(term in ln for term in _REVERSE_ANCHOR_TERMS)]
    section_indices = [i for i, ln in enumerate(norm_lines) if any(term in ln for term in _REVERSE_SECTION_TERMS)]

    anchor_near_set: set[int] = set()
    section_near_set: set[int] = set()

    for idx in anchor_indices:
        # Bias anchor context to the lines below the "소분류" row.
        for off in range(-1, near_anchor_window + 1):
            pos = idx + off
            if 0 <= pos < len(norm_lines):
                anchor_near_set.add(pos)

    for idx in section_indices:
        for off in range(-near_section_window, near_section_window + 1):
            pos = idx + off
            if 0 <= pos < len(norm_lines):
                section_near_set.add(pos)

    return norm_lines, anchor_near_set, section_near_set


def infer_sclass_candidates_reverse_dictionary(
    jd_text: str,
    hint_terms: list[str] | None = None,
    doc_name: str = "",
    max_items: int = 8,
) -> list[dict[str, Any]]:
    """Dictionary-first reverse recognition for sclass.

    Instead of extracting arbitrary words first, this scans a predefined sclass
    dictionary (official names + aliases) against document text and ranks
    candidates by weighted hit score.
    """
    txt = _repair_mojibake(str(jd_text or ""))
    if not txt.strip():
        return []
    norm_txt = _norm_text(txt)
    norm_doc = _norm_text(doc_name or "")
    if not norm_txt:
        return []

    catalog = load_sclass_catalog_from_csv()
    if not catalog:
        return []

    synonym_pack = load_sclass_synonym_dictionary()
    synonym_by_code = synonym_pack.get("by_code_no", {})
    synonym_by_name = synonym_pack.get("by_name", {})

    norm_hints = {_norm_text(t) for t in (hint_terms or []) if _norm_text(t)}
    norm_lines, anchor_near_set, section_near_set = _build_reverse_line_context(txt)
    scored: list[tuple[float, dict[str, str], str]] = []
    for row in catalog:
        code_no = str(row.get("ncs_code_no", "")).strip()
        name = str(row.get("ncs_sclass_name", "")).strip()
        if not (code_no and name):
            continue

        aliases = _build_sclass_aliases(
            sclass_name=name,
            code_no=code_no,
            synonym_by_code=synonym_by_code,
            synonym_by_name=synonym_by_name,
        )

        hit_score = 0.0
        hit_count = 0
        anchor_hits = 0
        section_hits = 0
        official_norm = _norm_text(name)
        for a in aliases:
            na = _norm_text(a)
            if len(na) < 2:
                continue
            if na not in norm_txt:
                continue

            for i, ln in enumerate(norm_lines):
                cnt = ln.count(na)
                if cnt <= 0:
                    continue
                hit_count += cnt
                # exact official name gets higher weight than alias.
                w = 1.2 if na == official_norm else 0.9
                if i in anchor_near_set:
                    w += 0.9
                    anchor_hits += cnt
                elif i in section_near_set:
                    w += 0.5
                    section_hits += cnt
                hit_score += float(cnt) * w

        if hit_count <= 0:
            continue

        # hints and file-name matches help for ambiguous docs.
        term_bonus = 0.0
        for a in aliases:
            na = _norm_text(a)
            if na in norm_hints:
                term_bonus += 0.8
            if na and na in norm_doc:
                term_bonus += 0.6
        total = hit_score + term_bonus
        scored.append(
            (
                total,
                row,
                (
                    "reverse-dict:"
                    f"hit={hit_count},anchor={anchor_hits},section={section_hits},bonus={round(term_bonus,2)}"
                ),
            )
        )

    if not scored:
        return []

    scored.sort(key=lambda x: (x[0], len(str(x[1].get("ncs_sclass_name", "")))), reverse=True)
    out: list[dict[str, Any]] = []
    seen = set()
    for score, row, evidence in scored:
        key = (
            str(row.get("ncs_code_no", "")).strip(),
            str(row.get("ncs_lclass_code", "")).strip(),
            str(row.get("ncs_mclass_code", "")).strip(),
            str(row.get("ncs_sclass_code", "")).strip(),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(
            {
                "sclass_name": str(row.get("ncs_sclass_name", "")).strip(),
                "ncs_sclass_code": key[3],
                "ncs_lclass_code": key[1],
                "ncs_mclass_code": key[2],
                "ncs_code_no": key[0],
                "confidence": float(min(1.0, max(0.62, score / 4.5))),
                "evidence": evidence,
            }
        )
        if len(out) >= max_items:
            break
    return out


def extract_small_categories_from_jd(jd_text: str) -> list[str]:
    """Extract NCS small-category labels from JD text (robust for OCR/noisy text).

    Improvements:
    - Increased line processing from 5 to 20 lines
    - Better tokenization with variable length support
    - Comma-separated category handling
    - NCS dataset validation
    - Improved stop word filtering
    """
    raw = jd_text or ""
    repaired = _repair_mojibake(raw)
    src = raw if re.search(r"[가-힣]", raw) else repaired

    # 1-a) 헤더 위치 기반 추출 (세로형/가로형 레이아웃 직접 처리)
    focus_lines = _collect_classification_lines(src, max_lines=90)
    known_categories = _load_ncs_small_categories()

    # Kordoc's HTML-table output retains the exact 소분류 row even when the
    # line-oriented text has tags and colspan boundaries mixed into labels.
    html_structural = _extract_small_categories_from_html_table(src, known_categories)
    if html_structural:
        return html_structural[:15]

    for idx, line in enumerate(focus_lines):
        if "소분류" not in line:
            continue
        plain_terms: list[str] = []
        for candidate_line in focus_lines[idx + 1 : idx + 20]:
            if any(header in candidate_line for header in ("대분류", "중분류", "세분류", "직무수행")):
                break
            term = _clean_category_value(candidate_line)
            if term in known_categories:
                plain_terms.append(term)
        if len(plain_terms) >= 2:
            return _dedup_keep_order(plain_terms)[:15]

    # 1-a0) 세로/가로 레이아웃이 크게 깨진 표 전용 복원
    vertical_blocks = _extract_small_categories_by_vertical_blocks(focus_lines, max_items=15)
    positional = _extract_sclass_by_header_position(focus_lines)
    # 1-b) 표 구조(가로/세로/혼합)에서 코드-명칭 패턴으로 소분류 열 복원
    structural = _extract_small_categories_by_code_pairs(focus_lines)
    # 1-c) 소분류 앵커 주변 방향성 스캔
    anchored = _extract_small_categories_by_anchor_direction(focus_lines, max_items=15)

    # 1-d) 후보들 중 실제 소분류 매핑이 가장 좋은 집합을 선택
    # vertical_blocks는 표 붕괴가 강한 경우에만 사용(그 외에는 과추출 위험).
    candidates_pool: list[tuple[str, list[str], int]] = []
    if len(_dedup_keep_order(vertical_blocks)) >= 4:
        candidates_pool.append(("vertical_blocks", vertical_blocks, 4))
    # 일반 케이스는 anchored를 structural보다 우선 순위로 둔다.
    candidates_pool.extend(
        [
            ("anchored", anchored, 3),
            ("structural", structural, 2),
            ("positional", positional, 1),
        ]
    )
    best_source = ""
    best_terms: list[str] = []
    best_key = (-1, -1, -1)
    mclass_keys = set(_build_mclass_to_sclass_keys_index().keys())
    for source, terms, priority in candidates_pool:
        uniq_terms = _dedup_keep_order(terms)[:15]
        if not uniq_terms:
            continue
        mapped = lookup_ncs_codes_by_sclass(uniq_terms)
        mapped_names = {str(x.get("sclass_name", "")).strip() for x in mapped if str(x.get("sclass_name", "")).strip()}
        mclass_noise = sum(1 for n in mapped_names if _sclass_norm_key(n) in mclass_keys)
        effective_mapped = max(0, len(mapped_names) - mclass_noise)
        key = (effective_mapped, len(mapped_names), priority)
        if key > best_key:
            best_key = key
            best_terms = uniq_terms
            best_source = source
    if best_terms:
        if best_source == "vertical_blocks":
            # Keep parsed table labels as-is for broken layouts (user expectation).
            return _dedup_keep_order(best_terms)[:15]
        # If mapped names cover most terms, return mapped canonical names for consistency.
        mapped = lookup_ncs_codes_by_sclass(best_terms)
        mapped_names = _dedup_keep_order([str(x.get("sclass_name", "")).strip() for x in mapped if str(x.get("sclass_name", "")).strip()])
        if mapped_names:
            mapped_keys = {_sclass_norm_key(x) for x in mapped_names if _sclass_norm_key(x)}
            extras: list[str] = []
            for t in best_terms:
                tt = _clean_category_value(t)
                if not tt:
                    continue
                preserve_explicit = _sclass_norm_key(tt) in {"학사운영"}
                if _sclass_norm_key(tt) in mapped_keys and not preserve_explicit:
                    continue
                # Keep explicit unmapped labels read from the 소분류 region.
                disp = re.sub(r"[·‧･ㆍ•∙⋅]", "", tt)
                if disp:
                    extras.append(disp)
            merged = _dedup_keep_order(mapped_names + extras)
            for explicit in _dedup_keep_order(structural + vertical_blocks):
                explicit_clean = _clean_category_value(explicit)
                if _sclass_norm_key(explicit_clean) in {"학사운영"} and explicit_clean not in merged:
                    merged.append(explicit_clean)
            return merged[:15]
        return best_terms[:15]

    # 2) 구조 복원이 실패하면 단어 기반 백업 추출
    lines = [ln.strip() for ln in src.splitlines() if ln.strip()]
    combined = "\n".join(focus_lines) if focus_lines else src[:2000]
    tokens = re.findall("[\uAC00-\uD7A3]{1,20}", combined)

    stop = {
        "소분류",
        "세분류",
        "대분류",
        "중분류",
        "분류체계",
        "직무수행",
        "직무수행내용",
        "능력단위",
        "필요지식",
        "필요기술",
        "채용분야",
        "직업기초능력",
        "관련자격",
        "관련전공과목",
        "기간제계약직",
        "휴직대체",
        "참고사이트",
        "비고",
    }
    out: list[str] = []
    seen = set()
    for tok in tokens:
        t = tok.strip()
        if not t or t in stop:
            continue
        if t in known_categories:
            if t not in seen:
                seen.add(t)
                out.append(t)
            continue
        best_match = None
        best_match_len = 0
        for known in known_categories:
            if t in known and len(t) >= 2 and len(known) > best_match_len:
                best_match = known
                best_match_len = len(known)
            elif known in t and len(known) >= 2 and len(known) > best_match_len:
                best_match = known
                best_match_len = len(known)
        if best_match and best_match not in seen:
            seen.add(best_match)
            out.append(best_match)
        if len(out) >= 15:
            break

    # 3) 마지막 안전장치: 소분류 라벨 뒤 텍스트 직접 파싱
    if not out:
        marker_patterns = [r"소\s*분\s*류", "소분류"]
        for line in focus_lines or lines:
            if any(re.search(p, line, re.IGNORECASE) for p in marker_patterns):
                parts = re.split(r"[:：]\s*", line, maxsplit=1)
                if len(parts) > 1:
                    vals = [_clean_category_value(x) for x in re.split(r"[,/|]", parts[1])]
                    out.extend([v for v in vals if v and v not in stop])
                break

    return _dedup_keep_order(out)[:15]


def extract_detail_categories_from_jd(jd_text: str) -> list[str]:
    """
    Extract 세분류 labels (e.g., 총무/자산관리/사무행정/회계감사) from JD text.
    """
    src = _repair_mojibake(jd_text)
    lines = [ln.strip() for ln in src.splitlines() if ln.strip()]
    candidate = src
    for i, ln in enumerate(lines):
        if "세분류" in ln or "?몃텇瑜?" in ln:
            candidate = " ".join(lines[i : min(i + 4, len(lines))])
            break

    known = [
        "총무",
        "자산관리",
        "사무행정",
        "회계감사",
        "회계처리",
        "문서관리",
        "계약관리",
        "구매관리",
        "물품관리",
        "재물조사",
        "비품관리",
        "행정지원",
    ]
    alias_hits = {
        "珥앸Т": "총무",
        "먯궛愿由?": "자산관리",
        "?щТ?됱젙": "사무행정",
        "?뚭퀎쨌媛먯궗": "회계감사",
        "?뚭퀎": "회계처리",
    }

    out: list[str] = []
    combined = f"{candidate}\n{src}"
    for k in known:
        if k in combined and k not in out:
            out.append(k)
    for broken, fixed in alias_hits.items():
        if broken in combined and fixed not in out:
            out.append(fixed)

    priority = {
        "총무": 0,
        "자산관리": 1,
        "사무행정": 2,
        "회계감사": 3,
        "회계처리": 4,
    }
    out = sorted(set(out), key=lambda x: priority.get(x, 99))
    return out[:10]


def infer_keywords_from_subcategory_ai(subcategory_text: str, jd_text: str) -> list[str]:
    """
    Infer NCS keywords from 소분류/JD context (local-fast path).

    Notes:
    - No job-family-specific hardcoded priority is applied by default.
    - Priority can be injected externally via env `NCS_KEYWORD_PRIORITY`.
    """
    sub = _repair_mojibake(subcategory_text or "")
    jd = _repair_mojibake(jd_text or "")
    combined = f"{sub}\n{jd}"
    out: list[str] = []

    def _push(term: str) -> None:
        t = str(term or "").strip()
        t = re.sub(r"\s+", " ", t).strip(" ,:;|/-")
        if t in {"대분류", "중분류", "소분류", "세분류", "분류체계"}:
            return
        if len(t) < 2:
            return
        if t not in out:
            out.append(t)

    # 1) Parse explicit subcategory lines first.
    for ln in [x.strip() for x in sub.splitlines() if x.strip()]:
        compact = re.sub(r"\s+", "", ln)
        if "소분류후보" in compact or "소분류" in compact or "세분류" in compact:
            right = re.split(r"[:：]", ln, maxsplit=1)
            rhs = right[1] if len(right) > 1 else ln
            for seg in re.split(r"[,/|]", rhs):
                for tok in re.findall(r"[가-힣A-Za-z0-9()+\-]{2,30}", seg):
                    _push(tok)
            if len(out) >= 20:
                break

    # 2) Recover readable terms from mojibake aliases without fixed rank.
    for broken, fixed in MOJIBAKE_ALIAS.items():
        if broken in combined or fixed in combined:
            _push(fixed)

    # 3) Add frequent Korean tokens from text.
    token_stop = {
        "분류체계",
        "대분류",
        "중분류",
        "소분류",
        "세분류",
        "능력단위",
        "직무",
        "업무",
        "수행",
        "관련",
        "채용",
        "기준",
        "필요지식",
        "필요기술",
    }
    freq_tokens = [t for t in re.findall(r"[\uac00-\ud7a3]{2,16}", combined) if t not in token_stop]
    for tok, _ in Counter(freq_tokens).most_common(60):
        _push(tok)
        if len(out) >= 20:
            break

    # 4) Prefer exact CSV small-category matches, but avoid fuzzy broad matching.
    catalog = load_sclass_catalog_from_csv()
    if catalog and out:
        norm_set = {_norm_text(x) for x in out if _norm_text(x)}
        exact_csv_hits: list[str] = []
        for row in catalog:
            name = str(row.get("ncs_sclass_name", "")).strip()
            if not name:
                continue
            if _norm_text(name) in norm_set and name not in exact_csv_hits:
                exact_csv_hits.append(name)
        out = _dedup_keep_order(exact_csv_hits + out)

    # 5) Optional external priority injection (comma-separated).
    env_priority_raw = os.getenv("NCS_KEYWORD_PRIORITY", "").strip()
    if env_priority_raw:
        priority = [x.strip() for x in env_priority_raw.split(",") if x.strip()]
        prioritized = [x for x in priority if x in out]
        tail = [x for x in out if x not in prioritized]
        out = prioritized + tail

    return out[:12]


def build_local_question_pack(jd_text: str, strengths: str, ncs_matches: list[dict[str, Any]]) -> dict[str, Any]:
    interview_questions = _generate_questions_with_openai_from_ncs(
        jd_text=jd_text,
        strengths=strengths,
        ncs_matches=ncs_matches,
        target_count=20,
        mode="local_pack",
    )
    by_comp = _build_interview_by_competency_from_questions(interview_questions)
    return {"interview_by_competency": by_comp, "interview_questions": interview_questions}


def _build_ksa_driven_question_pack(
    ncs_matches: list[dict[str, Any]],
    ncs_ksa: list[dict[str, Any]] | None = None,
    strengths: str = "",
) -> list[dict[str, Any]]:
    interview_questions = _generate_questions_with_openai_from_ncs(
        jd_text="",
        strengths=strengths,
        ncs_matches=ncs_matches,
        ncs_ksa=ncs_ksa,
        target_count=min(max(len(ncs_matches or []) * 4, 8), 32),
        mode="ksa_driven",
    )
    return _build_interview_by_competency_from_questions(interview_questions)


def _official_question_grounding(
    row: dict[str, Any],
    ncs_ksa: list[dict[str, Any]] | None,
    *,
    default_code: str = "",
) -> tuple[str, list[str], str]:
    question_code = str(
        row.get("ncsClCd") or row.get("ncs_code") or default_code
    ).strip()
    official_rows = [
        item
        for item in (ncs_ksa or [])
        if isinstance(item, dict)
        and str(item.get("ncsClCd") or "").strip() == question_code
        and str(item.get("factorName") or "").strip()
    ]
    official_by_key = {
        re.sub(r"\s+", "", str(item.get("factorName") or "")).lower(): str(
            item.get("factorName") or ""
        ).strip()
        for item in official_rows
    }
    official_by_id = {
        stable_ksa_evidence_id(item): str(item.get("factorName") or "").strip()
        for item in official_rows
    }
    raw_refs = (
        list(row.get("ksa_refs") or [])
        if isinstance(row.get("ksa_refs"), list)
        else []
    )
    verified_refs = [
        official_by_key[key]
        for value in raw_refs
        if (key := re.sub(r"\s+", "", str(value or "")).lower()) in official_by_key
    ]
    raw_focus = _primary_question_focus(row)
    focus_key = re.sub(r"\s+", "", raw_focus).lower()
    verified_focus = official_by_key.get(focus_key) or (
        verified_refs[0] if verified_refs else ""
    )
    # ``_server_selected_evidence_id`` is added only after the provider's
    # declared id has been resolved against an official MCP row.  It lets the
    # public result recover when a capable model returns the right evidence id
    # but omits the redundant ``ksa_refs`` array, without trusting arbitrary
    # provider-supplied factor labels.
    server_selected_id = str(row.get("_server_selected_evidence_id") or "").strip()
    server_selected_focus = official_by_id.get(server_selected_id, "")
    if not verified_focus and server_selected_focus:
        verified_focus = server_selected_focus
        verified_refs = [server_selected_focus]
    if verified_focus:
        verified_refs = [
            verified_focus,
            *[ref for ref in verified_refs if ref != verified_focus],
        ]
    source = "official_ksa" if verified_focus and verified_refs else (
        "synthetic_template"
        if str(row.get("question_focus_source") or "").strip() == "synthetic_template"
        else "unverified_model_output"
    )
    if source == "synthetic_template" and not verified_focus:
        verified_focus = raw_focus
    return verified_focus, list(dict.fromkeys(verified_refs)), source


def _server_official_ksa_evidence(
    rows: list[dict[str, Any]] | None,
) -> list[dict[str, Any]]:
    """Build the server-owned evidence registry used by public quality gates.

    Question-local evidence metadata originates in model output and must never
    attest itself.  These rows come only from the NCS_MCP lookup performed by
    the surrounding service, and their stable identifiers are recomputed here
    before the result crosses the public endpoint boundary.
    """

    registry: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for raw_row in rows or []:
        if not isinstance(raw_row, dict):
            continue
        row = dict(raw_row)
        ncs_code = str(row.get("ncsClCd") or row.get("unit_code") or "").strip()
        factor = str(row.get("factorName") or row.get("factor_name") or "").strip()
        if not ncs_code or not factor:
            continue
        evidence_id = stable_ksa_evidence_id(row)
        if evidence_id in seen_ids:
            continue
        row["evidence_id"] = evidence_id
        registry.append(row)
        seen_ids.add(evidence_id)
    return registry


def generate_personalized_interview_questions(
    ncs_code: str,
    competency_name: str = "",
    job_posting: str = "",
    user_profile: str = "",
    target_count: int = 12,
    api_key_override: str = "",
    generation_model: str = "",
    generation_provider: str = "openai_api",
) -> dict[str, Any]:
    comp_name = competency_name or f"NCS-{ncs_code}"
    ncs_matches = [{"ncsClCd": ncs_code, "compeUnitName": comp_name}]
    ncs_ksa = _safe_fetch_ncs_ksa_by_units(
        ncs_matches=ncs_matches,
        max_units=1,
        max_factors_per_unit=8,
    )
    generated = _generate_questions_with_openai_from_ncs(
        jd_text=job_posting,
        strengths=user_profile,
        ncs_matches=ncs_matches,
        ncs_ksa=ncs_ksa,
        target_count=min(max(target_count, 1), 20),
        mode="personalized",
        extra_context=f"user_profile={user_profile[:3000]}",
        api_key_override=api_key_override,
        generation_model=generation_model,
        generation_provider=generation_provider,
    )
    questions: list[dict[str, Any]] = []
    for q in generated:
        grounded_row = dict(q)
        grounded_row["ncsClCd"] = ncs_code
        verified_focus, verified_refs, focus_source = _official_question_grounding(
            grounded_row,
            ncs_ksa,
            default_code=ncs_code,
        )
        questions.append({
            "question": str(q.get("question", "")).strip(),
            "question_type": str(q.get("type", "면접질문")).strip() or "면접질문",
            "type": str(q.get("type", "면접질문")).strip() or "면접질문",
            "competency": comp_name,
            "ncs_code": ncs_code,
            "ncsClCd": ncs_code,
            "question_focus": verified_focus,
            "question_focus_source": focus_source,
            "question_focus_surface": str(q.get("question_focus_surface") or "").strip(),
            "question_task_frame": dict(q.get("question_task_frame") or {}),
            "question_evidence_id": str(q.get("question_evidence_id") or "").strip(),
            "question_evidence_required": bool(q.get("question_evidence_id")),
            "follow_ups": list(q.get("follow_ups", []) or []) if isinstance(q.get("follow_ups"), list) else [],
            "evaluation_points": list(q.get("evaluation_points", []) or []) if isinstance(q.get("evaluation_points"), list) else [],
            "eval_points": list(q.get("evaluation_points", []) or []) if isinstance(q.get("evaluation_points"), list) else [],
            "ksa_refs": list(dict.fromkeys(verified_refs)),
            "question_source": str(q.get("question_source") or generation_provider).strip(),
            "model_question_preserved": True,
            "candidate_selection_policy": str(q.get("candidate_selection_policy") or "").strip(),
            "candidate_pool_count": int(q.get("candidate_pool_count") or 0),
            "candidate_quality_score": float(q.get("candidate_quality_score") or 0.0),
            "candidate_selection_score": float(q.get("candidate_selection_score") or 0.0),
            "candidate_diversity_axes": dict(q.get("candidate_diversity_axes") or {}),
        })
    _refresh_ncs_code_main_question_repeat_metadata(questions)
    all_questions_grounded = bool(questions) and all(
        row.get("question_focus_source") == "official_ksa"
        and bool(row.get("ksa_refs"))
        for row in questions
    )
    return {
        "ncs_code": ncs_code,
        "competency_name": comp_name,
        "generation_mode": "ai_personalized_ncs",
        "company_from_posting": "",
        "requirements_from_posting": "",
        "skills_from_profile": "",
        "questions": questions,
        "question_count": len(generated),
        "generation_provider": generation_provider,
        "provider_generation_model": str(
            (generated[0] if generated else {}).get("provider_generation_model")
            or generation_model
        ).strip(),
        "provider_candidate_variant_count": int(
            (generated[0] if generated else {}).get("provider_candidate_variant_count") or 0
        ),
        "provider_candidate_variant_received_count": int(
            (generated[0] if generated else {}).get(
                "provider_candidate_variant_received_count"
            )
            or 0
        ),
        "ncs_ksa_available": bool(ncs_ksa) and all_questions_grounded,
        "official_ksa_evidence": _server_official_ksa_evidence(ncs_ksa),
        "warning": (
            ""
            if ncs_ksa and all_questions_grounded
            else "official_ncs_ksa_unavailable_or_question_grounding_failed"
        ),
        "note": "NCS 컨텍스트 기반 생성형 AI 자율 생성 결과입니다.",
    }


def generate_diverse_interview_questions(
    ncs_code: str,
    competency_name: str = "",
    job_posting: str = "",
    target_count: int = 6,
    extra_context: str = "",
    api_key_override: str = "",
    generation_model: str = "",
    generation_provider: str = "openai_api",
) -> dict[str, Any]:
    comp_name = competency_name or f"NCS-{ncs_code}"
    ncs_matches = [{"ncsClCd": ncs_code, "compeUnitName": comp_name}]
    ncs_ksa = _safe_fetch_ncs_ksa_by_units(ncs_matches=ncs_matches, max_units=1, max_factors_per_unit=6)
    generated = _generate_questions_with_openai_from_ncs(
        jd_text=job_posting,
        ncs_matches=ncs_matches,
        ncs_ksa=ncs_ksa,
        target_count=min(max(target_count, 1), 6),
        mode="diverse",
        extra_context=extra_context,
        api_key_override=api_key_override,
        generation_model=generation_model,
        generation_provider=generation_provider,
    )
    questions_list: list[dict[str, Any]] = []
    for i, q in enumerate(generated, 1):
        grounded_row = dict(q)
        grounded_row["ncsClCd"] = ncs_code
        verified_focus, verified_refs, focus_source = _official_question_grounding(
            grounded_row,
            ncs_ksa,
            default_code=ncs_code,
        )
        raw_fu = q.get("follow_ups")
        if isinstance(raw_fu, list):
            follow_ups = [str(x).strip() for x in raw_fu if str(x).strip()]
        else:
            one = str(q.get("follow_up", "")).strip()
            follow_ups = [one] if one else []
        questions_list.append(
            {
                "number": i,
                "type": str(q.get("type", "면접질문")).strip() or "면접질문",
                "competency": str(q.get("competency", comp_name)).strip() or comp_name,
                "ncs_code": ncs_code,
                "ncsClCd": ncs_code,
                "question": str(q.get("question", "")).strip(),
                "question_focus": verified_focus,
                "question_focus_source": focus_source,
                "question_focus_surface": str(q.get("question_focus_surface") or "").strip(),
                "question_task_frame": dict(q.get("question_task_frame") or {}),
                "question_evidence_id": str(q.get("question_evidence_id") or "").strip(),
                "question_evidence_required": bool(q.get("question_evidence_id")),
                "follow_ups": follow_ups,
                "follow_up": (follow_ups[0] if follow_ups else ""),
                "eval_points": list(q.get("evaluation_points", []) or []),
                "ksa_refs": verified_refs,
                "question_source": str(q.get("question_source") or generation_provider).strip(),
                "model_question_preserved": True,
                "candidate_selection_policy": str(q.get("candidate_selection_policy") or "").strip(),
                "candidate_pool_count": int(q.get("candidate_pool_count") or 0),
                "candidate_quality_score": float(q.get("candidate_quality_score") or 0.0),
                "candidate_selection_score": float(q.get("candidate_selection_score") or 0.0),
                "candidate_diversity_axes": dict(q.get("candidate_diversity_axes") or {}),
            }
        )
    all_questions_grounded = bool(questions_list) and all(
        row.get("question_focus_source") == "official_ksa"
        and bool(row.get("ksa_refs"))
        for row in questions_list
    )
    return {
        "ncs_code": ncs_code,
        "competency_name": comp_name,
        "generation_mode": "ai_autonomous_ncs",
        "questions": questions_list,
        "question_count": len(questions_list),
        "generation_provider": generation_provider,
        "provider_generation_model": str(
            (generated[0] if generated else {}).get("provider_generation_model")
            or generation_model
        ).strip(),
        "provider_candidate_variant_count": int(
            (generated[0] if generated else {}).get("provider_candidate_variant_count") or 0
        ),
        "provider_candidate_variant_received_count": int(
            (generated[0] if generated else {}).get(
                "provider_candidate_variant_received_count"
            )
            or 0
        ),
        "ncs_ksa_available": bool(ncs_ksa) and all_questions_grounded,
        "official_ksa_evidence": _server_official_ksa_evidence(ncs_ksa),
        "warning": (
            ""
            if ncs_ksa and all_questions_grounded
            else "official_ncs_ksa_unavailable_or_question_grounding_failed"
        ),
        "note": "NCS 컨텍스트 기반 생성형 AI 자율 생성 결과입니다.",
    }


_NCS_CODE_TEMPLATE_FALLBACK_METHODS = (
    "경험면접",
    "상황면접",
    "직무지식면접",
    "인바스켓면접",
    "발표면접",
    "토론면접",
    "창의적 문제해결력면접",
)


def _safe_fetch_ncs_ksa_by_units(**kwargs: Any) -> list[dict[str, Any]]:
    try:
        return fetch_ncs_ksa_by_units(**kwargs)
    except NcsMcpError:
        return []


def _primary_question_focus(row: dict[str, Any]) -> str:
    refs = row.get("ksa_refs")
    first_ref = ""
    if isinstance(refs, list) and refs:
        first_ref = str(refs[0] or "").strip()
    return str(row.get("question_focus") or first_ref).strip()


def _compact_question_intent_text(value: Any) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return ""
    raw = re.sub(r"[^0-9a-z가-힣]+", " ", raw)
    return re.sub(r"\s+", "", raw)


def _ncs_code_question_intent_key(question: Any) -> str:
    return classify_question_intent(question, unknown="other")


def _ncs_code_repeat_signature(row: dict[str, Any]) -> str:
    question = str((row or {}).get("question") or "").strip()
    intent = _ncs_code_question_intent_key(question)
    if not intent:
        return ""
    method = str((row or {}).get("question_type") or (row or {}).get("type") or "").strip()
    method_key = _compact_question_intent_text(method)
    focus = _primary_question_focus(row)
    if intent in GENERAL_QUESTION_INTENTS and not (
        intent in FOCUS_SCOPED_GENERAL_QUESTION_INTENTS and focus
    ):
        return f"{intent}|general"
    if focus:
        return f"{intent}|{method_key}|focus:{_compact_question_intent_text(focus)[:80]}"
    subject = str((row or {}).get("ncsClCd") or (row or {}).get("competency") or "").strip()
    return f"{intent}|{method_key}|{_compact_question_intent_text(subject)[:80]}"


def _refresh_ncs_code_main_question_repeat_metadata(rows: list[dict[str, Any]]) -> None:
    seen_by_signature: dict[str, list[str]] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        question = str(row.get("question") or "").strip()
        signature = _ncs_code_repeat_signature(row)
        previous_questions = seen_by_signature.get(signature, []) if signature else []
        repeat_duplicate = False
        if signature.endswith("|general") and previous_questions:
            repeat_duplicate = True
        elif signature:
            repeat_duplicate = any(
                normalize_question_dedup_key(question) == normalize_question_dedup_key(previous)
                or is_similar_question_text(question, previous)
                for previous in previous_questions
            )
        row["question_intent"] = _ncs_code_question_intent_key(question)
        row["question_repeat_signature"] = signature
        row["question_repeat_duplicate"] = bool(repeat_duplicate)
        if signature and question:
            seen_by_signature.setdefault(signature, []).append(question)


def _job_context_excerpt(
    value: Any,
    *,
    anchors: list[str] | None = None,
    limit: int = 220,
) -> str:
    """Extract a short operational excerpt from the uploaded job materials.

    Deterministic recovery must remain useful when the model provider times
    out.  The excerpt is deliberately source-derived and bounded; it is not a
    domain profile or an invented occupation-specific scenario.
    """

    compact = re.sub(r"\s+", " ", str(value or "").strip())
    compact = re.sub(
        r"\[(?:공고문|직무기술서|담당업무|지원자격|우대사항|면접평가항목|발표자료)[^\]]*\]\s*",
        "",
        compact,
    )
    if not compact:
        return ""
    segments = [
        re.sub(r"^[-•*○◦▪●\s]+", "", part).strip(" -·•○◦▪●")
        for part in re.split(
            r"(?<=[.!?。])\s+|(?=[○◦•▪●]\s)|(?=\d+[.)]\s)|\s+[|｜]\s+",
            compact,
        )
        if len(part.strip()) >= 12
    ]
    if not segments:
        segments = [compact]
    normalized_anchors = [
        re.sub(r"\s+", "", str(anchor or "")).casefold()
        for anchor in (anchors or [])
        if len(re.sub(r"\s+", "", str(anchor or ""))) >= 2
    ]
    operational_markers = (
        "담당", "수행", "운영", "관리", "점검", "유지", "처리", "작성", "분석",
        "설계", "개선", "보고", "민원", "현장", "설비", "자료", "기록",
    )
    scored: list[tuple[int, int, str]] = []
    for position, segment in enumerate(segments):
        normalized = re.sub(r"\s+", "", segment).casefold()
        anchor_hits = sum(1 for anchor in normalized_anchors if anchor in normalized)
        marker_hits = sum(1 for marker in operational_markers if marker in segment)
        # Source order is intentional: callers place 담당업무 before JD and
        # 공고문, while anchors keep the selected NCS work element preferred.
        score = (anchor_hits * 10) + min(marker_hits, 5)
        if score:
            scored.append((score, -position, segment))
    selected = [item[2] for item in sorted(scored, reverse=True)[:2]] if scored else segments[:1]
    excerpt = " ".join(dict.fromkeys(selected)).strip()
    if len(excerpt) <= max(40, int(limit or 220)):
        return excerpt
    clipped = excerpt[: max(40, int(limit or 220))].rsplit(" ", 1)[0].rstrip(" ,·")
    return f"{clipped}…"


def _build_ncs_code_template_fallback_question(
    *,
    unit: dict[str, Any] | None,
    comp_name: str,
    ncs_code: str,
    ksa_terms: list[str],
    evidence_terms: list[str] | None = None,
    evidence_rows: list[dict[str, Any]] | None = None,
    index: int,
    method_override: str | None = None,
    case_slot_id: str | None = None,
    case_slot_signature: str | None = None,
    presentation_material_text: str = "",
    job_context_text: str = "",
) -> dict[str, Any]:
    method = str(method_override or "").strip() or _NCS_CODE_TEMPLATE_FALLBACK_METHODS[
        index % len(_NCS_CODE_TEMPLATE_FALLBACK_METHODS)
    ]
    label = str((unit or {}).get("compeUnitName") or comp_name or "해당 직무").strip()
    code = str((unit or {}).get("ncsClCd") or ncs_code).strip()
    detail = str(
        (unit or {}).get("ncsSubdCdnm")
        or (unit or {}).get("ncsSclasCdnm")
        or (unit or {}).get("matchedDetailName")
        or ""
    ).strip()
    context_label = label
    if detail and re.sub(r"\s+", "", detail).lower() not in re.sub(r"\s+", "", label).lower():
        context_label = f"{detail} {label}".strip()
    pool = [str(x).strip() for x in ksa_terms if str(x).strip()]
    if not pool:
        pool = ["업무 우선순위 설정", "이해관계자 협업", "성과 점검 및 개선"]
    evidence_pool = [str(x).strip() for x in (evidence_terms or []) if str(x).strip()]
    k1 = evidence_pool[index % len(evidence_pool)] if evidence_pool else pool[index % len(pool)]
    k1_index = pool.index(k1) if k1 in pool else index % len(pool)
    k2 = pool[(k1_index + 1) % len(pool)] if len(pool) > 1 else "성과 점검 및 개선"
    official_terms = {
        re.sub(r"\s+", "", str(term or "")).lower()
        for term in (evidence_terms if evidence_terms is not None else pool)
        if str(term or "").strip()
    }
    evidence_refs = [
        term
        for term in (k1, k2)
        if re.sub(r"\s+", "", term).lower() in official_terms
    ]

    evidence_row = next(
        (
            row
            for row in (evidence_rows or [])
            if isinstance(row, dict)
            and re.sub(r"\s+", "", str(row.get("factorName") or "")).lower()
            == re.sub(r"\s+", "", k1).lower()
        ),
        None,
    )
    task_frame = build_question_task_frame(
        evidence_row=evidence_row,
        factor_name=k1,
        ksa_type=(evidence_row or {}).get("ksaTypeName") or (evidence_row or {}).get("factorType") or "",
        element_name=(evidence_row or {}).get("elementName") or "",
        competency_name=label,
        competency_definition=str((unit or {}).get("compeUnitDef") or "").strip(),
    )
    surface_focus = str(task_frame.get("task_object") or "업무 판단과 수행 기준").strip()
    # The traceable task frame can end in a mechanical suffix such as
    # "관련 실무 적용·검증 절차".  Keep that wording in internal evidence,
    # but trim only the generic suffix from candidate-facing copy.
    candidate_surface_focus = re.sub(
        r"\s*(?:관련\s*)?(?:실무\s*)?(?:적용|활용)(?:·|ㆍ|/)?\s*검증\s*(?:절차|기준)\s*$",
        "",
        surface_focus,
    ).strip(" ·ㆍ/") or "해당 업무 수행 기준"
    surface_focus = candidate_surface_focus
    ksa_kind = str(task_frame.get("ksa_type") or "").strip()
    # Provider-free fallback requests can contain up to five slots for one
    # competency.  Reusing one static sentence after the KSA pool is exhausted
    # makes the UI look like it generated duplicate questions.  Keep the same
    # official focus while rotating the observable decision angle by slot.
    variation_axes = (
        "요구사항과 기준을 먼저 확인한",
        "마감·우선순위를 조정한",
        "누락·오류 자료를 대조한",
        "협업·보고 순서를 정한",
        "결과를 검증하고 재발을 막은",
    )
    variation_axis = variation_axes[index % len(variation_axes)]
    normalized_slot = _compact_question_intent_text(case_slot_id or "")
    normalized_code = _compact_question_intent_text(code)
    fallback_slot_id = (
        normalized_slot
        if normalized_slot
        else f"{normalized_code}:{_compact_question_intent_text(method)}"
    )
    fallback_slot_signature = (
        str(case_slot_signature or "").strip() or fallback_slot_id
    )

    # The full notice/JD and official NCS labels are retained in the request
    # trace and review payload, but neither belongs in candidate-facing
    # fallback text. OCR/provider failures can turn recruitment boilerplate or
    # a taxonomy label into a prompt that reads like an instruction. Model-
    # backed paths receive validated job context through their own leak guard;
    # the deterministic fallback stays generic and visibly degraded while the
    # trace panel keeps the exact unit.
    candidate_context = "해당 직무"

    if method == "상황면접":
        scenario_variants = (
            (
                "핵심 자료 오류와 일정 지연이 동시에 발생한 상황",
                "먼저 확인할 사실과 판단 기준, 위험을 통제할 행동·보고 순서",
                "먼저 확인해야 할 사실과 기준은 무엇입니까?",
                "결과가 기대와 다르면 어떤 후속 조치를 하시겠습니까?",
            ),
            (
                "자료 오류로 점검 기록과 현장 측정값이 일치하지 않고 일정 지연과 교대 인수인계가 겹친 상황",
                "기록의 신뢰도와 즉시 조치 범위를 가르는 판단 기준, 인수인계 전 행동 순서",
                "기록과 측정값 중 우선 대조할 자료와 그 이유는 무엇입니까?",
                "인수인계 뒤 새 정보가 확인되면 조치와 보고를 어떻게 수정하시겠습니까?",
            ),
            (
                "자료 오류가 섞인 안전 위험 민원과 설비 복구 요청이 동시에 접수되어 일정 지연이 우려되는 상황",
                "안전·서비스 영향의 우선순위, 확인 절차와 작업 중지·재개 기준",
                "안전 위험과 복구 요청의 우선순위를 정할 때 어떤 사실을 확인합니까?",
                "초기 위험 판단이 틀렸다고 드러나면 누구에게 무엇을 다시 보고하시겠습니까?",
            ),
            (
                "자료 오류를 바로잡아야 하는데 상급자의 신속 처리 지시와 규정상 승인 절차가 충돌해 일정 지연이 우려되는 상황",
                "권한 범위와 예외 승인 가능성을 확인한 뒤 위험을 줄이는 행동·보고 순서",
                "지시와 규정이 충돌한다는 사실을 확인할 근거는 무엇입니까?",
                "승인권자의 답변이 지연되면 업무를 어디까지 보류·진행하시겠습니까?",
            ),
            (
                "자료 오류와 일정 지연이 겹친 가운데 협업 부서마다 점검 기준이 달라 마감 전 결과를 확정해야 하는 상황",
                "공통 사실과 부서별 기준의 차이를 구분하고 결과 확정·보류를 결정하는 절차",
                "부서별 기준이 다를 때 공통으로 확인할 자료와 조정 주체는 누구입니까?",
                "확정 뒤 반대 근거가 나오면 기록과 다음 점검 계획을 어떻게 고치시겠습니까?",
            ),
        )
        scenario, action_focus, first_follow_up, final_follow_up = scenario_variants[
            index % len(scenario_variants)
        ]
        question = (
            f"{candidate_context}에서 {surface_focus}에 따라 판단해야 하는데, {scenario}입니다. "
            f"{action_focus}를 설명하고 위험을 통제할 행동 순서를 제시해 주세요."
        )
        follow_ups = [
            first_follow_up,
            f"{candidate_context}에서 {surface_focus}에 따라 그 행동을 선택한 이유와 직접 남길 기록은 무엇입니까?",
            final_follow_up,
        ]
        evaluation_points = ["사실 확인", "판단 기준", "행동 순서와 보고", "위험 통제와 후속점검"]
    elif method == "직무지식면접":
        question = (
            f"{candidate_context}에서 {variation_axis} {surface_focus}에 관해 확인해야 할 절차와 기준, 적용 범위, 산출물, "
            "예외상황 대응 및 오류 예방 방법을 설명해 주세요."
        )
        follow_ups = [
            f"{candidate_context}에서 {surface_focus}의 근거가 되는 문서나 사실은 무엇입니까?",
            "예외상황에서 기준 적용을 어떻게 조정하겠습니까?",
            "산출물 품질과 오류 예방은 어떻게 점검하겠습니까?",
        ]
        evaluation_points = ["절차·기준 이해", "예외상황 판단", "산출물 품질", "오류 예방"]
    elif method == "인바스켓면접":
        question = (
            f"[인바스켓과제] {candidate_context}에서 {variation_axis} 상황에 관련 요청, 오류 정정, 보고 문서가 동시에 들어왔습니다. "
            f"{surface_focus}에 따라 우선순위와 보고, 위임, 직접처리 판단을 제시하고, "
            "첫 조치와 기록 산출물을 포함해 주세요."
        )
        follow_ups = [
            f"{candidate_context}에서 {surface_focus}에 따라 가장 먼저 처리할 문서와 보류할 요청은 무엇입니까?",
            "보고, 위임, 직접처리를 나눈 판단 근거는 무엇입니까?",
            "처리 이후 기록과 후속 확인은 어떻게 남기겠습니까?",
        ]
        evaluation_points = ["우선순위 판단", "문서·요청 분류", "보고·위임·직접처리", "시간관리"]
    elif method == "발표면접":
        presentation_variants = (
            "현황 자료와 오류 사례를 바탕으로 원인과 대안을 제시",
            "운영 지표 변화와 민원 기록을 비교해 개선 우선순위를 제시",
            "절차 준수와 처리 속도 사이의 개선 대안을 비교해 제시",
            "품질 점검 결과를 근거로 실행계획과 성과지표를 제시",
            "제한된 인력·예산 안에서 단계별 개선 로드맵을 제시",
        )
        presentation_focus = presentation_variants[index % len(presentation_variants)]
        material_hint = ""
        presentation_source_text = str(presentation_material_text or "").strip()
        packet_task_hint = ""
        if presentation_source_text:
            # The presentation packet is built from this request's notice/JD
            # and selected NCS evidence. Surface its generated main task in
            # the candidate-facing question so the screen is not reduced to a
            # generic "analyze the materials" placeholder. Keep the fallback
            # bounded even when the packet contains long source excerpts.
            for line in presentation_source_text.splitlines():
                cleaned_line = re.sub(r"^[-•*]\s*", "", str(line or "").strip())
                if cleaned_line.startswith("발표 메인 과제:"):
                    raw_task_hint = cleaned_line.split(":", 1)[1].strip()
                    packet_task_hint = raw_task_hint[:200]
                    if len(raw_task_hint) > 200:
                        packet_task_hint = packet_task_hint.rsplit(" ", 1)[0].rstrip(" ,·") + "…"
                    break
            # Keep official KSA evidence in the expandable packet, but avoid
            # exposing the raw taxonomy label as if it were a candidate-facing
            # instruction. The question still names the concrete task and
            # points the candidate to the supplied source material.
            for term in dict.fromkeys(pool):
                cleaned_term = str(term or "").strip()
                if cleaned_term:
                    packet_task_hint = packet_task_hint.replace(cleaned_term, "해당 평가기준")
            material_hint = " 제공된 발표 자료의 수치·사실을 우선 사용하고 자료명을 근거로 밝혀 주세요."
        if packet_task_hint:
            # The packet task already carries the selected job's concrete
            # duty wording.  Do not prepend the official NCS detail/unit label
            # (for example, ``<detail> <unit> 업무``) to the candidate-facing
            # prompt; that taxonomy belongs in the traceability panel and is
            # otherwise flagged as a label leak by the realism gate.
            candidate_context = "제공된 직무 자료"
            question = (
                f"[발표과제] {candidate_context}에서 {packet_task_hint}"
                f"{material_hint} 위 자료에서 확인되는 사실과 제약을 근거로 대안 2가지와 선택 기준, "
                f"현황과 원인을 진단하고 {candidate_surface_focus} 관점의 실행계획·성과지표를 발표해 주세요. "
                "대안 간 합의가 어렵다면 결정권자에게 보고할 기준과 질의응답 답변도 제시해 주세요."
            )
        else:
            question = (
                f"[발표과제] {candidate_context}에서 {surface_focus}를 기준으로 {presentation_focus}합니다."
                f"{material_hint} 자료를 바탕으로 현황을 진단하고 대안 2가지, 실행계획과 성과지표를 발표한 뒤 "
                "질의응답에 답변해 주세요."
            )
        follow_ups = (
            [
                f"방금 발표에서 다룬 {candidate_surface_focus} 기준과 핵심 사실이 달라진다면 우선순위와 판단을 어떻게 수정하시겠습니까?",
                "앞서 선택한 대안의 실행 과정에서 일정 또는 품질 지표가 예상과 다르면 어떤 조치를 먼저 하시겠습니까?",
                "질의응답에서 자료와 반대되는 근거가 나오면 어떤 부분을 재검증하고 발표안을 수정하시겠습니까?",
            ]
            if packet_task_hint
            else [
                f"{candidate_context}의 {surface_focus} 판단에 사용한 핵심 자료와 그 자료를 선택한 이유는 무엇입니까?",
                "제시한 대안 중 우선안을 고른 기준과 예상되는 부작용은 무엇입니까?",
                "질의응답에서 자료와 반대되는 근거가 나오면 어떤 부분을 재검증하고 발표안을 수정하시겠습니까?",
            ]
        )
        evaluation_points = ["자료 근거 분석", "핵심 판단의 논리적 구조화", "대안 실행가능성", "성과지표와 질의응답"]
    elif method == "토론면접":
        approval_change_focus = "승인" in k1 and "변경" in k1
        if approval_change_focus:
            scenario = (
                "변경심의는 3영업일 뒤, 기존 범위 검수 준비 마감은 5영업일 뒤이고, "
                "변경 범위 착수·최종 인력배치·비용 집행에는 사전 승인이 필요하지만 "
                "영향분석·가용성 확인·잠정안 작성의 착수 해당 여부는 절차에 명시되지 않았습니다"
            )
            opposing_positions = (
                "절차 해석이 확정될 때까지 핵심 인력을 기존 승인 범위에 전담시키고 PM만 문서 영향분석을 하자는 입장과 "
                "변경 범위 실행·확정·비용 집행은 하지 않되 핵심 인력의 하루 2시간을 영향분석과 가용성 확인에 배정하자는 입장"
            )
        elif ksa_kind == "기술":
            scenario = "오류 위험과 처리 지연이 함께 발생했고 두 대안 모두 업무상 비용이 있는 상황입니다"
            opposing_positions = (
                "오류를 막기 위해 모든 건에 표준 절차를 끝까지 적용하자는 입장과 "
                "긴급·저위험 건은 핵심 단계만 먼저 처리하고 사후 점검하자는 입장"
            )
        elif ksa_kind == "태도":
            scenario = "원칙의 일관성과 협업 지연 위험을 동시에 고려해야 하는 상황입니다"
            opposing_positions = (
                "일관성과 책임을 위해 원칙을 예외 없이 지키자는 입장과 "
                "협업 지연을 줄이기 위해 상황별 재량을 허용하자는 입장"
            )
        else:
            scenario = "핵심 근거 일부가 늦게 도착해 검증 정확성과 처리 지연 위험을 함께 판단해야 하는 상황입니다"
            opposing_positions = (
                "근거가 모두 확인될 때까지 적용을 보류하자는 입장과 "
                "긴급·저위험 건은 조건부로 먼저 처리한 뒤 사후 검증하자는 입장"
            )
        debate_angle = (
            "초기 사실 확인 순서",
            "대안별 일정·품질 영향",
            "예외 적용과 권한 범위",
            "공통 실행안의 책임 배분",
            "합의 후 검증·후속점검",
        )[index % 5]
        question = (
            f"[토론과제] {candidate_context}에서 {surface_focus}에 따라 판단해야 하는 가운데 {scenario}. "
            f"{opposing_positions}이 충돌합니다. 각 입장의 근거와 위험, 타당성을 검토하세요. "
            f"특히 {debate_angle}을 쟁점으로 삼아 합의할 수 있다면 공통 실행안을, "
            "합의가 어렵다면 미합의 쟁점과 결정권자 이송 기준을 제시해 주세요."
        )
        follow_ups = [
            f"{candidate_context}의 {surface_focus}에 관한 초기 입장을 정하기 전에 어떤 문서와 사실을 확인하겠습니까?",
            "상대 입장에서 수용할 부분과 수용하기 어려운 부분을 어떤 기준으로 구분하겠습니까?",
            "공통안의 적용 범위·예외·검증·실행 책임 또는 미합의 이송 기준을 어떻게 정하겠습니까?",
        ]
        evaluation_points = [
            "사실·규정에 근거한 초기 입장",
            "대안별 영향 비교",
            "반대 근거 검토와 쟁점 조정",
            "공통안 또는 미합의 이송안의 실행 가능성",
        ]
    elif method == "창의적 문제해결력면접":
        creative_angle = (
            "반복 오류의 원인 가설을 재구성하고",
            "이해관계자 요구 변화까지 예측하고",
            "제한된 자원 안에서 대안을 설계하고",
            "새로운 자료·도구를 적용해 검증하고",
            "실행 후 재발 신호를 조기에 탐지하고",
        )[index % 5]
        question = (
            f"창의적 문제해결력과제에서 {candidate_context}의 {surface_focus}와 관련된 오류가 최근 3건 반복되고 "
            "필수 확인자료 1건이 누락된 채 마감까지 2시간 남았습니다. "
            f"전면 재작성과 우선 보완 중 선택해야 하는 압박 속에서 {creative_angle} "
            "미래예측 관점으로 문제를 정의하고 원인 가설, 창의적 대안 2가지, 검증 방법과 "
            "실현가능성, 의사결정 기준, 실행계획·성과지표·리스크 보완을 제시해 주세요."
        )
        follow_ups = [
            "방금 제시한 문제 정의에서 먼저 확인할 변화 신호가 달라지면 무엇을 수정하시겠습니까?",
            f"앞서 말씀하신 {surface_focus} 원인 가설을 뒷받침할 자료가 부족하거나 반대 결과가 나오면 검증 순서를 어떻게 바꾸시겠습니까?",
            "선택한 대안의 리스크가 예상보다 커졌다면 실행계획과 성과지표를 어떻게 조정하시겠습니까?",
        ]
        evaluation_points = ["미래예측과 문제 정의", "창의적 사고와 대안 도출", "검증 방법과 실현가능성", "의사결정과 실행계획"]
    else:
        experience_angle = (
            "요구사항과 기준을 확인한 뒤",
            "마감 압박 속 우선순위를 정한 뒤",
            "누락·오류 자료를 대조한 뒤",
            "협업 부서와 보고 순서를 조정한 뒤",
            "결과를 검증하고 재발을 막은 뒤",
        )[index % 5]
        if ksa_kind == "태도":
            experience_angle = (
                "압박과 이해 충돌 속에서 요구사항과 기준을 확인한 뒤",
                "마감 압박과 상충하는 요구 속에서 우선순위를 정한 뒤",
                "누락·오류 자료와 책임 범위를 대조한 뒤",
                "협업 부서와 이해 충돌 속에서 보고 순서를 조정한 뒤",
                "결과에 대한 책임을 확인하고 재발을 막은 뒤",
            )[index % 5]
        question = (
            f"{candidate_context}에서 업무를 수행하던 실제 상황 하나를 골라 {experience_angle} {candidate_surface_focus}에 따라 어떤 판단과 행동을 했는지 말씀해 주세요. "
            "그 결과를 문서·수치·기록·피드백으로 어떻게 확인했으며 이후 무엇을 개선했는지도 설명해 주세요."
        )
        follow_ups = [
            "방금 답변에서 당시 상황과 본인 역할, 직접 맡은 범위를 구분해 설명해 주세요.",
            f"앞서 말씀하신 {candidate_surface_focus} 관련 행동의 근거 자료나 기준이 달랐다면 어떤 부분을 바꾸시겠습니까?",
            "앞서 제시한 결과를 문서·수치·기록·피드백 중 무엇으로 확인했으며, 확인값이 다르면 어떻게 조정하시겠습니까?",
        ]
        evaluation_points = ["상황과 역할", "판단 근거", "실행 행동", "성과와 학습"]

    generic_case_facts = [
        f"{context_label} 관련 긴급 요청 1건과 일반 요청 2건이 같은 처리일에 접수됨",
        f"{surface_focus} 판단에 필요한 필수 확인자료 1건이 누락됨",
        "결재 또는 보고 마감까지 2시간이 남아 있음",
    ]
    approval_change_case_facts = [
        "변경심의 결과는 3영업일 뒤 확정될 예정임",
        "기존 승인 범위의 검수 준비 마감은 5영업일 뒤임",
        "핵심 인력 1명은 기존 승인 범위 작업에 배정되어 있음",
        "변경 범위 착수·최종 인력배치·비용 집행에는 사전 승인이 필요함",
        "영향분석·가용성 확인·잠정안 작성의 착수 해당 여부는 절차에 명시되지 않음",
    ]
    generic_case_materials = [
        {"source": "접수 현황표", "field": "요청 건수", "value": "긴급 1건, 일반 2건"},
        {"source": "자료 점검표", "field": "누락 자료", "value": "필수 확인자료 1건"},
        {"source": "처리 일정표", "field": "남은 시간", "value": "결재 또는 보고 마감까지 2시간"},
    ]
    approval_change_case_materials = [
        {"source": "변경요청서", "field": "심의 상태", "value": "승인 대기, 결정 예정 D+3"},
        {"source": "기존 범위 일정표", "field": "검수 준비 마감", "value": "D+5"},
        {"source": "역할·책임표", "field": "핵심 인력 가용성", "value": "1명, 기존 승인 범위 작업 중"},
        {"source": "변경관리 절차서", "field": "사전 승인 대상", "value": "변경 범위 착수·최종 인력배치·비용 집행"},
        {"source": "변경관리 절차서", "field": "해석 공백", "value": "영향분석·가용성 확인·잠정안 작성의 착수 해당 여부 미기재"},
    ]

    task_conditions: dict[str, Any] = {
        "candidate_instruction": "제시된 직무 상황을 기준으로 판단 근거, 구체적 행동과 결과 확인 방법을 구분해 답변하십시오.",
        "time_plan": [],
        "provided_materials": ["별도 자료 없음"],
        "required_outputs": ["판단 근거", "구체적 행동", "결과 확인 또는 후속점검"],
        "case_slot_id": fallback_slot_id,
        "case_signature": fallback_slot_signature,
        "standardization": "모든 지원자에게 동일한 자료, 기본 과제, 시간 조건과 허용된 후속질문 범위를 적용합니다.",
        "timing_basis": "기관 운영기준에서 동일 응답시간을 사전 확정합니다.",
    }
    if method == "토론면접":
        task_conditions.update(
            {
                "candidate_instruction": "근거 있는 초기 입장을 밝히고 상대 의견을 검토하십시오. 공통 실행안을 찾되 합의가 어렵다면 남은 쟁점과 결정권자 이송 기준을 제시하십시오.",
                "time_plan": [
                    {"phase": "개별 입장발표", "minutes": 1},
                    {"phase": "전체 토론", "minutes": 20},
                ],
                "provided_materials": (
                    ["변경요청서", "기존 범위 일정표", "역할·책임표", "변경관리 절차서", "토론 쟁점과 상반된 입장"]
                    if approval_change_focus
                    else ["접수 현황표", "자료 점검표", "처리 일정표", "토론 쟁점과 상반된 입장"]
                ),
                "case_facts": approval_change_case_facts if approval_change_focus else generic_case_facts,
                "case_materials": approval_change_case_materials if approval_change_focus else generic_case_materials,
                "required_outputs": [
                    "초기 입장과 확인 근거",
                    "반대 입장의 수용·불수용 기준",
                    "공통안의 적용 범위·예외·검증·실행 책임 또는 미합의 이송 기준",
                ],
                "timing_basis": "기관 운영기준에 따라 사전 확정한 동일 시간구조를 적용합니다.",
            }
        )
    elif method == "상황면접":
        task_conditions.update(
            {
                "candidate_instruction": "상황카드의 사실만을 사용해 확인 항목, 판단 기준, 행동·보고·후속조치 순서를 설명하십시오.",
                "time_plan": [{"phase": "개별 답변", "minutes": 5}],
                "provided_materials": ["접수 현황표", "자료 점검표", "처리 일정표"],
                "case_facts": generic_case_facts,
                "case_materials": generic_case_materials,
                "required_outputs": ["사실 확인 항목", "판단 기준과 위험 통제", "행동·보고·후속조치 순서"],
                "timing_basis": "기관 운영기준에 따라 사전 확정한 동일 응답시간을 적용합니다.",
            }
        )
    elif method == "발표면접":
        task_conditions.update(
            {
                "candidate_instruction": "제공자료를 근거로 현황과 원인을 구분하고 대안·실행계획·성과지표를 발표한 뒤 질의응답에 답하십시오.",
                "time_plan": [
                    {"phase": "준비", "minutes": 20},
                    {"phase": "발표", "minutes": 5},
                    {"phase": "질의응답", "minutes": 5},
                ],
                "provided_materials": ["접수 현황표", "자료 점검표", "처리 일정표"],
                "case_facts": [
                    *generic_case_facts,
                    "최근 4주 동안 같은 유형의 오류가 3건 반복됨",
                ],
                "case_materials": generic_case_materials,
                "required_outputs": ["현황·원인 진단", "대안 2가지와 우선순위", "실행계획과 성과지표"],
                "timing_basis": "기관 운영기준에 따라 사전 확정한 동일 시간구조를 적용합니다.",
            }
        )
        provided_material_text = str(presentation_material_text or "").strip()
        if provided_material_text:
            # Keep the candidate-facing question compact while exposing the
            # supplied source in the expandable task-conditions panel.
            task_conditions["provided_materials"] = list(dict.fromkeys([
                *task_conditions.get("provided_materials", []),
                "응시자 제공 발표 자료",
            ]))
            task_conditions["case_materials"] = [
                *list(task_conditions.get("case_materials") or []),
                {
                    "source": "응시자 제공 발표 자료",
                    "field": "원문",
                    "value": provided_material_text[:2400],
                },
            ]
            task_conditions["case_facts"] = list(dict.fromkeys([
                *task_conditions.get("case_facts", []),
                "응시자가 제공한 발표 자료의 사실·수치·제약을 우선 검토함",
            ]))
    elif method == "인바스켓면접":
        task_conditions.update(
            {
                "candidate_instruction": "동시에 접수된 문서를 분류해 처리 우선순위와 보고·위임·직접처리 결정을 기록하십시오.",
                "time_plan": [{"phase": "문서 검토 및 의사결정", "minutes": 30}],
                "provided_materials": ["접수 현황표", "자료 점검표", "처리 일정표", "업무분장표", "전결규정"],
                "case_facts": [
                    *generic_case_facts,
                    "상급자 보고가 필요한 예외 요청 1건이 포함되어 있음",
                    "응시자 역할은 실무 담당자이며 사실 확인·단순 자료 정정·담당자 협조 요청은 직접처리할 수 있음",
                    "대외 회신·일정 변경·예외 승인은 팀장 결재가 필요함",
                ],
                "case_materials": [
                    *generic_case_materials,
                    {"source": "업무분장표", "field": "응시자 역할·직접처리 범위", "value": "실무 담당자, 사실 확인·단순 자료 정정·담당자 협조 요청"},
                    {"source": "전결규정", "field": "팀장 결재 대상", "value": "대외 회신·일정 변경·예외 승인"},
                ],
                "required_outputs": ["문서별 우선순위", "보고·위임·직접처리 판단", "첫 조치와 후속점검"],
                "timing_basis": "기관 운영기준에 따라 사전 확정한 동일 시간구조를 적용합니다.",
            }
        )
    elif method == "창의적 문제해결력면접":
        task_conditions.update(
            {
                "candidate_instruction": "변화 신호를 근거로 문제를 재정의하고 복수 대안을 비교해 검증·실행·보완 계획을 설명하십시오.",
                "time_plan": [
                    {"phase": "준비", "minutes": 20},
                    {"phase": "해결안 설명", "minutes": 7},
                    {"phase": "질의응답", "minutes": 5},
                ],
                "case_materials": generic_case_materials,
                "provided_materials": ["접수 현황표", "자료 점검표", "처리 일정표"],
                "case_facts": [
                    *generic_case_facts,
                    "현재 방식으로 처리하면 다음 달 동일 오류가 2건 이상 재발할 가능성이 있음",
                ],
                "required_outputs": ["문제 정의와 원인 가설", "대안 2가지와 검증방법", "실행계획·성과지표·리스크 보완"],
                "timing_basis": "기관 운영기준에 따라 사전 확정한 동일 시간구조를 적용합니다.",
            }
        )

    if method in {"상황면접", "토론면접", "인바스켓면접", "창의적 문제해결력면접"}:
        authority_facts = [
            "응시자 역할은 실무 담당자이며 사실 확인·자료 수집·초안 작성·담당자 협조 요청은 직접 수행할 수 있음",
            "대외 회신·일정 변경·예외 승인은 팀장 결재가 필요하고 자료 대조는 품질담당자에게 협조 요청할 수 있음",
        ]
        authority_materials = [
            {
                "source": "업무분장표",
                "field": "응시자 역할·직접 수행 범위",
                "value": "실무 담당자, 사실 확인·자료 수집·초안 작성·담당자 협조 요청",
            },
            {
                "source": "전결규정",
                "field": "팀장 결재·협조 요청 범위",
                "value": "대외 회신·일정 변경·예외 승인 / 자료 대조 협조",
            },
        ]
        task_conditions["case_facts"] = list(dict.fromkeys([
            *task_conditions.get("case_facts", []),
            *authority_facts,
        ]))
        existing_rows = [
            row
            for row in task_conditions.get("case_materials", [])
            if isinstance(row, dict)
        ]
        existing_sources = {
            str(row.get("source") or "").strip()
            for row in existing_rows
        }
        task_conditions["case_materials"] = [
            *existing_rows,
            *[row for row in authority_materials if row["source"] not in existing_sources],
        ]
        task_conditions["provided_materials"] = list(dict.fromkeys([
            *task_conditions.get("provided_materials", []),
            "업무분장표",
            "전결규정",
        ]))

    return {
        "question": question,
        "type": method,
        "competency": label,
        "ncsClCd": code,
        "ncs_detail": detail,
        "question_focus": k1,
        "question_variation_axis": variation_axis,
        "question_focus_surface": candidate_surface_focus,
        "question_task_frame": task_frame,
        "question_evidence_id": str(task_frame.get("evidence_id") or ""),
        "question_evidence_required": bool(task_frame.get("evidence_id")),
        "question_focus_source": "official_ksa" if evidence_refs and evidence_refs[0] == k1 else "synthetic_template",
        "evaluation_points": evaluation_points,
        "ksa_refs": evidence_refs,
        "follow_ups": follow_ups,
        "task_conditions": task_conditions,
        "question_source": "template_fallback",
        "model_question_preserved": False,
    }


def generate_interview_questions_by_ncs_code(
    ncs_code: str,
    competency_name: str = "",
    target_count: int = 10,
    include_followups: bool = True,
    extra_context: str = "",
    api_key_override: str = "",
    generation_model: str = "",
    generation_provider: str = "openai_api",
) -> dict[str, Any]:
    code = str(ncs_code or "").strip()
    comp_name = competency_name or f"NCS-{code}"
    ncs_matches: list[dict[str, Any]] = [{"ncsClCd": code, "compeUnitName": comp_name}]
    is_sclass_mode = bool(code.isdigit() and len(code) == 6)
    sclass_units: list[dict[str, Any]] = []

    # Expand 6-digit small-category codes through NCS_MCP only. Local XLSX and
    # legacy HRDK APIs are not authoritative runtime sources for public
    # question generation.
    code6 = re.sub(r"[^0-9]", "", code)[:6]
    if is_sclass_mode:
        try:
            candidates = suggest_units_by_text([comp_name, code], max_units=80)
        except NcsMcpError:
            candidates = []
        sclass_units = [
            unit
            for unit in candidates
            if re.sub(r"\D", "", str(unit.get("ncsClCd") or ""))[:6] == code6
            or (
                comp_name
                and _norm_text(str(unit.get("ncsSclasCdnm") or ""))
                == _norm_text(comp_name)
            )
        ]
        ncs_matches = (
            sclass_units[: max(8, min(12, len(sclass_units)))]
            if sclass_units
            else []
        )
        if sclass_units:
            comp_name = (
                str(sclass_units[0].get("ncsSclasCdnm", "")).strip()
                or comp_name
            )

    desired_count = min(max(target_count, 1), 25)
    allow_template_fallback = str(os.getenv("NCS_ALLOW_TEMPLATE_FALLBACK", "false")).strip().lower() in {"1", "true", "yes", "y"}
    try:
        ai_topup_attempts = int(str(os.getenv("NCS_AI_TOPUP_ATTEMPTS", "4")).strip() or "4")
    except Exception:
        ai_topup_attempts = 2
    ai_topup_attempts = max(0, min(5, ai_topup_attempts))
    # The public single-question path must issue one semantic generation only.
    # Retrying an underfilled one-question result here used to hide as many as
    # four additional provider calls behind one accepted HTTP request.
    if desired_count == 1:
        ai_topup_attempts = 0
    used_template_fallback = False
    ncs_ksa = _safe_fetch_ncs_ksa_by_units(
        ncs_matches=ncs_matches,
        max_units=min(max(1, len(ncs_matches)), 8),
        max_factors_per_unit=6,
    )
    slot_units = [u for u in ncs_matches if str(u.get("ncsClCd", "")).strip()]
    if is_sclass_mode and sclass_units:
        ordered_slot_units: list[dict[str, Any]] = []
        seen_slot_codes: set[str] = set()
        for unit in sclass_units:
            unit_code = str(unit.get("ncsClCd", "")).strip()
            if not unit_code or unit_code in seen_slot_codes:
                continue
            seen_slot_codes.add(unit_code)
            ordered_slot_units.append(unit)
        slot_units = ordered_slot_units or slot_units
    if not slot_units:
        slot_units = [{"ncsClCd": code, "compeUnitName": comp_name}]
    slot_unit_codes: list[str] = []
    for unit in slot_units:
        unit_code = str(unit.get("ncsClCd", "")).strip()
        if unit_code and unit_code not in slot_unit_codes:
            slot_unit_codes.append(unit_code)
    if not slot_unit_codes:
        slot_unit_codes = [str(code).strip()]
    slot_unit_map: dict[str, dict[str, Any]] = {}
    for unit in slot_units:
        unit_code = str(unit.get("ncsClCd", "")).strip()
        if unit_code and unit_code not in slot_unit_map:
            slot_unit_map[unit_code] = unit
    if str(code).strip() and str(code).strip() not in slot_unit_map:
        slot_unit_map[str(code).strip()] = {"ncsClCd": str(code).strip(), "compeUnitName": comp_name}
    method_signature_to_label = {
        _compact_question_intent_text(method): method
        for method in _NCS_CODE_TEMPLATE_FALLBACK_METHODS
    }
    fallback_methods = list(_NCS_CODE_TEMPLATE_FALLBACK_METHODS)

    def _normalize_case_slot(value: Any) -> str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        code_part, _, method_part = raw.partition(":")
        if not method_part:
            return _compact_question_intent_text(raw)
        return f"{_compact_question_intent_text(code_part)}:{_compact_question_intent_text(method_part)}"

    def _slot_signature_for_row(row: dict[str, Any], fallback_index: int = 0) -> str:
        conditions = row.get("task_conditions") if isinstance(row, dict) else {}
        raw_slot = (
            str((conditions or {}).get("case_slot_id") or "").strip()
            if isinstance(conditions, dict)
            else ""
        )
        if raw_slot:
            return _normalize_case_slot(raw_slot)
        q_type = str((row or {}).get("type") or (row or {}).get("question_type") or "").strip()
        method_token = _compact_question_intent_text(q_type)
        if not method_token and fallback_methods:
            method_token = _compact_question_intent_text(fallback_methods[fallback_index % len(fallback_methods)])
        q_code = _compact_question_intent_text(str((row or {}).get("ncsClCd") or code).strip() or str(code))
        return f"{q_code}:{method_token}"

    def _required_case_slots(goal_count: int) -> list[str]:
        max_slots = max(1, min(goal_count, len(fallback_methods)))
        slots: list[str] = []
        for idx in range(max_slots):
            method_token = _compact_question_intent_text(fallback_methods[idx % len(fallback_methods)])
            unit_code = (
                _compact_question_intent_text(slot_unit_codes[idx % len(slot_unit_codes)])
                if slot_unit_codes
                else _compact_question_intent_text(str(code).strip())
            )
            slots.append(f"{unit_code}:{method_token}")
        return slots

    def _method_from_slot(slot_id: str) -> str:
        token = (
            _compact_question_intent_text(slot_id.split(":", 1)[1])
            if ":" in slot_id
            else _compact_question_intent_text(slot_id)
        )
        return method_signature_to_label.get(token, fallback_methods[0])

    generated_raw = _generate_questions_with_openai_from_ncs(
        jd_text="",
        ncs_matches=ncs_matches,
        ncs_ksa=ncs_ksa,
        target_count=desired_count,
        mode="ncs_code_only",
        extra_context=extra_context,
        api_key_override=api_key_override,
        generation_model=generation_model,
        generation_provider=generation_provider,
    )
    generated: list[dict[str, Any]] = []
    seen_question_keys: set[str] = set()
    allowed_ncs_codes = {
        str(item.get("ncsClCd", "")).strip()
        for item in ncs_matches
        if str(item.get("ncsClCd", "")).strip()
    }

    def _merge_generated(rows: list[dict[str, Any]] | None) -> int:
        added = 0
        for row in (rows or []):
            if not isinstance(row, dict):
                continue
            q_text = str(row.get("question", "")).strip()
            q_key = normalize_question_dedup_key(q_text)
            if not q_key or q_key in seen_question_keys:
                continue
            seen_question_keys.add(q_key)
            r = dict(row)
            row_code = str(r.get("ncsClCd", "")).strip()
            if len(allowed_ncs_codes) == 1:
                row_code = next(iter(allowed_ncs_codes))
            elif row_code not in allowed_ncs_codes:
                continue
            r["ncsClCd"] = row_code or code
            generated.append(r)
            added += 1
            if len(generated) >= desired_count:
                break
        return added

    _merge_generated(generated_raw)
    required_case_slots = _required_case_slots(desired_count)

    for _ in range(ai_topup_attempts):
        if len(generated) >= desired_count:
            break
        remaining = desired_count - len(generated)
        existing_questions = [str(x.get("question", "")).strip() for x in generated if str(x.get("question", "")).strip()]
        dedup_hint = ""
        if existing_questions:
            dedup_hint = "[중복 금지 - 이미 생성된 질문]\n" + "\n".join(f"- {q}" for q in existing_questions[:12])
        if extra_context:
            dedup_hint = f"{extra_context}\n{dedup_hint}".strip()
        extra_raw = _generate_questions_with_openai_from_ncs(
            jd_text="",
            ncs_matches=ncs_matches,
            ncs_ksa=ncs_ksa,
            target_count=min(desired_count, remaining + 2),
            mode="ncs_code_only",
            extra_context=dedup_hint,
            api_key_override=api_key_override,
            generation_model=generation_model,
            generation_provider=generation_provider,
        )
        _merge_generated(extra_raw)

    # Sclass mode rule:
    # 1) one main question per unit first
    # 2) only if units < desired_count, then allow duplicates to fill.
    if is_sclass_mode and sclass_units and allow_template_fallback:
        ordered_units = [u for u in sclass_units if str(u.get("ncsClCd", "")).strip()]
        unique_units: list[dict[str, Any]] = []
        seen_units: set[str] = set()
        for u in ordered_units:
            uc = str(u.get("ncsClCd", "")).strip()
            if uc in seen_units:
                continue
            seen_units.add(uc)
            unique_units.append(u)
        required_unique = min(desired_count, len(unique_units))

        by_code: dict[str, list[dict[str, Any]]] = {}
        for g in generated:
            gc = str(g.get("ncsClCd", "")).strip()
            by_code.setdefault(gc, []).append(g)

        distributed: list[dict[str, Any]] = []
        for u in unique_units[:required_unique]:
            uc = str(u.get("ncsClCd", "")).strip()
            picked = None
            if by_code.get(uc):
                picked = by_code[uc].pop(0)
            if not picked and allow_template_fallback:
                unit_ksa_rows = _safe_fetch_ncs_ksa_by_units(ncs_matches=[u], max_units=1, max_factors_per_unit=3)
                unit_evidence = [
                    str(x.get("factorName", "")).strip()
                    for x in unit_ksa_rows
                    if str(x.get("factorName", "")).strip()
                ]
                unit_ksa = list(unit_evidence)
                if len(unit_ksa) < 2:
                    unit_ksa = unit_ksa + ["업무 우선순위 설정", "이해관계자 협업"]
                used_template_fallback = True
                picked = _build_ncs_code_template_fallback_question(
                    unit=u,
                    comp_name=str(u.get("compeUnitName", "")).strip() or comp_name,
                    ncs_code=uc,
                    ksa_terms=unit_ksa,
                    evidence_terms=unit_evidence,
                    evidence_rows=unit_ksa_rows,
                    index=len(distributed),
                )
            if picked:
                distributed.append(picked)

        # Fill remainder only when unique units are fewer than desired count.
        if len(unique_units) < desired_count:
            leftovers: list[dict[str, Any]] = []
            for arr in by_code.values():
                leftovers.extend(arr)
            for g in leftovers:
                if len(distributed) >= desired_count:
                    break
                distributed.append(g)

        generated = distributed[:desired_count]

    if len(generated) < desired_count and allow_template_fallback:
        used_template_fallback = True
        fallback_focuses = ["업무 우선순위 설정", "이해관계자 협업", "성과 점검 및 개선"]
        ksa_by_code: dict[str, list[dict[str, Any]]] = {}
        ksa_by_name: dict[str, list[dict[str, Any]]] = {}
        unscoped_ksa: list[dict[str, Any]] = []
        for row in ncs_ksa or []:
            factor = str(row.get("factorName", "")).strip()
            if not factor:
                continue
            unit_code = str(row.get("ncsClCd", "")).strip()
            unit_name = str(row.get("compeUnitName", "")).strip()
            if unit_code:
                ksa_by_code.setdefault(unit_code, []).append(row)
            if unit_name:
                ksa_by_name.setdefault(unit_name, []).append(row)
            if not unit_code and not unit_name:
                unscoped_ksa.append(row)
        existing = {normalize_question_dedup_key(str(x.get("question", ""))) for x in generated}
        idx = 0
        while len(generated) < desired_count and idx < desired_count * 4:
            unit = ncs_matches[idx % len(ncs_matches)] if ncs_matches else {}
            unit_code = str(unit.get("ncsClCd", "")).strip()
            unit_name = str(unit.get("compeUnitName", "")).strip()
            unit_evidence_rows = list(ksa_by_code.get(unit_code) or ksa_by_name.get(unit_name) or [])
            if not unit_evidence_rows and len(ncs_matches) <= 1:
                unit_evidence_rows = list(unscoped_ksa)
            unit_evidence = [
                str(row.get("factorName") or "").strip()
                for row in unit_evidence_rows
                if str(row.get("factorName") or "").strip()
            ]
            unit_ksa = list(unit_evidence) or list(fallback_focuses)
            q = _build_ncs_code_template_fallback_question(
                unit=unit,
                comp_name=comp_name,
                ncs_code=unit_code or code,
                ksa_terms=unit_ksa,
                evidence_terms=unit_evidence,
                evidence_rows=unit_evidence_rows,
                index=idx,
            )
            qtext = str(q.get("question", "")).strip()
            key = normalize_question_dedup_key(qtext)
            idx += 1
            if not key or key in existing:
                continue
            existing.add(key)
            generated.append(q)

    if used_template_fallback and required_case_slots:
        current_slots = [
            _slot_signature_for_row(row, idx)
            for idx, row in enumerate(generated)
        ]
        slot_presence = {slot: idx for idx, slot in enumerate(current_slots)}
        missing_case_slots = [slot for slot in required_case_slots if slot not in slot_presence]
        if missing_case_slots:
            required_slot_set = set(required_case_slots)
            for missing_slot in missing_case_slots:
                unit_code = missing_slot.split(":", 1)[0] if ":" in missing_slot else str(code).strip()
                target_unit = slot_unit_map.get(unit_code) or slot_unit_map.get(slot_unit_codes[0]) or {"ncsClCd": str(code).strip(), "compeUnitName": comp_name}
                unit_ksa_rows = [
                    row for row in (ncs_ksa or [])
                    if str(row.get("ncsClCd", "")).strip() == str(unit_code).strip()
                ]
                unit_evidence = [
                    str(row.get("factorName", "")).strip()
                    for row in unit_ksa_rows
                    if str(row.get("factorName", "")).strip()
                ]
                official_unit_evidence = list(unit_evidence)
                if not unit_evidence:
                    unit_evidence = ["업무 우선순위 설정", "이해관계자 협업", "성과 점검 및 개선"]
                replacement_method = _method_from_slot(missing_slot)

                replacement = _build_ncs_code_template_fallback_question(
                    unit=target_unit,
                    comp_name=str(target_unit.get("compeUnitName", "")).strip() or comp_name,
                    ncs_code=unit_code or str(code).strip(),
                    ksa_terms=unit_evidence,
                    evidence_terms=official_unit_evidence,
                    evidence_rows=unit_ksa_rows,
                    index=len(generated),
                    method_override=replacement_method,
                    case_slot_id=missing_slot,
                    case_slot_signature=f"coverage:{missing_slot}",
                )
                replacement["type"] = replacement_method

                replace_target: int | None = None
                if replace_target is None:
                    for idx, slot in enumerate(current_slots):
                        if slot not in required_slot_set:
                            replace_target = idx
                            break
                if replace_target is None:
                    counts: dict[str, int] = {}
                    for slot in current_slots:
                        counts[slot] = counts.get(slot, 0) + 1
                    for idx, slot in enumerate(current_slots):
                        if counts.get(slot, 0) > 1 and slot != missing_slot:
                            replace_target = idx
                            break
                if replace_target is None and len(generated) < desired_count:
                    replace_target = len(generated)
                    generated.append(replacement)
                    current_slots.append(missing_slot)
                elif replace_target is None:
                    continue
                else:
                    generated[replace_target] = replacement
                    current_slots[replace_target] = missing_slot
            slot_presence = {slot: idx for idx, slot in enumerate(current_slots)}

    generated = _apply_entry_level_policy_to_questions(generated)
    grounded_generated: list[dict[str, Any]] = []
    for raw_question in generated:
        question = dict(raw_question)
        verified_focus, verified_refs, focus_source = _official_question_grounding(
            question,
            ncs_ksa,
            default_code=code,
        )
        question["question_focus"] = verified_focus
        question["ksa_refs"] = verified_refs
        question["question_focus_source"] = focus_source
        grounded_generated.append(question)
    generated = grounded_generated

    for raw_question in generated:
        if not isinstance(raw_question, dict):
            continue
        task_conditions = dict(raw_question.get("task_conditions") or {})
        if not task_conditions.get("case_facts"):
            task_conditions["case_facts"] = [
                "사례 기반 판단의 확인자료와 근거를 제시해야 함",
            ]
            raw_question["task_conditions"] = task_conditions

    main_questions = [
        {
            "question": str(q.get("question", "")).strip(),
            "evaluation_points": list(q.get("evaluation_points", []) or []),
            "question_type": str(q.get("type", "면접질문")).strip() or "면접질문",
            "type": str(q.get("type", "면접질문")).strip() or "면접질문",
            "competency": str(q.get("competency") or comp_name).strip() or comp_name,
            "ncsClCd": str(q.get("ncsClCd", "")).strip(),
            "ncs_detail": str(q.get("ncs_detail") or "").strip(),
            "question_focus": _primary_question_focus(q),
            "question_focus_surface": str(q.get("question_focus_surface") or "").strip(),
            "question_task_frame": dict(q.get("question_task_frame") or {}),
            "question_evidence_id": str(q.get("question_evidence_id") or "").strip(),
            "question_evidence_required": bool(q.get("question_evidence_required")),
            "question_focus_source": str(q.get("question_focus_source", "")).strip(),
            "ksa_refs": list(q.get("ksa_refs", []) or []) if isinstance(q.get("ksa_refs"), list) else [],
            "follow_ups": list(q.get("follow_ups", []) or []),
            "task_conditions": dict(q.get("task_conditions") or {}),
            "question_source": str(q.get("question_source", "")).strip(),
            "model_question_preserved": bool(q.get("model_question_preserved")),
            "candidate_selection_policy": str(q.get("candidate_selection_policy") or "").strip(),
            "candidate_pool_count": int(q.get("candidate_pool_count") or 0),
            "candidate_quality_score": float(q.get("candidate_quality_score") or 0.0),
            "candidate_selection_score": float(q.get("candidate_selection_score") or 0.0),
            "candidate_diversity_axes": dict(q.get("candidate_diversity_axes") or {}),
        }
        for q in generated
    ]
    _refresh_ncs_code_main_question_repeat_metadata(main_questions)

    follow_up_questions: list[dict[str, Any]] = []
    if include_followups:
        for i, q in enumerate(generated):
            raw_fu = q.get("follow_ups")
            if isinstance(raw_fu, list):
                follow_ups = [str(x).strip() for x in raw_fu if str(x).strip()]
            else:
                one = str(q.get("follow_up", "")).strip()
                follow_ups = [one] if one else []
            if len(follow_ups) < 3:
                fallback_fus = [
                    "당시 상황과 본인 역할을 구체적으로 설명해 주세요.",
                    "가장 어려웠던 지점과 대응 방식을 말씀해 주세요.",
                    "결과와 배운 점, 다음 개선 계획을 말씀해 주세요.",
                ]
                for f in fallback_fus:
                    if len(follow_ups) >= 3:
                        break
                    follow_ups.append(f)
            for j, fu in enumerate(follow_ups[:3], start=1):
                follow_up_questions.append(
                    {
                        "follow_up": fu,
                        "for_question_index": i,
                        "step": j,
                        "purpose": "심층 확인",
                    }
                )

    generation_mode = "ai_autonomous_ncs_code_only"
    if used_template_fallback:
        generation_mode = "hybrid_ai_with_template_fallback"
    elif not main_questions:
        generation_mode = "ai_generation_empty_no_fallback"

    all_questions_grounded = bool(main_questions) and all(
        row.get("question_focus_source") == "official_ksa"
        and bool(row.get("ksa_refs"))
        and bool(str(row.get("ncsClCd") or "").strip())
        for row in main_questions
    )
    final_slots = [
        _slot_signature_for_row(row, idx)
        for idx, row in enumerate(generated)
    ]
    used_slots_set = set(final_slots)
    covered_required_slots = [
        slot for slot in required_case_slots
        if slot in used_slots_set
    ]
    case_coverage = {
        "required_slots": required_case_slots,
        "required_count": len(required_case_slots),
        "used_slots": sorted(used_slots_set),
        "used_count": len(used_slots_set),
        "missing_slots": [slot for slot in required_case_slots if slot not in used_slots_set],
        "covered_required_count": len(covered_required_slots),
        "coverage_ratio": round(len(covered_required_slots) / len(required_case_slots), 4) if required_case_slots else 0.0,
    }
    result = {
        "ncs_code": code,
        "competency_name": comp_name,
        "generation_mode": generation_mode,
        "generation_provider": generation_provider,
        "provider_generation_model": str(
            (generated[0] if generated else {}).get("provider_generation_model")
            or generation_model
        ).strip(),
        "provider_candidate_variant_count": int(
            (generated[0] if generated else {}).get("provider_candidate_variant_count") or 0
        ),
        "provider_candidate_variant_received_count": int(
            (generated[0] if generated else {}).get(
                "provider_candidate_variant_received_count"
            )
            or 0
        ),
        "case_coverage": case_coverage,
        "main_questions": main_questions,
        "follow_up_questions": follow_up_questions,
        "question_count": len(main_questions),
        "follow_up_count": len(follow_up_questions),
        "total_count": len(main_questions) + len(follow_up_questions),
        "template_fallback_used": used_template_fallback,
        "ncs_ksa_available": bool(ncs_ksa) and all_questions_grounded,
        "official_ksa_evidence": _server_official_ksa_evidence(ncs_ksa),
        "warning": (
            ""
            if ncs_ksa and all_questions_grounded
            else "official_ncs_ksa_unavailable_or_question_grounding_failed"
        ),
    }
    if not main_questions:
        result["error"] = "ai_generation_empty"
    return result


def _is_generic_interview_set(items: list[dict[str, Any]]) -> bool:
    return False


def _build_flat_interview_questions(
    ncs_matches: list[dict[str, Any]],
    ncs_ksa: list[dict[str, Any]] | None = None,
    ncs_context: dict[str, Any] | None = None,
    strengths: str = "",
    target_count: int = 50,
    run_seed: int | None = None,
) -> list[dict[str, Any]]:
    _ = run_seed
    return _generate_questions_with_openai_from_ncs(
        jd_text="",
        strengths=strengths,
        ncs_matches=ncs_matches,
        ncs_ksa=ncs_ksa,
        ncs_context=ncs_context,
        target_count=target_count,
        mode="fallback_flat",
    )


def build_flat_interview_questions_fallback(
    ncs_matches: list[dict[str, Any]],
    ncs_ksa: list[dict[str, Any]] | None = None,
    ncs_context: dict[str, Any] | None = None,
    strengths: str = "",
    target_count: int = 50,
) -> list[dict[str, Any]]:
    return _build_flat_interview_questions(
        ncs_matches=ncs_matches,
        ncs_ksa=ncs_ksa,
        ncs_context=ncs_context,
        strengths=strengths,
        target_count=target_count,
        run_seed=None,
    )


_QUESTION_SIMILARITY_STOPWORDS: set[str] = {
    "그리고",
    "또한",
    "또는",
    "대해",
    "관련",
    "경우",
    "업무",
    "직무",
    "상황",
    "질문",
    "설명",
    "해주세요",
    "주십시오",
    "무엇",
    "어떻게",
    "이유",
    "기준",
    "경험",
    "있나요",
    "있다면",
}


_ENTRY_LEVEL_TRIGGER_RE = re.compile(
    r"(수행\s*경험|경험이\s*있다면|해본\s*경험|참여했던|담당했던|실무에서|업무를\s*수행|수립한\s*경험|운영한\s*경험)"
)
_ENTRY_LEVEL_ALREADY_RE = re.compile(r"(유사\s*사례|가정\s*상황|가정해|가정하여|가정하고)")


def _needs_entry_level_softening(question: str) -> bool:
    q = str(question or "").strip()
    if not q:
        return False
    if _ENTRY_LEVEL_ALREADY_RE.search(q):
        return False
    return bool(_ENTRY_LEVEL_TRIGGER_RE.search(q))


def _soften_entry_level_question(question: str) -> str:
    q = str(question or "").strip()
    if not q:
        return q
    if not _needs_entry_level_softening(q):
        return q

    replacements: list[tuple[str, str]] = [
        (r"수행\s*경험에서", "수행했거나 유사 상황을 가정한 사례에서"),
        (r"수행\s*경험을", "수행했거나 유사 상황을 가정한 사례를"),
        (r"수립한\s*경험", "수립했거나 유사 상황을 가정한 사례"),
        (r"운영한\s*경험", "운영했거나 유사 상황을 가정한 사례"),
        (r"경험이\s*있다면", "경험이나 유사 사례(가정 상황 포함)가 있다면"),
        (r"경험에\s*대해", "경험 또는 유사 사례(가정 상황 포함)에 대해"),
        (r"참여했던", "참여했거나 유사한"),
        (r"담당했던", "담당했거나 유사한"),
    ]
    out = q
    for pattern, repl in replacements:
        new_q = re.sub(pattern, repl, out, count=1)
        if new_q != out:
            out = new_q
            break
    if out == q:
        out = re.sub(r"경험", "경험 또는 유사 사례(가정 상황 포함)", q, count=1)
    return out


def _apply_entry_level_policy_to_questions(items: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in (items or []):
        if not isinstance(row, dict):
            continue
        r = dict(row)
        r["question"] = _soften_entry_level_question(str(r.get("question", "")).strip())

        fus = r.get("follow_ups")
        if isinstance(fus, list):
            r["follow_ups"] = [_soften_entry_level_question(str(x).strip()) for x in fus if str(x).strip()]
            if r["follow_ups"]:
                r["follow_up"] = r["follow_ups"][0]
        else:
            one = str(r.get("follow_up", "")).strip()
            if one:
                r["follow_up"] = _soften_entry_level_question(one)
        out.append(r)
    return out


def normalize_question_dedup_key(text: str) -> str:
    raw = str(text or "").strip().lower()
    if not raw:
        return ""
    raw = re.sub(r"\[[^\]]+\]", " ", raw)
    raw = re.sub(r"[^0-9a-z가-힣 ]+", " ", raw)
    raw = re.sub(r"\s+", " ", raw).strip()
    return raw


def _question_token_set(text: str) -> set[str]:
    key = normalize_question_dedup_key(text)
    if not key:
        return set()
    return {
        tok for tok in key.split()
        if len(tok) >= 2 and tok not in _QUESTION_SIMILARITY_STOPWORDS
    }


def _char_ngram_set(text: str, size: int = 3) -> set[str]:
    raw = normalize_question_dedup_key(text).replace(" ", "")
    if len(raw) < size:
        return set()
    return {raw[i:i + size] for i in range(len(raw) - size + 1)}


def is_similar_question_text(
    text_a: str,
    text_b: str,
    seq_ratio_threshold: float = 0.92,
    jaccard_threshold: float = 0.80,
    min_token_overlap: int = 4,
) -> bool:
    a = normalize_question_dedup_key(text_a)
    b = normalize_question_dedup_key(text_b)
    if not a or not b:
        return False
    if a == b:
        return True

    seq_ratio = SequenceMatcher(None, a, b).ratio()
    if seq_ratio >= seq_ratio_threshold:
        return True

    a_tokens = _question_token_set(a)
    b_tokens = _question_token_set(b)
    if not a_tokens or not b_tokens:
        return False

    if a_tokens and b_tokens:
        inter = a_tokens.intersection(b_tokens)
        if inter:
            min_size = min(len(a_tokens), len(b_tokens))
            if min_size > 0 and (len(inter) / min_size) >= 0.62 and len(inter) >= 5:
                return True
        if len(inter) >= min_token_overlap:
            union = a_tokens.union(b_tokens)
            if union and (len(inter) / len(union)) >= jaccard_threshold:
                return True

    a_ngrams = _char_ngram_set(a, size=3)
    b_ngrams = _char_ngram_set(b, size=3)
    if not a_ngrams or not b_ngrams:
        return False
    n_union = a_ngrams.union(b_ngrams)
    if not n_union:
        return False
    return (len(a_ngrams.intersection(b_ngrams)) / len(n_union)) >= 0.68


def _normalize_question_key(q: dict[str, Any]) -> str:
    text_key = normalize_question_dedup_key(str((q or {}).get("question", "")))
    if text_key:
        return text_key
    # follow_ups(배열) 또는 follow_up(레거시 문자열) 모두 지원
    fus = (q or {}).get("follow_ups")
    fallback = fus[0] if isinstance(fus, list) and fus else str((q or {}).get("follow_up", ""))
    return normalize_question_dedup_key(fallback)


def _ensure_diverse_question_set(
    generated: list[dict[str, Any]] | None,
    fallback_pool: list[dict[str, Any]],
    target_count: int = 50,
) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    seen: set[str] = set()
    seen_questions: list[str] = []
    for src in (generated or []) + (fallback_pool or []):
        key = _normalize_question_key(src)
        if not key or key in seen:
            continue
        q_text = str(src.get("question", "")).strip()
        if any(is_similar_question_text(q_text, prev) for prev in seen_questions):
            continue
        seen.add(key)
        seen_questions.append(q_text)
        # follow_ups(배열) 우선 사용, 없으면 follow_up(문자열) 호환
        raw_fu = src.get("follow_ups")
        if isinstance(raw_fu, list) and raw_fu:
            follow_ups = [str(f).strip() for f in raw_fu if str(f).strip()]
        else:
            single = str(src.get("follow_up", "")).strip()
            follow_ups = [single] if single else []
        merged.append(
            {
                "type": str(src.get("type", "상황면접")).strip() or "상황면접",
                "competency": str(src.get("competency", "")).strip(),
                "ncsClCd": str(src.get("ncsClCd", "")).strip(),
                "question": str(src.get("question", "")).strip(),
                "follow_ups": follow_ups,
                "evaluation_points": list(src.get("evaluation_points", []) or []),
                "question_source": str(src.get("question_source", "")).strip(),
                "question_evidence_id": str(src.get("question_evidence_id", "")).strip(),
                "question_evidence_required": bool(src.get("question_evidence_id")),
                "question_focus_surface": str(src.get("question_focus_surface", "")).strip(),
                "question_focus": str(src.get("question_focus", "")).strip(),
                "ksa_refs": [
                    str(value).strip()
                    for value in (src.get("ksa_refs") or [])
                    if str(value).strip()
                ]
                if isinstance(src.get("ksa_refs"), list)
                else [],
            }
        )
        if len(merged) >= target_count:
            break
    return merged[:target_count]


def _build_interview_by_competency_from_questions(
    questions: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for q in (questions or []):
        comp = str((q or {}).get("competency", "")).strip() or "핵심 직무"
        code = str((q or {}).get("ncsClCd", "")).strip()
        key = (comp, code)
        grouped.setdefault(key, []).append(
            {
                "question": str((q or {}).get("question", "")).strip(),
                "follow_ups": list((q or {}).get("follow_ups", []) or []),
                "evaluation_points": list((q or {}).get("evaluation_points", []) or []),
            }
        )
    out: list[dict[str, Any]] = []
    for (comp, code), qset in grouped.items():
        out.append({"competency": comp, "ncsClCd": code, "questions": qset})
    return out


def rank_ncs_matches_by_jd(
    jd_text: str,
    ncs_items: list[dict[str, Any]],
    top_k: int = 8,
    preferred_sclass: list[str] | None = None,
    per_sclass_limit: int | None = None,
) -> list[dict[str, Any]]:
    """
    Stage-1 unit ranking:
    - Query(JD + duty/evaluation context) vs unit text TF-IDF(char n-gram) cosine
    - Blend with source score
    - Diversify by sclass to avoid one broad sclass dominating top-k
    """
    query_text = _repair_mojibake(jd_text or "")
    if not query_text.strip():
        return []

    rows: list[dict[str, Any]] = []
    seen_codes: set[str] = set()
    for it in (ncs_items or []):
        if not isinstance(it, dict):
            continue
        code = str(it.get("ncsClCd", "")).strip()
        if not code or code in seen_codes:
            continue
        seen_codes.add(code)
        rows.append(dict(it))
    if not rows:
        return []

    keep_n = max(1, int(top_k or 8))

    sim_w = 0.88
    src_w = 0.12
    try:
        sim_w = float(str(os.getenv("NCS_UNIT_SIMILARITY_WEIGHT", "0.88")).strip())
    except Exception:
        sim_w = 0.88
    try:
        src_w = float(str(os.getenv("NCS_UNIT_SOURCE_WEIGHT", "0.12")).strip())
    except Exception:
        src_w = 0.12
    sim_w = max(0.0, sim_w)
    src_w = max(0.0, src_w)
    if sim_w <= 0 and src_w <= 0:
        sim_w = 1.0
    total_w = sim_w + src_w
    sim_w = sim_w / total_w
    src_w = src_w / total_w

    min_similarity = 0.02
    try:
        min_similarity = float(str(os.getenv("NCS_UNIT_MIN_SIMILARITY", "0.02")).strip())
    except Exception:
        min_similarity = 0.02
    min_similarity = max(0.0, min(0.5, min_similarity))

    preferred_keys = {
        _sclass_norm_key(x)
        for x in (preferred_sclass or [])
        if _sclass_norm_key(str(x or ""))
    }

    # Normalize source scores.
    raw_scores: list[float] = []
    for row in rows:
        try:
            raw_scores.append(float(row.get("score", 0.0) or 0.0))
        except Exception:
            raw_scores.append(0.0)
    score_min = min(raw_scores) if raw_scores else 0.0
    score_max = max(raw_scores) if raw_scores else 1.0

    def _norm_source_score(v: float) -> float:
        if score_max > score_min:
            return (v - score_min) / (score_max - score_min)
        if 0.0 <= v <= 1.0:
            return v
        return 1.0 if v > 0 else 0.0

    query_tf = _char_ngram_tf(query_text, ngram_min=2, ngram_max=4)
    doc_tfs: list[Counter[str]] = []
    doc_texts: list[str] = []
    for row in rows:
        doc_text = _repair_mojibake(
            " ".join(
                [
                    str(row.get("ncsSclasCdnm", "")).strip(),
                    str(row.get("ncsSubdCdnm", "")).strip(),
                    str(row.get("compeUnitName", "")).strip(),
                    str(row.get("compeUnitDef", "")).strip(),
                ]
            )
        )
        doc_texts.append(doc_text)
        doc_tfs.append(_char_ngram_tf(doc_text, ngram_min=2, ngram_max=4))

    similarity_scores: list[float] = [0.0] * len(rows)
    if query_tf and any(doc_tfs):
        df: Counter[str] = Counter()
        for tf in doc_tfs:
            df.update(tf.keys())

        doc_count = max(1, len(doc_tfs))
        idf = {term: (math.log((doc_count + 1) / (freq + 1)) + 1.0) for term, freq in df.items()}

        query_w: dict[str, float] = {}
        for term, cnt in query_tf.items():
            if term not in idf:
                continue
            query_w[term] = (1.0 + math.log(max(1, cnt))) * idf[term]
        query_norm = math.sqrt(sum(v * v for v in query_w.values())) if query_w else 0.0

        if query_norm > 0:
            for i, tf in enumerate(doc_tfs):
                if not tf:
                    continue
                dot = 0.0
                doc_norm_sq = 0.0
                for term, cnt in tf.items():
                    weight = (1.0 + math.log(max(1, cnt))) * idf.get(term, 0.0)
                    if weight <= 0:
                        continue
                    doc_norm_sq += weight * weight
                    qv = query_w.get(term)
                    if qv:
                        dot += qv * weight
                doc_norm = math.sqrt(doc_norm_sq)
                if doc_norm > 0 and dot > 0:
                    similarity_scores[i] = dot / (query_norm * doc_norm)

    # Human-readable hit keywords for diagnostics/UI.
    focus_terms = [t for t in _extract_focus_terms(query_text) if str(t or "").strip()][:100]
    if not focus_terms:
        focus_terms = re.findall(r"[\uAC00-\uD7A3]{2,12}", query_text)[:100]

    scored_rows: list[dict[str, Any]] = []
    for i, row in enumerate(rows):
        code = str(row.get("ncsClCd", "")).strip()
        sclass_nm = str(row.get("ncsSclasCdnm", "")).strip()
        sclass_key = _sclass_norm_key(sclass_nm)
        sim = float(similarity_scores[i] or 0.0)
        try:
            src_raw = float(row.get("score", 0.0) or 0.0)
        except Exception:
            src_raw = 0.0
        src_norm = _norm_source_score(src_raw)
        final_score = (sim_w * sim) + (src_w * src_norm)
        # Penalize candidates outside preferred sclass set when that set exists.
        if preferred_keys and sclass_key and sclass_key not in preferred_keys:
            final_score -= 0.15

        doc_text = doc_texts[i]
        hit = [k for k in focus_terms if k and k in doc_text][:8]
        scored_rows.append(
            {
                "ncsClCd": code,
                "compeUnitName": str(row.get("compeUnitName", "")).strip(),
                "compeUnitLevel": str(row.get("compeUnitLevel", "")).strip(),
                "ncsSclasCdnm": sclass_nm,
                "ncsSubdCdnm": str(row.get("ncsSubdCdnm", "")).strip(),
                "matchedDetailName": str(row.get("matchedDetailName", "")).strip(),
                "compeUnitDef": str(row.get("compeUnitDef", "")).strip(),
                "score": round(final_score, 6),
                "matched_keywords": hit,
                "similarityScore": round(sim, 6),
                "sourceScore": round(src_norm, 6),
                "__sclass_key": sclass_key,
            }
        )

    scored_rows.sort(
        key=lambda x: (
            float(x.get("score", 0.0) or 0.0),
            float(x.get("similarityScore", 0.0) or 0.0),
            float(x.get("sourceScore", 0.0) or 0.0),
        ),
        reverse=True,
    )

    if per_sclass_limit is None:
        bucket_count = len(preferred_keys) if preferred_keys else len(
            {str(x.get("__sclass_key", "")).strip() for x in scored_rows if str(x.get("__sclass_key", "")).strip()}
        )
        bucket_count = max(1, bucket_count)
        if bucket_count <= 1:
            effective_cap = keep_n
        else:
            effective_cap = max(1, min(4, int(math.ceil(keep_n / bucket_count))))
    else:
        try:
            effective_cap = max(1, int(per_sclass_limit))
        except Exception:
            effective_cap = 2

    selected: list[dict[str, Any]] = []
    seen_selected: set[str] = set()
    per_sclass_count: dict[str, int] = {}

    def _try_add(row: dict[str, Any], enforce_cap: bool = True) -> bool:
        code = str(row.get("ncsClCd", "")).strip()
        if not code or code in seen_selected:
            return False
        s_key = str(row.get("__sclass_key", "")).strip() or "__none__"
        if enforce_cap and per_sclass_count.get(s_key, 0) >= effective_cap:
            return False
        copied = dict(row)
        copied.pop("__sclass_key", None)
        selected.append(copied)
        seen_selected.add(code)
        per_sclass_count[s_key] = per_sclass_count.get(s_key, 0) + 1
        return True

    # Pass 1: high-similarity rows first.
    for row in scored_rows:
        if float(row.get("similarityScore", 0.0) or 0.0) < min_similarity:
            continue
        _try_add(row, enforce_cap=True)
        if len(selected) >= keep_n:
            return selected[:keep_n]

    # Pass 2: fill with remaining rows, still respecting diversity cap.
    for row in scored_rows:
        _try_add(row, enforce_cap=True)
        if len(selected) >= keep_n:
            return selected[:keep_n]

    # Pass 3: if still short, relax diversity cap.
    for row in scored_rows:
        _try_add(row, enforce_cap=False)
        if len(selected) >= keep_n:
            break
    return selected[:keep_n]


def _parse_ai_rerank_codes(text: str) -> list[str]:
    raw = str(text or "").strip()
    if not raw:
        return []

    if "```" in raw:
        m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", raw, flags=re.S)
        if m:
            raw = m.group(1).strip()

    try:
        obj = json.loads(raw)
    except Exception:
        m = re.search(r"\{.*\}", raw, flags=re.S)
        if not m:
            return []
        try:
            obj = json.loads(m.group(0))
        except Exception:
            return []

    if not isinstance(obj, dict):
        return []

    arr = obj.get("ordered_codes") or obj.get("ranked_codes") or obj.get("codes") or []
    if not isinstance(arr, list):
        return []

    out: list[str] = []
    seen: set[str] = set()
    for v in arr:
        code = re.sub(r"[^\d]", "", str(v or "").strip())
        if len(code) < 6 or code in seen:
            continue
        seen.add(code)
        out.append(code)
    return out


def _ai_rerank_ncs_matches(
    jd_text: str,
    ranked_items: list[dict[str, Any]],
    top_k: int = 8,
    api_key_override: str = "",
    generation_provider: str = "openai_api",
) -> list[dict[str, Any]]:
    enabled = os.getenv("ENABLE_AI_RERANK", "true").strip().lower() in {"1", "true", "yes", "y"}
    if not enabled:
        return []

    generation_provider = normalize_generation_provider(generation_provider)
    api_key = (
        settings.resolve_openrouter_key(api_key_override)
        if generation_provider == OPENROUTER_PROVIDER
        else settings.resolve_openai_key(api_key_override)
    )
    if not api_key or len(ranked_items) < 2:
        return []

    if generation_provider == OPENROUTER_PROVIDER:
        net_ok, _ = _check_openai_connectivity(
            api_key=api_key,
            ttl_sec=60,
            provider=generation_provider,
        )
    else:
        # Preserve the long-standing OpenAI call contract so local adapters
        # and existing integrations that accept only api_key/ttl_sec continue
        # to work. OpenRouter still receives the explicit provider boundary.
        net_ok, _ = _check_openai_connectivity(
            api_key=api_key,
            ttl_sec=60,
        )
    if not net_ok:
        return []

    model = provider_model(generation_provider, (
        os.getenv("OPENAI_RERANK_MODEL", "").strip()
        or os.getenv("OPENAI_MODEL", "").strip()
        or "gpt-4o-mini"
    ))

    candidates = []
    for it in ranked_items[:20]:
        candidates.append(
            {
                "ncsClCd": str(it.get("ncsClCd", "")).strip(),
                "compeUnitName": str(it.get("compeUnitName", "")).strip(),
                "ncsSubdCdnm": str(it.get("ncsSubdCdnm", "")).strip(),
                "compeUnitDef": str(it.get("compeUnitDef", "")).strip()[:240],
                "keyword_score": float(it.get("score", 0.0) or 0.0),
            }
        )

    if len(candidates) < 2:
        return []

    payload = {
        "model": model,
        "temperature": 0.0,
        "response_format": {"type": "json_object"},
        "messages": [
            {
                "role": "system",
                "content": (
                    "너는 NCS 매칭 재정렬기다. 반드시 JSON만 출력한다. "
                    "스키마: {\"ordered_codes\":[\"...\"]}"
                ),
            },
            {
                "role": "user",
                "content": (
                    "직무기술서와 후보 NCS를 보고 적합한 순서대로 ncsClCd를 정렬하세요.\n"
                    f"JD:\n{_repair_mojibake(jd_text or '')[:1800]}\n\n"
                    f"Candidates:\n{json.dumps(candidates, ensure_ascii=False)}\n\n"
                    f"반드시 최대 {max(1, int(top_k or 8))}개 코드만 ordered_codes에 넣으세요."
                ),
            },
        ],
    }

    try:
        data = post_chat_completions_with_retries(
            payload=prepare_chat_payload(payload, generation_provider),
            api_key=api_key,
            timeout_sec=provider_timeout_sec(generation_provider, 15.0),
            max_attempts=2,
            provider=generation_provider,
        )
        content = str(((data.get("choices") or [{}])[0].get("message") or {}).get("content", ""))
    except Exception:
        return []

    ordered_codes = _parse_ai_rerank_codes(content)
    if not ordered_codes:
        return []

    def _digits(value: Any) -> str:
        return re.sub(r"[^\d]", "", str(value or "").strip())

    by_code: dict[str, dict[str, Any]] = {}
    for it in ranked_items:
        code = str(it.get("ncsClCd", "")).strip()
        if not code:
            continue
        by_code[code] = it
        d_code = _digits(code)
        if d_code:
            by_code[d_code] = it

    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for code in ordered_codes:
        item = by_code.get(code) or by_code.get(_digits(code))
        raw_code = str((item or {}).get("ncsClCd", "")).strip()
        if not item or not raw_code or raw_code in seen:
            continue
        merged = dict(item)
        merged["rerank_method"] = "ai"
        out.append(merged)
        seen.add(raw_code)
        if len(out) >= top_k:
            break

    for item in ranked_items:
        code = str(item.get("ncsClCd", "")).strip()
        if not code or code in seen:
            continue
        merged = dict(item)
        merged["rerank_method"] = "keyword"
        out.append(merged)
        seen.add(code)
        if len(out) >= top_k:
            break
    return out


def _ocr_image_with_windows_ocr(image_bytes: bytes, lang: str = "ko") -> str:
    """OCR a PNG/JPEG image via built-in Windows OCR (offline, no API key)."""
    if os.name != "nt" or not image_bytes:
        return ""
    try:
        td = _safe_tmp_dir()
        try:
            img_path = os.path.join(td, "in.png")
            ps_path = os.path.join(td, "ocr.ps1")
            with open(img_path, "wb") as f:
                f.write(image_bytes)

            ps_code = (
                "param([string]$ImagePath,[string]$Lang='ko')\n"
                "$ErrorActionPreference='Stop'\n"
                "Add-Type -AssemblyName System.Runtime.WindowsRuntime\n"
                "$null=[Windows.Globalization.Language, Windows, ContentType=WindowsRuntime]\n"
                "$null=[Windows.Media.Ocr.OcrEngine, Windows, ContentType=WindowsRuntime]\n"
                "$null=[Windows.Graphics.Imaging.BitmapDecoder, Windows, ContentType=WindowsRuntime]\n"
                "$null=[Windows.Storage.Streams.InMemoryRandomAccessStream, Windows, ContentType=WindowsRuntime]\n"
                "$null=[Windows.Storage.Streams.DataWriter, Windows, ContentType=WindowsRuntime]\n"
                "$bytes=[System.IO.File]::ReadAllBytes($ImagePath)\n"
                "$stream=New-Object Windows.Storage.Streams.InMemoryRandomAccessStream\n"
                "$writer=New-Object Windows.Storage.Streams.DataWriter($stream)\n"
                "$writer.WriteBytes($bytes)\n"
                "[System.WindowsRuntimeSystemExtensions]::AsTask($writer.StoreAsync()).Result | Out-Null\n"
                "$writer.DetachStream() | Out-Null\n"
                "$writer.Dispose()\n"
                "$stream.Seek(0)\n"
                "$decoder=[System.WindowsRuntimeSystemExtensions]::AsTask([Windows.Graphics.Imaging.BitmapDecoder]::CreateAsync($stream)).Result\n"
                "$bitmap=[System.WindowsRuntimeSystemExtensions]::AsTask($decoder.GetSoftwareBitmapAsync()).Result\n"
                "$engine=$null\n"
                "if ($Lang) { try { $engine=[Windows.Media.Ocr.OcrEngine]::TryCreateFromLanguage((New-Object Windows.Globalization.Language($Lang))) } catch {} }\n"
                "if ($null -eq $engine) { $engine=[Windows.Media.Ocr.OcrEngine]::TryCreateFromUserProfileLanguages() }\n"
                "if ($null -eq $engine) { Write-Output ''; exit 0 }\n"
                "$result=[System.WindowsRuntimeSystemExtensions]::AsTask($engine.RecognizeAsync($bitmap)).Result\n"
                "Write-Output ($result.Text)\n"
            )
            with open(ps_path, "w", encoding="utf-8") as f:
                f.write(ps_code)

            p = subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-File",
                    ps_path,
                    img_path,
                    str(lang or "ko"),
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="ignore",
                timeout=60,
                check=False,
            )
            if p.returncode != 0:
                return ""
            return _repair_mojibake((p.stdout or "").strip())
        finally:
            shutil.rmtree(td, ignore_errors=True)
    except Exception:
        return ""


def _extract_pdf_text_via_windows_ocr(file_bytes: bytes, max_pages: int = 2) -> str:
    images = _render_pdf_pages_png_py313(file_bytes=file_bytes, max_pages=max_pages)
    if not images:
        return ""
    parts: list[str] = []
    for img in images:
        txt = _ocr_image_with_windows_ocr(img, lang=os.getenv("WINDOWS_OCR_LANG", "ko").strip() or "ko")
        if txt:
            parts.append(txt)
    return "\n".join(parts).strip()


def rerank_ncs_matches(
    jd_text: str,
    ncs_items: list[dict[str, Any]],
    top_k: int = 8,
    preferred_sclass: list[str] | None = None,
    openai_api_key: str = "",
    generation_provider: str = "openai_api",
) -> tuple[list[dict[str, Any]], str]:
    rank_pool_k = max(top_k, 12)
    diversity_cap: int | None = None
    if preferred_sclass:
        pref_keys = {
            _sclass_norm_key(x)
            for x in preferred_sclass
            if _sclass_norm_key(str(x or ""))
        }
        if pref_keys:
            diversity_cap = max(1, int(math.ceil(max(1, int(top_k or 1)) / len(pref_keys))))
    try:
        ranked = rank_ncs_matches_by_jd(
            jd_text=jd_text,
            ncs_items=ncs_items,
            top_k=rank_pool_k,
            preferred_sclass=preferred_sclass,
            per_sclass_limit=diversity_cap,
        )
    except TypeError as exc:
        if "unexpected keyword argument" not in str(exc):
            raise
        ranked = rank_ncs_matches_by_jd(
            jd_text=jd_text,
            ncs_items=ncs_items,
            top_k=rank_pool_k,
        )
    if not ranked:
        return [], "keyword"

    ai_ranked = _ai_rerank_ncs_matches(
        jd_text=jd_text,
        ranked_items=ranked,
        top_k=top_k,
        api_key_override=openai_api_key,
        generation_provider=generation_provider,
    )
    if ai_ranked:
        return ai_ranked[:top_k], "ai"

    out: list[dict[str, Any]] = []
    for it in ranked[:top_k]:
        row = dict(it)
        row["rerank_method"] = "keyword"
        out.append(row)
    return out, "keyword"


def _build_rule_based_questions_from_ncs(
    ncs_matches: list[dict[str, Any]] | None,
    ncs_ksa: list[dict[str, Any]] | None = None,
    target_count: int = 24,
) -> list[dict[str, Any]]:
    return []


def build_strategy_with_rule_fallback(
    ncs_matches: list[dict[str, Any]] | None,
    ncs_ksa: list[dict[str, Any]] | None = None,
    error_message: str = "",
    target_count: int = 24,
) -> dict[str, Any]:
    obj: dict[str, Any] = {
        "interview_questions": [],
        "interview_by_competency": [],
        "ncs_link": [
            {
                "ncsClCd": str(x.get("ncsClCd", "")).strip(),
                "compeUnitName": str(x.get("compeUnitName", "")).strip(),
                "why": "NCS 매핑 결과",
            }
            for x in (ncs_matches or [])[:6]
        ],
        "question_generation_policy": "model_only_no_template_fallback",
    }
    if error_message:
        obj["error"] = error_message
    return obj


def _check_openai_connectivity(
    api_key: str,
    ttl_sec: int = 60,
    provider: str = "openai_api",
) -> tuple[bool, str]:
    """Check a request credential without retaining a key-derived cache id."""

    enabled = os.getenv("OPENAI_NET_CHECK_ENABLED", "true").strip().lower() in {"1", "true", "yes", "y"}
    if not enabled:
        return True, "disabled"

    # ``ttl_sec`` remains in the signature for compatibility with existing
    # callers.  A BYOK credential (or a fingerprint derived from it) must not
    # survive the current request in a process-global cache.
    del ttl_sec

    msg = ""
    ok = False
    connect_timeout = 5.0
    read_timeout = 15.0
    write_timeout = 5.0
    pool_timeout = 2.5
    try:
        connect_timeout = max(0.5, float(str(os.getenv("OPENAI_NET_CHECK_CONNECT_TIMEOUT_SEC", "5.0")).strip()))
    except Exception:
        pass
    try:
        read_timeout = max(1.0, float(str(os.getenv("OPENAI_NET_CHECK_READ_TIMEOUT_SEC", "15.0")).strip()))
    except Exception:
        pass
    try:
        write_timeout = max(1.0, float(str(os.getenv("OPENAI_NET_CHECK_WRITE_TIMEOUT_SEC", "5.0")).strip()))
    except Exception:
        pass
    try:
        pool_timeout = max(0.5, float(str(os.getenv("OPENAI_NET_CHECK_POOL_TIMEOUT_SEC", "2.5")).strip()))
    except Exception:
        pass

    try:
        timeout = httpx.Timeout(
            connect=connect_timeout,
            read=read_timeout,
            write=write_timeout,
            pool=pool_timeout,
        )
        ok, msg = check_openai_connectivity_with_retries(
            api_key=api_key,
            timeout=timeout,
            max_attempts=1,
            provider=provider,
        )
    except Exception as e:
        ok = False
        msg = str(e)

    return ok, msg


def _fallback_structured_interview_guide_summary() -> str:
    return (
        "원칙: 주질문1개+꼬리질문3개(사례구체화/어려움대처/결과교훈). "
        "type=경험면접/상황면접/발표면접/토론면접/인바스켓면접/직무지식면접/창의적 문제해결력면접 중 선택. "
        "경험면접은 STAR 행동증거, 상황면접은 판단 기준과 행동 순서, 발표면접은 자료에 근거한 핵심 판단 하나와 그 판단을 기록하는 산출물 하나, "
        "토론면접은 구체적 입장 충돌·근거 검토·경청·조정·합의, 인바스켓면접은 복수 문서의 우선순위와 첫 조치, "
        "직무지식면접은 절차·기준·산출물·예외상황 적용, 창의적 문제해결력면접은 미래예측·실현가능성·의사결정을 검증. 개방형 단일의도."
    )


_PROMPT_INTERVIEW_METHODS = (
    "경험면접",
    "상황면접",
    "발표면접",
    "토론면접",
    "인바스켓면접",
    "직무지식면접",
    "창의적 문제해결력면접",
)

_SELECTED_METHOD_PROMPT_RULES = {
    "경험면접": (
        "- 경험면접(STAR): 실제 사건, 당시 본인 역할, 본인이 택한 선택 또는 직접 행동 1개, "
        "관찰된 결과만 묻습니다. STAR는 지원자가 답을 구성하는 흐름이며 질문에 영문 약자를 "
        "나열하게 하는 암기 검사가 아닙니다.\n"
        "- 경험면접 주질문 하나만 읽어도 S·T·A·R 답변이 나오게 하세요. 반드시 ① 언제·어떤 "
        "직무 사건이었는지 ② 당시 맡은 역할·목표·책임 ③ 배정된 KSA로 직접 판단하거나 수행한 "
        "핵심 행동 한 가지와 그 근거 ④ 수치·문서·피드백 등으로 확인한 결과를 모두 요구하세요. "
        "'관련 경험을 말씀해 주세요' 뒤에 일반적인 행동·결과만 붙이는 문장은 실패입니다.\n"
        "- 지식은 당시 확인한 규정·문서·개념과 적용/제외 근거, 기술은 실제 사용 자료·도구·수행 "
        "순서와 산출물 품질, 태도는 압박·이해 충돌 속에서 고른 행동과 상충효과가 주질문에 "
        "드러나야 합니다. 배정된 업무 대상 없이 일반 협업·문제해결 경험으로 바꾸지 마세요.\n"
        "- S(Situation)는 사건의 시점·맥락·제약, T(Task)는 당시 맡은 역할·목표·책임, "
        "A(Action)는 본인이 실제로 선택하고 수행한 행동과 판단 근거, R(Result)는 관찰 가능한 "
        "결과·증거와 학습 또는 전이를 각각 끌어내야 합니다.\n"
        "- follow_ups가 3개이면 1번은 빠진 S/T 사실, 2번은 답변에서 언급한 A의 선택 이유와 "
        "배정 KSA가 드러난 실제 행동, 3번은 R의 수치·기록·피드백 등 결과 증거와 학습을 "
        "답변 연동형으로 묻습니다. evaluation_points 4개도 S·T·A·R을 서로 겹치지 않게 하나씩 "
        "평가합니다. 새 산출물을 요구하지 않고 기록·증빙은 답변 연동 follow_ups로 확인합니다.\n"
    ),
    "상황면접": (
        "- 상황면접: 오류·불일치·충돌이 있는 구체 자료와 현실적 제약을 제시하고, "
        "첫 판단 1개와 그 판단을 기록하는 최소 산출물 1개를 묻습니다.\n"
    ),
    "발표면접": (
        "- 발표면접은 답변 형식이지 과제 범위를 넓히는 면접이 아닙니다. 표·보고서·민원기록의 "
        "핵심 이상징후를 제시합니다. 분석 기술을 볼 때는 가장 중요한 차이·원인 판정 하나와 "
        "분석표 하나만, 배분·공정성 같은 선택 태도를 볼 때는 배분 결정 하나와 배분안 하나만 "
        "발표하게 합니다. 발표형 산출물은 KSA를 식별하는 최소 필드만 두고 원칙적으로 3개 이하로 "
        "제한합니다. 자원 총량의 검증 숫자가 없으면 상대적 우선순위·범위·원칙만 묻고 정확한 "
        "배분량이나 수치화를 요구하지 않으며, 4개 이상의 세부 항목, 별도 로드맵, 추가 보고서를 "
        "한꺼번에 요구하지 않습니다.\n"
    ),
    "토론면접": (
        "- 토론면접: '[토론과제]'로 시작하고 현장 사건과 서로 양립하기 어려운 두 정책 대안, "
        "검토할 사실, 공동 합의안 또는 미합의 쟁점의 이송 기준을 묻습니다. 합의를 강제하지 않습니다.\n"
    ),
    "인바스켓면접": (
        "- 인바스켓면접: 마감과 권한이 충돌하는 구체 문서·요청을 제시하고 우선순위와 "
        "보고·위임·직접처리 판단 및 첫 조치를 묻습니다.\n"
    ),
    "직무지식면접": (
        "- 직무지식면접: 실제 산출물이나 오류 사례를 제시하고 적용 근거, 예외 처리, "
        "품질 확인 방법을 묻습니다.\n"
    ),
    "창의적 문제해결력면접": (
        "- 창의적 문제해결력면접: 반복 현상과 자원 제약을 제시하고 원인 가설, 작은 검증 "
        "실험, 채택 또는 중단 기준을 묻습니다.\n"
    ),
}


def _selected_prompt_methods(method_names: list[str] | None) -> list[str]:
    selected = [
        method
        for method in _PROMPT_INTERVIEW_METHODS
        if method in {str(value or "").strip() for value in (method_names or [])}
    ]
    return selected or list(_PROMPT_INTERVIEW_METHODS)


def _selected_method_prompt_contract(method_names: list[str] | None) -> str:
    methods = _selected_prompt_methods(method_names)
    return "[선택 면접기법별 규칙]\n" + "".join(
        _SELECTED_METHOD_PROMPT_RULES[method] for method in methods
    )


def _structured_interview_guide_path() -> str:
    return os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "STRUCTURED_INTERVIEW_GUIDE.md"))


def _legacy_model_question_gate_contract() -> str:
    return (
        "[모델 질문 보존 게이트]\n"
        "- 아래 필수어는 주질문(question)에 직접 포함하세요. 빠지면 템플릿으로 교체되어 model-origin 품질 실패로 기록됩니다.\n"
        "- 경험면접 question 필수어: 경험, 상황, 본인, 행동, 결과. evaluation_points에는 구체적 상황 설명 또는 본인 역할과 행동 또는 성과와 학습 포함.\n"
        "- 상황면접 question 필수어: 상황, 판단, 기준, 순서, 위험. evaluation_points에는 판단 기준 또는 위험요인 인식 또는 이해관계자 대응 포함.\n"
        "- 발표면접 question 필수어: 발표, 진단, 대안, 실행, 성과지표, 질의응답. evaluation_points에는 자료 분석력 또는 논리적 구조화 또는 대안의 실행가능성 또는 질의응답 대응 포함.\n"
        "- 토론면접 question 필수어: 토론, 충돌, 입장, 근거, 합의. evaluation_points에는 입장발표 근거 또는 경청과 상호작용 또는 갈등 조정 또는 최종 합의안 도출 포함.\n"
        "- 인바스켓면접 question 필수어: 인바스켓, 문서, 우선순위, 보고, 위임, 직접처리. evaluation_points에는 우선순위 판단 또는 문서·요청 분류 또는 시간관리 포함.\n"
        "- 직무지식면접 question 필수어: 절차, 기준, 산출물, 예외상황. evaluation_points에는 절차·기준 이해 또는 직무지식 적용 또는 산출물 품질 포함.\n"
        "- 창의적 문제해결력면접 question 필수어: 창의적, 미래예측, 문제, 정의, 대안, 검증, 실현가능성, 의사결정, 실행. evaluation_points에는 미래예측과 문제 정의 또는 창의적 사고와 대안 도출 또는 검증 방법 또는 실현가능성 또는 의사결정과 실행계획 포함.\n"
        "- type과 question의 형식이 충돌하면 안 됩니다. 예: type=발표면접이면 question은 반드시 발표과제여야 합니다.\n"
        "- question과 follow_ups에 글자 그대로 'KSA'라고 쓰지 마세요. 반드시 [NCS평가요소]의 factorName 원문 중 하나를 복사해 넣으세요.\n"
        "- factorName은 근거 라벨이지 완성 문장이 아닙니다. 지식은 '활용/근거로', 기술은 '발휘/수행하여', 태도는 '요구된 상황에서 행동으로 보여준' 방식으로 연결하고, '태도를 적용했다'처럼 어색하게 쓰지 마세요.\n"
        "- 금지 질문: '{factorName} 능력과 관련하여 실제 경험이 있으십니까? 말씀해 주세요.'처럼 factorName을 능력명으로 되묻기만 하는 문장. 경험 유무 자체가 아니라 관찰 가능한 판단·행동·산출물을 과제로 요구하세요.\n"
        "- 지식 factorName은 주어진 자료·예외상황에서 무엇을 판단 근거로 활용했고 적용 범위와 오류 위험을 어떻게 설명하는지 측정하세요.\n"
        "- 기술 factorName은 실제 수행 순서, 구체 조치, 사용한 도구·자료, 산출물과 품질 확인을 측정하세요.\n"
        "- 태도 factorName은 마감 압박·이해관계 충돌·품질 위험 같은 선택 상황에서 어떤 행동을 일관되게 택하는지 측정하세요. '태도 경험이 있습니까'나 '태도를 적용했습니까'로 묻지 마세요.\n"
        "- 면접기법이 경험면접이 아니면 과거 경험 유무를 묻지 말고, 해당 기법의 과제 수행 결과로 factorName을 관찰하세요.\n"
        "- 필수어를 체크리스트처럼 나열하지 말고 하나의 구체적인 직무 상황, 판단, 행동과 결과 흐름으로 묶으세요.\n"
        "- 각 질문은 자신에게 배정된 능력단위명(compeUnitName)과 factorName을 함께 반영해야 합니다. 다른 세분류나 다른 능력단위의 직무 맥락을 섞지 마세요.\n"
        "- follow_ups 중 최소 1개에는 질문에 배정된 능력단위명 또는 required_job_context 원문을 직접 포함하세요. 주질문에만 직무명을 넣고 follow_ups를 일반론으로 쓰면 실패입니다.\n"
        "- [질문별 생성 순서]에 required_factorName이 있으면 question과 지정 follow_up slot에 그 문자열을 원문 그대로 포함하세요.\n"
        "- required_factorName을 작성 전 임시 변수 F처럼 그대로 복사한 뒤, question과 지정 follow_up slot에 F를 한 글자도 바꾸지 말고 붙여 넣으세요.\n"
        "- [질문별 생성 순서]에 required_scenario_frame이 있으면 question은 그 상황 프레임을 직접 반영해야 합니다. 같은 면접기법에서 같은 상황 프레임을 반복하지 마세요.\n"
        "- 지정 follow_up slot은 required_followup_focus_slot 값입니다. 값이 없으면 기본 slot은 follow_ups[1]입니다.\n"
        "- 발표·토론·인바스켓·직무지식면접은 지정 slot이 보통 follow_ups[0]입니다. 이 경우 follow_ups[1]이 아니라 follow_ups[0]에 F와 required_job_context를 넣으세요.\n"
        "- 경험·상황·창의적 문제해결력면접은 지정 slot이 보통 follow_ups[1]입니다. 이 경우 follow_ups[1]에 F와 required_job_context를 넣으세요.\n"
        "- [질문별 생성 순서]에 required_followup_focus_example이 있으면 지정 slot의 follow_up은 example과 같은 구조로 쓰고, F와 required_job_context를 모두 포함하세요.\n"
        "- required_followup_focus_example이 있으면 지정 slot 문장을 새로 만들 때 그 예시의 F와 required_job_context 순서를 유지하세요. 직무맥락을 생략하고 '현황 진단', '선택한 대안', '제안한 대안' 같은 일반 표현만 쓰면 실패입니다.\n"
        "- 경험면접 follow_ups[1]에는 반드시 F와 required_job_context를 모두 쓰세요. follow_ups[0]에만 직무명을 쓰고 follow_ups[1]을 '당시 어려움은 무엇입니까?'처럼 쓰면 실패입니다.\n"
        "- 경험면접에서 F를 follow_ups[0]에 쓰고 required_job_context를 follow_ups[1]에 따로 쓰는 것도 실패입니다. F와 required_job_context는 같은 follow_ups[1] 문장 안에 함께 있어야 합니다.\n"
        "- 상황면접 follow_ups[1]에는 반드시 F와 required_job_context를 모두 쓰세요. follow_ups[1]을 '그 판단 기준은 무엇입니까?'처럼 쓰면 실패입니다.\n"
        "- 상황면접은 모델이 자주 F를 빠뜨리므로 follow_ups[0]이나 follow_ups[1] 중 최소 1개는 반드시 F 원문으로 시작하고, 같은 문장에 required_job_context를 함께 넣으세요.\n"
        "- required_factorName이 있으면 의미가 비슷한 대체어, 능력단위명, 세분류명을 factorName 대신 쓰지 마세요. 원문 불일치는 실패입니다.\n"
        "- 토론면접 question은 반드시 '[토론과제]'로 시작하고, F 관련 두 입장이 충돌한다는 구조를 써야 합니다. '입장발표의 근거는 무엇입니까?'를 주질문으로 쓰면 실패입니다.\n"
        "- 인바스켓면접 follow_ups 중 최소 1개는 F와 required_job_context를 포함하면서 문서·요청 우선순위와 보고·위임·직접처리 판단을 함께 물어야 합니다.\n"
        "- follow_ups는 최소 3개이며 서로 중복되면 안 됩니다. 최소 1개는 직무/NCS/KSA 핵심어를 직접 포함해야 합니다.\n"
        "- follow_ups도 기법별로 달라야 합니다: 경험=상황·역할·행동·성과, 상황=확인·기준·위험·후속, 발표=근거자료·대안·반대의견·질의응답·성과지표, 토론=입장발표·반대·조정·합의, 인바스켓=문서분류·우선순위·보고/위임/직접처리, 직무지식=기준·예외·산출물·품질, 창의적 문제해결력=미래예측·문제정의·원인가설·검증·대안·실현가능성·의사결정.\n"
        "- follow_ups[0]은 상황/자료/문서/기준 확인 질문, follow_ups[1]은 판단 이유·행동·우선순위 질문, follow_ups[2]는 결과·후속점검·리스크 보완 질문으로 쓰세요.\n"
        "- 지정 follow_up slot에는 question에 쓴 factorName 원문을 반드시 그대로 포함하세요. '그 판단', '이 절차', '관련 자료', '선택한 대안'처럼 지시어만 쓰면 실패입니다.\n"
        "- follow_ups[0]이나 follow_ups[1]에는 compeUnitName 또는 required_job_context도 함께 넣어 직무 맥락을 유지하세요.\n"
        "- follow_ups 중 하나는 반드시 question에 사용한 factorName 표현을 원문 그대로 반복하세요. 예: '문서 요구사항 파악을 위해 어떤 자료를 확인하고 어떤 행동을 했습니까?'\n"
        "- 실패 예시: 지정 slot follow_up='그 판단 기준은 무엇입니까?' 또는 '선택한 대안의 이유는 무엇입니까?'처럼 F가 없으면 실패입니다.\n"
        "- 통과 예시: 지정 slot follow_up='문서 요구사항 파악을 기준으로 그 판단이나 행동을 선택한 이유는 무엇입니까?'처럼 F 원문이 직접 들어가야 합니다.\n"
        "- 출력 전 자체검사: question에 실제 factorName이 있는가, 지정 follow_up slot에 같은 factorName이 있는가, follow_ups 3개가 서로 다른 평가항목을 묻는가.\n"
        "- 아래 질문 골격을 그대로 따르되, {직무}, {능력단위}, {KSA}는 반드시 실제 NCS/KSA 표현으로 바꾸고 placeholder를 남기지 마세요.\n"
        "- 경험면접 question 골격: {직무}에서 {KSA}가 실제로 필요했던 문제를 해결하거나 성과를 낸 경험을 말씀해 주세요. 당시 상황, 본인 역할, 선택한 행동, 판단 근거, 결과 지표와 학습을 포함해 설명해 주세요.\n"
        "- 상황면접 question 골격: {직무} 중 {KSA}를 실제로 판단·수행해야 하는 구체적 상황입니다. 어떤 판단 기준으로 위험을 통제하고, 사실 확인부터 보고와 실행까지 어떤 순서로 행동하시겠습니까? 지식이면 판단 근거, 기술이면 구체 조치와 산출물, 태도이면 압박 속 선택 행동을 답하도록 요구하세요.\n"
        "- 발표면접 question 골격: [발표과제] {직무}에서 {KSA}를 실제로 적용해야 하는 자료가 주어졌습니다. 현황을 진단하고 대안 2가지, 실행계획과 성과지표를 발표한 뒤 질의응답에 답하게 하세요. 지식이면 판단 근거, 기술이면 구체 조치와 산출물, 태도이면 제약 속 선택 행동을 발표하도록 요구하세요.\n"
        "- 토론면접 question 골격: [토론과제] {직무}에서 {KSA} 관련 두 입장이 충돌합니다. 반대 의견 검토, 조정 방식과 최종 합의안을 토론하게 하세요.\n"
        "- 인바스켓면접 question 골격: [인바스켓과제] {직무} 관련 여러 문서와 요청이 동시에 들어왔습니다. {KSA}를 실제로 발휘해 우선순위, 보고, 위임, 직접처리 판단, 첫 조치와 작성할 산출물을 제시하게 하세요.\n"
        "- 준비·발표·토론·질의응답 시간과 제출 방식은 question 본문에 쓰지 말고 별도 task_conditions에만 둡니다.\n"
        "- 직무지식면접 question 골격: {직무}에서 {KSA}와 관련해 확인해야 할 절차, 기준, 산출물, 예외상황 대응과 품질 점검 방법을 설명해 주세요.\n"
        "- 창의적 문제해결력면접 question 골격: [창의적 문제해결력과제] {직무}에서 {KSA} 관련 복합 문제가 발생했습니다. 미래예측 관점에서 핵심 문제를 정의하고 원인 가설, 창의적 대안 2가지, 검증 방법, 실현가능성, 의사결정 기준, 실행계획과 성과지표를 제시해 주세요.\n"
        "- follow_ups 골격 예시: 1) 먼저 확인할 상황·자료·문서는 무엇입니까? 2) {KSA}를 기준으로 그 판단이나 행동을 선택한 이유는 무엇입니까? 3) 결과 확인, 후속점검, 리스크 보완은 어떻게 하겠습니까?\n"
        "- 면접기법별 지정 slot 예시: 발표 follow_ups[0]={KSA}를 발표 쟁점으로 볼 때 {직무} 현황 진단의 근거자료는 무엇입니까? / 토론 follow_ups[0]={KSA}를 토론 쟁점으로 볼 때 {직무} 입장발표 근거는 무엇입니까? / 인바스켓 follow_ups[0]={KSA}를 처리 기준으로 삼아 {직무} 우선순위를 정한 이유는 무엇입니까? / 직무지식 follow_ups[0]={KSA}와 관련한 기준으로 {직무} 절차를 어떻게 확인하겠습니까? / 창의적 문제해결력 follow_ups[1]={KSA}와 관련한 원인과 대안 관점에서 {직무} 문제를 어떻게 검증하겠습니까?\n"
    )


def _model_question_gate_contract(method_names: list[str] | None = None) -> str:
    """Contract for evidence-grounded questions without exposing raw KSA labels."""

    selected_methods = set(_selected_prompt_methods(method_names))
    selected_method_rules = _selected_method_prompt_contract(method_names)
    debate_follow_up_rule = (
        "- 토론면접 꼬리질문은 ① 확인 자료·사실 ② 수용·불수용 경계와 기준 "
        "③ 합의안의 적용 범위·예외·검증 기준과 실행 책임·후속점검을 각각 검증하세요.\n"
        if "토론면접" in selected_methods
        else ""
    )
    selected_examples = ""
    if "경험면접" in selected_methods:
        selected_examples += (
            "- 좋은 경험형 설계: 각 index의 required_scenario_frame을 실제 사건으로 사용하고, "
            "required_task_statement에 해당하는 본인 행동이 사건의 핵심이 되게 하세요. 상황·역할·행동·결과를 "
            "묻되 다른 index의 사건이나 일반 협업 경험으로 바꾸지 마세요.\n"
        )
    if "상황면접" in selected_methods:
        selected_examples += (
            "- 좋은 상황형: '연구협약서 초안의 정산 조항과 내부 지침이 서로 다르고 협약 마감이 오늘입니다. "
            "어떤 문서를 먼저 대조하고 누구에게 쟁점을 확인한 뒤 수정안을 확정하겠습니까?'\n"
        )
    if "발표면접" in selected_methods:
        selected_examples += (
            "- 나쁜 발표형: '원인을 진단하고 두 대안을 비교해 우선안을 선택한 뒤 세부 실행계획과 성과지표를 모두 발표하세요.' "
            "진단·비교·선택·로드맵·검증을 한 번에 묻습니다.\n"
            "- 좋은 분석 발표형: '목표와 실적이 엇갈린 자료에서 가장 중요한 차이 한 건을 판정하고, "
            "목표값·실적값·차이 근거가 보이는 분석표 한 장을 발표하세요.' 대안과 후속 조치는 꼬리질문에서 묻습니다.\n"
            "- 좋은 배분 발표형: '총량 수치가 제공되지 않은 동결 조건에서 어느 사업을 상대적으로 우선할지 정하고, "
            "공통 원칙과 조정 가능한 범위가 보이는 배분안 한 장을 발표하세요.' 정확한 배분량·수치화와 반발 조정은 요구하지 않습니다.\n"
        )
    return (
        "[모델 질문 보존 게이트 v2]\n"
        "- 이 게이트는 앞의 일반 면접 가이드보다 우선합니다. 가이드에 나열된 진단·대안·실행·검증 요소는 한 주질문에 모두 넣을 필수 목록이 아니라 서로 다른 문항 또는 follow_ups에 배치할 선택지입니다.\n"
        "- required_factorName, required_task_statement, required_observable_behavior, required_job_context, compeUnitName, detail은 내부 의미 힌트입니다. 공식 KSA·NCS 라벨을 question, follow_ups, evaluation_points에 원문으로 복사하거나 조사만 붙여 쓰지 마세요.\n"
        "- required_surface_focus는 공식 KSA 라벨이 아니라 지원자에게 보여도 되는 업무 대상·행동 힌트입니다. 그 의미가 주질문에서 식별되게 실제 문서·자료·행동으로 한 번만 자연스럽게 풀어 쓰고, '절차·절차'나 '관련 실무 적용·검증 절차' 같은 기계적 suffix를 반복하지 마세요.\n"
        "- question_focus와 ksa_refs에는 required_factorName을, question_focus_surface에는 required_surface_focus를 정확히 보존하세요. 지원자 문장에는 required_factorName과 evidence_id를 노출하지 마세요.\n"
        "- 내부 추적 필드의 원문은 question, follow_ups, evaluation_points에는 노출하지 마세요.\n"
        "- question_evidence_id에는 배정된 evidence_id를 정확히 저장하세요. evidence_id는 문장에 노출하지 않고, 내부 근거와 생성 문항의 추적 연결을 바꾸지 마세요.\n"
        "- 내부 힌트의 뜻을 실제 사건·문서·데이터·이해관계자·제약·판단·관찰 가능한 결과로 자유롭게 번역하세요. 직무명이나 추상 라벨로 문장을 시작하지 마세요.\n"
        "- 문항을 쓰기 전에 반사실 검사를 하세요: 배정된 KSA가 없는 유능한 일반 행정 담당자도 같은 답을 할 수 있다면 사건을 다시 설계하세요. 단순 우선순위·자료 확인·협업만으로 답할 수 있는 문항은 해당 KSA 측정 문항이 아닙니다.\n"
        "- 지식 KSA는 그 지식만의 정의·적용 근거·범위·예외 중 하나를 실제 판단에 사용하게 하세요. 법·규정 지식은 목적 제한·적법 근거·최소 범위처럼 구별되는 적용 논리를, 지표 지식은 포함 범위·중복 처리·측정 기간처럼 지표 정의를 답하게 해야 하며 일반적인 검토 순서로 대체하지 마세요.\n"
        "- 기술 KSA는 그 기술만의 변환·대조·작성·협상 절차와 도메인 산출물을 직접 만들거나 설명하게 하세요. 일반 우선순위표가 아니라 보고서 구조·계획 대비 차이 분석·서식 보완·등록 필드와 변경 이력 등 배정 기술을 식별할 수 있는 흔적이 있어야 합니다.\n"
        f"{_neutral_attitude_prompt_contract()}"
        f"{_editorial_realism_prompt_contract()}"
        "- 필수 산출물의 이름만 그럴듯하게 붙이지 마세요. 산출물에 요구한 필드·구조·판정 규칙을 보면 어떤 KSA를 측정하는지 구별되어야 합니다. 보고서 작성 요령 지식이라면 본문·주석·잠정값·증빙 처리 같은 보고서 구성 규칙을 적용하게 해야 하며, 문서 처리 순서나 담당자 배정표만으로 대체하지 마세요.\n"
        "- required_factor가 회계·분석·검토보고서 작성 요령이면 과제형 question 자체의 판단과 산출물에 ① 확정값/잠정값 구분 ② 본문/주석 배치 ③ 증빙 연결 중 최소 2개를 직접 적용하게 하세요. 단, 경험면접은 당시 실제 보고서 작성 행동과 관찰된 결과만 주질문에서 묻고 실제 구성·증빙 근거는 답변 연동 follow_ups로 확인합니다.\n"
        "- 보고서 작성 요령의 두 필드는 같은 보고서 한 장의 구성요소이며 별도 산출물을 추가하라는 뜻이 아닙니다.\n"
        "- 출력 전 KSA 대체 검사를 하세요: 내부 factor를 다른 KSA로 바꿔도 질문이 그대로 성립하면 실패입니다. raw factorName을 노출하지 않은 채 사건의 판단 근거·행동·산출물을 해당 KSA에만 맞게 고치세요.\n"
        "- 출력 직전 question만 단독으로 다시 읽으세요. follow_ups와 evaluation_points를 지웠을 때 required KSA 없이도 답할 수 있다면, 꼬리질문으로 보완하지 말고 question의 핵심 판단·산출물을 다시 작성하세요.\n"
        "- 경험면접을 제외한 과제형 주질문에는 핵심 판단 1개와 답변자가 반드시 제시할 최소 산출물 1개만 남기세요. 산출물은 반드시 그 핵심 판단을 기록해야 합니다. 원인 진단, 복수 대안 비교, 우선안 선택, 이해관계 조정, 실행 로드맵, 성과 검증은 서로 다른 판단 family이므로 둘 이상을 주질문에 직렬로 결합하지 말고 나머지는 서로 다른 follow_ups로 이동하세요.\n"
        f"{selected_method_rules}"
        "- 위 자산은 면접기법별 선택지입니다. 모든 문항에 자료·이상·제약·결정·산출물을 동일 체크리스트처럼 전부 나열하지 마세요.\n"
        "- 상황문에 정답 정책을 먼저 알려 주지 마세요. 예를 들어 '최소 정보만 제공하고 근거가 없으면 보류한다는 원칙에 따라 처리하라'처럼 올바른 선택을 완성해 제시하지 말고, 목적·권한·피해 위험이 충돌하는 사실을 주어 지원자가 적용 원칙과 처리 경계를 스스로 설명하게 하세요.\n"
        f"{_unverified_material_precision_prompt_contract()}"
        "- 발표·토론·인바스켓의 시간과 제출요건은 별도 task_conditions로 제공되므로 주질문에 체크리스트처럼 반복하지 마세요.\n"
        f"{debate_follow_up_rule}"
        "- follow_ups가 3개이면 최소 2개는 지원자의 직전 답변에서 언급한 내용, 빠뜨린 근거, 선택한 행동, 보고한 결과 중 하나를 명시적으로 받아 묻는 답변 연동 질문이어야 합니다. 나머지 1개만 모든 지원자에게 동일한 표준화 질문으로 둘 수 있습니다.\n"
        "- 답변 연동 질문은 '방금 말씀하신 선택', '앞서 언급한 결과', '답변에 수치 근거가 없다면'처럼 참조 대상이나 조건을 드러내세요. 단순히 '필요한 경우' 또는 '면접관 판단에 따라'라고만 쓰지 마세요.\n"
        "- evaluation_points는 정확히 4개 작성하세요. 네 항목 모두 question 또는 follow_ups가 실제로 답을 끌어내고 직접 관찰할 수 있는 서로 다른 핵심 근거·판단·행동·산출물이어야 하며, 질문하지 않은 숨은 기준이나 성향 라벨을 넣지 마세요.\n"
        "[대조 예시]\n"
        "- 나쁜 예: '시장환경 분석·판단 기준에 따라 사업계획을 수립한 경험을 말씀해 주세요.' 추상 surface 라벨을 질문 골격에 붙였을 뿐 사건이 없습니다.\n"
        "- 나쁜 예: '문서 요구사항 확인 절차에 따라 어떻게 처리하시겠습니까?' 실제 문서, 오류, 이해관계자, 제약이 없습니다.\n"
        f"{selected_examples}"
        "- 출력 전 자체검사: 공식 factorName·surface·능력단위명이 노출되지 않았는가, 직무 사건이 구체적인가, exact evidence_id가 유지됐는가, 꼬리질문 3개 중 2개 이상이 답변에 연동되는가, 다른 KSA로 바꿔도 그대로 답할 수 없는가.\n"
    )


def _pick_planned_unit_for_prompt(
    target_detail: str,
    offset: int,
    ncs_matches: list[dict[str, Any]] | None,
) -> dict[str, Any]:
    rows = [x for x in (ncs_matches or []) if isinstance(x, dict)]
    if not rows:
        return {}
    detail_key = _norm_text(target_detail)
    exact: list[dict[str, Any]] = []
    fallback: list[dict[str, Any]] = []
    if detail_key:
        for row in rows:
            authoritative_detail_keys = {
                _norm_text(str(row.get("matchedDetailName", ""))),
                _norm_text(str(row.get("reviewed_detail", ""))),
                _norm_text(str(row.get("confirmed_detail", ""))),
                _norm_text(str(row.get("ncs_detail", ""))),
                _norm_text(str(row.get("ncsSubdCdnm", ""))),
            }
            authoritative_detail_keys.discard("")
            if detail_key in authoritative_detail_keys:
                exact.append(row)
                continue
            sclass = _norm_text(str(row.get("ncsSclasCdnm", "")))
            matched = [
                _norm_text(str(x))
                for x in (row.get("matched_keywords") or [])
                if str(x).strip()
            ] if isinstance(row.get("matched_keywords"), list) else []
            if sclass == detail_key or detail_key in matched:
                fallback.append(row)
    pool = exact or fallback or rows
    return dict(pool[offset % len(pool)])


def _planned_factor_for_prompt(
    ncs_code: str,
    question_index: int,
    ncs_ksa: list[dict[str, Any]] | None,
) -> str:
    code = str(ncs_code or "").strip()
    if not code:
        return ""
    factors: list[str] = []
    seen: set[str] = set()
    for row in ncs_ksa or []:
        if not isinstance(row, dict) or str(row.get("ncsClCd", "")).strip() != code:
            continue
        factor = str(row.get("factorName") or "").strip()
        key = _norm_text(factor)
        if factor and key and key not in seen:
            seen.add(key)
            factors.append(factor[:120])
    if not factors:
        return ""
    return factors[(max(1, int(question_index or 1)) - 1) % len(factors)]


def _planned_factor_row_for_prompt(
    ncs_code: str,
    factor_name: str,
    ncs_ksa: list[dict[str, Any]] | None,
) -> dict[str, Any]:
    code = str(ncs_code or "").strip()
    factor_key = _norm_text(factor_name)
    for row in ncs_ksa or []:
        if not isinstance(row, dict) or str(row.get("ncsClCd") or "").strip() != code:
            continue
        if factor_key and _norm_text(str(row.get("factorName") or "")) == factor_key:
            return dict(row)
    return {}


def _planned_followup_focus_slot_for_prompt(method: str) -> int:
    return {
        "경험면접": 1,
        "상황면접": 1,
        "발표면접": 0,
        "토론면접": 0,
        "인바스켓면접": 0,
        "직무지식면접": 0,
        "창의적 문제해결력면접": 1,
    }.get(str(method or "").strip(), 1)


def _canonical_interview_method_for_prompt(method: str) -> str:
    value = str(method or "").strip()
    return {
        "experience": "경험면접",
        "situation": "상황면접",
        "presentation": "발표면접",
        "discussion": "토론면접",
        "inbasket": "인바스켓면접",
        "job_knowledge": "직무지식면접",
        "creative_problem_solving": "창의적 문제해결력면접",
    }.get(value, value)


def _planned_followup_focus_example_for_prompt(method: str, job_context: str, factor_name: str) -> str:
    """Return an answer-linked design brief, never a label-filled final sentence."""

    method = str(method or "").strip()
    # These values remain in the surrounding plan as trace metadata.  They are
    # intentionally not interpolated into a visible-sentence example because
    # doing so taught the model to copy NCS/surface labels verbatim.
    _ = job_context, factor_name
    method = _canonical_interview_method_for_prompt(method)
    briefs = {
        "경험면접": "답변 연동 설계: 꼬리1은 답변에서 빠진 상황·역할을, 꼬리2는 배정 과업에 해당하는 본인의 선택·행동과 근거를, 꼬리3은 관찰 가능한 결과 증거와 학습을 파고든다.",
        "상황면접": "답변 연동 설계: 지원자가 첫 조치로 고른 행동을 받아, 빠뜨린 위험이나 정보가 있을 때 판단을 어떻게 수정할지 묻는다.",
        "발표면접": "답변 연동 설계: 발표에서 근거로 든 수치나 선택한 대안을 받아 출처·반대 자료·성과 확인 방식을 묻는다.",
        "토론면접": "답변 연동 설계: 지원자가 수용한 상대 입장과 남겨 둔 예외를 받아 합의의 경계·책임·검증 방식을 묻는다.",
        "인바스켓면접": "답변 연동 설계: 지원자가 1순위로 둔 문서와 처리 주체를 받아, 보고·위임·직접처리 선택의 근거와 누락 위험을 묻는다.",
        "직무지식면접": "답변 연동 설계: 지원자가 적용한다고 한 근거나 절차를 받아 예외 사례와 산출물 오류 확인 방식을 묻는다.",
        "창의적 문제해결력면접": "답변 연동 설계: 지원자가 제시한 원인 가설이나 대안을 받아 반증 자료·중단 기준·관찰 결과를 묻는다.",
    }
    return briefs.get(
        method,
        "답변 연동 설계: 지원자가 방금 말한 선택이나 결과를 하나 집어 근거, 누락, 후속 확인을 묻는다.",
    )


def _planned_scenario_frame_for_prompt(method: str, offset: int) -> str:
    method = str(method or "").strip()
    frames_by_method = {
        "경험면접": (
            "일정 지연 또는 마감 압박 속에서 본인이 선택한 행동",
            "자료 불일치나 정보 부족을 확인하고 보완한 경험",
            "이해관계자 요청이 충돌한 상황에서 조정한 경험",
            "예외상황이나 오류를 발견하고 재발을 막은 경험",
            "제한된 자원 안에서 우선순위를 조정한 경험",
        ),
        "상황면접": (
            "동시에 들어온 요청과 마감 충돌",
            "자료 오류 또는 기준 불일치 발견",
            "민원, 안전, 품질 리스크가 함께 있는 상황",
            "상급자 지시와 현장 제약이 충돌하는 상황",
            "협업 부서와 처리 기준이 다른 상황",
        ),
        "발표면접": (
            "현황 자료와 오류 사례를 바탕으로 원인과 대안을 제시",
            "운영 지표 변화와 민원 기록을 분석해 개선안을 제시",
            "절차 준수와 처리 속도 사이의 개선 우선순위를 제시",
            "품질 점검 결과를 바탕으로 실행계획과 성과지표를 제시",
            "자원 제약 속 단계별 개선 로드맵을 제시",
        ),
        "토론면접": (
            "기준 강화 입장과 처리 효율 우선 입장의 충돌",
            "안전·품질 우선 입장과 일정 준수 입장의 충돌",
            "이용자 편의 우선 입장과 절차 준수 입장의 충돌",
            "정보 공유 확대 입장과 보안·책임성 강화 입장의 충돌",
            "단기 성과 우선 입장과 재발 방지 우선 입장의 충돌",
        ),
        "인바스켓면접": (
            "긴급 요청, 오류 정정, 보고 문서가 동시에 들어온 상황",
            "민원 확대 가능성과 마감 임박 업무가 겹친 상황",
            "상급자 보고, 협업 요청, 현장 확인이 동시에 필요한 상황",
            "안전·품질 이슈와 처리 속도 요구가 동시에 있는 상황",
            "자료 확인, 위임, 직접처리 판단이 동시에 필요한 상황",
        ),
        "직무지식면접": (
            "기준 적용 절차와 예외상황 처리",
            "산출물 품질 점검과 오류 예방",
            "관련 근거 확인과 기록 관리",
            "업무 도구 또는 자료를 활용한 검증",
            "신규 담당자에게 설명할 적용 순서",
        ),
        "창의적 문제해결력면접": (
            "반복 오류의 원인 가설과 검증 방법",
            "환경 변화에 따른 미래 리스크와 대안",
            "제한된 자원에서 실행 가능한 개선안 비교",
            "이해관계자 요구가 충돌하는 복합 문제",
            "성과지표와 후속 점검까지 포함한 개선 실험",
        ),
    }
    default_frames = (
        "일정 지연과 품질 요구가 함께 있는 상황",
        "자료 오류 또는 기준 불일치가 발견된 상황",
        "이해관계자 요구가 충돌하는 상황",
        "제한된 자원에서 우선순위를 정해야 하는 상황",
        "예외상황을 처리하고 재발을 막아야 하는 상황",
    )
    frames = frames_by_method.get(method, default_frames)
    if not frames:
        return ""
    return frames[max(0, int(offset or 0)) % len(frames)]


def _planned_ksa_scenario_frame_for_prompt(
    method: str,
    offset: int,
    *,
    job_context: str,
    task_frame: dict[str, str],
) -> str:
    """Anchor a scenario to the assigned KSA instead of an unrelated stock event."""

    method = _canonical_interview_method_for_prompt(method)
    task_statement = str(task_frame.get("task_statement") or "").strip()
    observable_behavior = str(task_frame.get("observable_behavior") or "").strip()
    context = str(job_context or "").strip() or "해당 업무"
    if not task_statement:
        return _planned_scenario_frame_for_prompt(method, offset)

    variants = (
        "마감 또는 정보 부족이라는 제약이 있었던 경우",
        "적용 기준이 애매하거나 예외를 판단해야 했던 경우",
        "기존 방식과 다른 요청의 타당성을 검토해야 했던 경우",
        "오류나 누락을 발견해 처음 결정을 수정해야 했던 경우",
        "결과를 확인하고 다음 업무에 반영해야 했던 경우",
    )
    variant = variants[max(0, int(offset or 0)) % len(variants)]
    core = (
        f"직무 맥락: {context}. 배정 KSA의 핵심 과업: {task_statement}. "
        f"사건 조건: {variant}. 관찰해야 할 행동·근거: {observable_behavior}."
    )
    if method == "경험면접":
        return (
            "실제 경험 사건으로 설계. "
            f"{core} 당시 상황·제약, 본인의 역할·목표, 본인이 직접 선택한 행동 하나와 판단 근거, "
            "수치·기록·피드백 등 관찰 가능한 결과 증거가 모두 답변에서 나오게 할 것."
        )
    return f"선택 면접기법의 과제로 설계. {core}"


def _planned_question_example_for_prompt(method: str, job_context: str, factor_name: str) -> str:
    """Return method-specific situation assets without copying prompt labels."""

    method = str(method or "").strip()
    _ = job_context, factor_name
    method = _canonical_interview_method_for_prompt(method)
    briefs = {
        "경험면접": "설계 자산: required_scenario_frame의 실제 사건 + 당시 역할·목표 + required_task_statement에 해당하는 본인의 판단·행동 하나 + 행동 전후의 관찰 가능한 결과 증거.",
        "상황면접": "설계 자산: 금액이나 조항이 서로 다른 두 문서 + 당일 마감 + 확인 가능한 담당자 + 첫 판단과 수정안.",
        "발표면접": "설계 자산: 월별 지표표와 민원 기록 + 급변한 수치 하나 + 제한된 예산 + KSA에 가장 가까운 핵심 판단 하나 + 그 판단을 기록하는 산출물 하나.",
        "토론면접": "설계 자산: 실제 운영 사건 + 속도 우선과 검증 우선의 양립하기 어려운 두 입장 + 확인할 사실 + 합의 적용 범위.",
        "인바스켓면접": "설계 자산: 오류 정정 요청·결재 문서·민원 회신 + 서로 다른 마감 + 제한된 결재 권한 + 처리 주체 결정.",
        "직무지식면접": "설계 자산: 오류가 있는 실제 산출물 한 종류 + 적용 근거 + 예외 사례 + 수정 완료를 확인할 기록.",
        "창의적 문제해결력면접": "설계 자산: 반복되는 이상 현상 + 상충하는 원인 자료 + 제한된 인력·예산 + 작은 검증 실험과 중단 기준.",
    }
    return briefs.get(method, "설계 자산: 구체 사건·대상·제약 중 필요한 것 + 지원자의 판단 + 확인 가능한 결과.")


_PLANNED_DIFFICULTY_AXES = (
    "기본: 핵심 근거 하나를 정확히 적용",
    "심화: 상충하는 근거·목표 두 개를 비교",
    "고난도: 불완전 정보와 권한·시간 제약 아래 예외를 판단",
)
_PLANNED_QUESTION_ANGLES = (
    "적용 근거와 범위",
    "직접 수행 행동과 도메인 산출물",
    "오류·예외 탐지와 처리 경계",
    "결과 검증과 수정 조건",
)
_PLANNED_CONSTRAINT_AXES = (
    "자료 또는 수치 불일치",
    "마감 또는 선후 의존성",
    "이해관계자 목표 충돌",
    "승인 권한 또는 규정 예외",
    "인력·예산 또는 품질 위험",
)


def _planned_question_sequence_for_prompt(
    question_plan: dict[str, Any] | None,
    method_names: list[str],
    target_count: int,
    ncs_matches: list[dict[str, Any]] | None = None,
    ncs_ksa: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    if not isinstance(question_plan, dict):
        return []
    raw_sequence = [x for x in (question_plan.get("question_sequence") or []) if isinstance(x, dict)]
    if not raw_sequence:
        return []
    methods = [str(x).strip() for x in (method_names or []) if str(x).strip()]
    limit = max(1, min(50, int(target_count or len(raw_sequence))))
    planned: list[dict[str, Any]] = []
    detail_offsets: dict[str, int] = {}
    factor_offsets_by_code: dict[str, int] = {}
    scenario_offsets_by_method: dict[str, int] = {}
    for idx, row in enumerate(raw_sequence[:limit], start=1):
        detail = str(row.get("detail") or "").strip()
        if not detail:
            continue
        method = methods[(idx - 1) % len(methods)] if methods else ""
        locked_code = str(row.get("ncsClCd") or "").strip()
        locked_evidence_id = str(row.get("evidence_id") or "").strip()
        locked_factor_row = next(
            (
                dict(candidate)
                for candidate in (ncs_ksa or [])
                if isinstance(candidate, dict)
                and str(candidate.get("ncsClCd") or "").strip() == locked_code
                and stable_ksa_evidence_id(candidate) == locked_evidence_id
            ),
            {},
        )
        locked_unit = next(
            (
                dict(candidate)
                for candidate in (ncs_matches or [])
                if isinstance(candidate, dict)
                and str(candidate.get("ncsClCd") or "").strip() == locked_code
            ),
            {},
        )
        locked_method = _canonical_interview_method_for_prompt(row.get("type") or "")
        selected_canonical_methods = {
            _canonical_interview_method_for_prompt(value) for value in methods
        }
        if locked_method and locked_method in selected_canonical_methods:
            method = locked_method
        detail_key = _norm_text(detail)
        offset = detail_offsets.get(detail_key, 0)
        unit = locked_unit or _pick_planned_unit_for_prompt(detail, offset, ncs_matches)
        detail_offsets[detail_key] = offset + 1
        planned_item = {
            "index": idx,
            "detail": detail,
            "type": method,
            "follow_up_count": max(0, min(5, int(row.get("follow_up_count", 3) or 0))),
            "required_difficulty": _PLANNED_DIFFICULTY_AXES[
                (idx - 1) % len(_PLANNED_DIFFICULTY_AXES)
            ],
            "required_question_angle": _PLANNED_QUESTION_ANGLES[
                (idx - 1) % len(_PLANNED_QUESTION_ANGLES)
            ],
            "required_constraint_axis": _PLANNED_CONSTRAINT_AXES[
                (idx - 1) % len(_PLANNED_CONSTRAINT_AXES)
            ],
        }
        scenario_offset = scenario_offsets_by_method.get(method, 0)
        scenario_frame = _planned_scenario_frame_for_prompt(method, scenario_offset)
        if method:
            scenario_offsets_by_method[method] = scenario_offset + 1
        if unit:
            ncs_code = str(unit.get("ncsClCd", "")).strip()
            compe_unit_name = str(unit.get("compeUnitName", "")).strip()
            ncs_sub_detail = str(unit.get("ncsSubdCdnm", "")).strip()
            factor_offset = factor_offsets_by_code.get(ncs_code, 0)
            required_factor = str(locked_factor_row.get("factorName") or "").strip()
            if not required_factor:
                required_factor = _planned_factor_for_prompt(ncs_code, factor_offset + 1, ncs_ksa)
            if ncs_code:
                factor_offsets_by_code[ncs_code] = factor_offset + 1
            required_context = compe_unit_name or ncs_sub_detail or detail
            factor_row = locked_factor_row or _planned_factor_row_for_prompt(
                ncs_code,
                required_factor,
                ncs_ksa,
            )
            task_frame = build_question_task_frame(
                evidence_row=factor_row or None,
                factor_name=required_factor,
                ksa_type=(
                    factor_row.get("ksaTypeName")
                    or factor_row.get("factorType")
                    or factor_row.get("ksa_type")
                    or ""
                ),
                element_name=factor_row.get("elementName") or factor_row.get("element_name") or "",
                competency_name=compe_unit_name,
                competency_definition=str(unit.get("compeUnitDef", "")).strip(),
                decision_dilemma=scenario_frame,
            )
            scenario_frame = _planned_ksa_scenario_frame_for_prompt(
                method,
                scenario_offset,
                job_context=required_context,
                task_frame=task_frame,
            )
            surface_focus = task_frame["task_object"]
            planned_item.update(
                {
                    "ncsClCd": ncs_code,
                    "compeUnitName": compe_unit_name,
                    "compeUnitDef": str(unit.get("compeUnitDef", "")).strip()[:240],
                    "ncsSubdCdnm": ncs_sub_detail,
                    "required_job_context": required_context,
                    "evidence_id": task_frame.get("evidence_id", ""),
                    "required_element_name": str(
                        factor_row.get("elementName")
                        or factor_row.get("element_name")
                        or ""
                    ).strip(),
                    "required_factorName": required_factor,
                    "required_ksa_type": task_frame["ksa_type"],
                    "required_surface_focus": surface_focus,
                    "required_task_statement": task_frame["task_statement"],
                    "required_observable_behavior": task_frame["observable_behavior"],
                    "required_scenario_frame": scenario_frame,
                    "required_question_example": _planned_question_example_for_prompt(
                        method,
                        required_context,
                        surface_focus,
                    ),
                    "required_followup_focus_slot": _planned_followup_focus_slot_for_prompt(method),
                    "required_followup_focus_example": _planned_followup_focus_example_for_prompt(
                        method,
                        required_context,
                        surface_focus,
                    ),
                }
            )
        elif scenario_frame:
            planned_item.update(
                {
                    "required_job_context": detail,
                    "required_scenario_frame": scenario_frame,
                    "required_followup_focus_slot": _planned_followup_focus_slot_for_prompt(method),
                }
            )
        planned.append(planned_item)
    return planned


def _load_structured_interview_guide_summary(max_chars: int = 1400) -> str:
    try:
        with open(_structured_interview_guide_path(), "r", encoding="utf-8") as f:
            guide_full = f.read()
        match = re.search(r"(## 3\. 질문 유형별 작성 기법.*?)(?=\n## [4-9]\.|\Z)", guide_full, re.DOTALL)
        guide_summary = (match.group(1) if match else guide_full).strip()
        if guide_summary:
            return guide_summary[: max(200, int(max_chars))]
    except Exception:
        pass
    return _fallback_structured_interview_guide_summary()


def _experience_only_generation_prompt(
    *,
    planned_sequence: list[dict[str, Any]],
    target_count: int,
    follow_up_count: int,
    notice_text: str,
    jd_text: str,
    duty_text: str,
    evaluation_text: str,
    extra_context: str,
) -> str:
    """Build a compact, non-conflicting STAR prompt for experience interviews."""

    slots: list[dict[str, Any]] = []
    for fallback_index, raw in enumerate(planned_sequence[:target_count], start=1):
        if not isinstance(raw, dict):
            continue
        slots.append(
            {
                "index": int(raw.get("index") or fallback_index),
                "type": "경험면접",
                "detail": str(raw.get("detail") or "").strip(),
                "ncsClCd": str(raw.get("ncsClCd") or "").strip(),
                "competency": str(raw.get("compeUnitName") or "").strip(),
                "competency_definition": str(raw.get("compeUnitDef") or "").strip(),
                "work_element": str(raw.get("required_element_name") or "").strip(),
                "evidence_id": str(raw.get("evidence_id") or "").strip(),
                "ksa_type": str(raw.get("required_ksa_type") or "").strip(),
                "task_semantics": str(raw.get("required_task_statement") or "").strip(),
                "observable_evidence": str(
                    raw.get("required_observable_behavior") or ""
                ).strip(),
            }
        )

    context = {
        "notice": str(notice_text or "")[:900],
        "jd": str(jd_text or "")[:900],
        "duties": str(duty_text or "")[:1200],
        "evaluation": str(evaluation_text or "")[:700],
    }
    retry_context = str(extra_context or "")[:1400]
    return (
        "JSON만 출력하세요. 공공기관 NCS 기반 경험(행동)면접 질문을 작성합니다.\n"
        f"interview_questions를 정확히 {target_count}개, 입력 slot 순서대로 작성하세요. "
        f"각 문항은 follow_ups를 정확히 {follow_up_count}개, evaluation_points를 정확히 4개 가집니다.\n"
        "[핵심 원칙]\n"
        "- 주질문 하나만 읽어도 지원자가 STAR로 답해야 합니다. 구체적인 과거 직무 사건(S)을 먼저 "
        "묻고, 당시 역할·목표(T), 직접 선택·수행한 핵심 행동과 판단 근거(A), 수치·문서·피드백으로 "
        "확인한 결과(R)를 자연스럽게 연결하세요. 주질문은 두 문장, 약 170자 이내로 쓰고 STAR를 "
        "체크리스트처럼 길게 나열하지 마세요. 부족한 세부 증거는 답변 연동 꼬리질문으로 확인합니다.\n"
        "- '관련 경험을 말씀해 주세요'처럼 일반 경험만 묻지 마세요. slot의 work_element, "
        "competency_definition, task_semantics를 이용해 실제 문서·자료·도구·이해관계자·산출물 중 "
        "최소 두 가지를 주질문에 넣으세요.\n"
        "- ksa_type=지식이면 당시 확인한 규정·문서·개념, 적용하거나 제외한 범위와 그 판단 근거가 "
        "행동을 어떻게 바꿨는지 물으세요. ksa_type=기술이면 사용한 자료·도구, 실제 수행 순서·조치, "
        "만든 산출물과 품질 확인 결과를 물으세요. ksa_type=태도이면 마감 압박·정확성·이해 충돌 중 "
        "구체 제약을 제시하고 본인이 고른 행동과 상충효과, 확인 결과를 물으세요.\n"
        "- task_semantics와 observable_evidence는 질문 설계용 의미입니다. 핵심 대상·행동을 자연스러운 "
        "직무 사건으로 바꾸고, '관련 실무 적용·검증 절차', '행동 기준', '절차·절차' 같은 메타 문구를 "
        "만들지 마세요. evidence_id, NCS 코드·능력단위명은 질문과 꼬리질문에 노출하지 마세요.\n"
        "- 꼬리1은 '방금 말씀하신'으로 시작해 답변에서 빠진 S/T의 사건 조건·본인 역할·목표를, "
        "꼬리2는 '앞서 언급한'으로 시작해 A의 선택 이유·실제 행동·사용 근거를, 꼬리3은 결과가 "
        "없거나 불명확한 경우를 열어 두고 R의 수치·문서·피드백과 학습·전이를 확인하세요. 서로 "
        "독립적인 새 질문을 만들지 마세요.\n"
        "- 같은 detail의 여러 slot도 work_element와 task_semantics에 맞춰 서로 다른 사건·판단·산출물을 "
        "물어야 합니다. 다른 slot의 질문 골격이나 '절차·절차' 같은 구를 반복하지 마세요.\n"
        "- evaluation_points는 질문과 꼬리질문에서 실제로 요구한 응답만 평가하도록 ① 당시 사건과 본인 "
        "역할 ② 해당 slot에 맞는 구체 판단 근거·적용 범위 또는 수행 순서 ③ 본인이 직접 한 행동과 "
        "산출물 ④ 문서·수치·기록·피드백으로 입증한 결과를 각각 한 문장으로 작성하세요. 'KSA 고유', "
        "'직무역량', '구체성' 같은 추상 평가어만 쓰지 마세요.\n"
        "[메타데이터 규칙]\n"
        "- 각 출력 row의 type='경험면접', competency, ncsClCd, question_evidence_id는 같은 slot 값을 "
        "정확히 복사하세요. 내부 NCS/KSA 명칭은 서버가 evidence_id로 복구하므로 "
        "question_focus_surface='', question_focus='', ksa_refs=[]로 출력하세요.\n"
        f"[질문 SLOT JSON]{json.dumps(slots, ensure_ascii=False, separators=(',', ':'))}\n"
        f"[신뢰하지 않는 채용 문맥 JSON]{json.dumps(context, ensure_ascii=False, separators=(',', ':'))}\n"
        + (
            f"[서버 재생성 문맥]{retry_context}\n"
            if retry_context
            else ""
        )
        + _untrusted_context_prompt_contract()
    )


_OPENAI_MODEL_OUTPUT_FAILURE_CODES = frozenset(
    {
        "model_response_not_object",
        "model_response_invalid_shape",
        "model_response_invalid_json",
        "model_response_truncated",
        "model_response_content_filtered",
        "model_response_refused",
        "model_question_count_mismatch",
        "model_question_content_missing",
        "model_question_diversity_mismatch",
        "question_set_count_or_diversity_failed",
    }
)


class _OpenRouterTimeoutRecoveryOutputError(ValueError):
    """The bounded medium timeout rescue returned unusable model output."""


def _decode_strategy_model_content(content: Any) -> Any:
    """Decode common OpenRouter JSON wrappers without inventing content.

    Ox Alpha occasionally surrounds an otherwise valid object with a Markdown
    fence or a short preamble, and OpenAI-compatible gateways may return text
    parts instead of one string. Those are transport-format differences, not
    semantic generation failures, so recover the balanced JSON value before
    applying the existing strict count/content checks.
    """

    if isinstance(content, list):
        content = "\n".join(
            str(part.get("text") or "").strip()
            for part in content
            if isinstance(part, dict) and str(part.get("text") or "").strip()
        )
    text = str(content or "").strip()
    extracted = _extract_json_text(text)
    candidates = (
        text,
        extracted,
        _slice_balanced_json(extracted),
        _slice_balanced_json(text),
    )
    decoded: Any = None
    seen: set[str] = set()
    for candidate in candidates:
        candidate = str(candidate or "").strip()
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        try:
            decoded = json.loads(candidate)
            break
        except json.JSONDecodeError:
            continue
    if decoded is None:
        raise ValueError("model_response_invalid_json")
    if isinstance(decoded, list):
        return {"interview_questions": decoded}
    if not isinstance(decoded, dict):
        return decoded

    normalized = dict(decoded)
    if not isinstance(normalized.get("interview_questions"), list):
        for alias in ("questions", "items"):
            if isinstance(normalized.get(alias), list):
                normalized["interview_questions"] = normalized[alias]
                break
        else:
            nested = normalized.get("data")
            if isinstance(nested, dict):
                nested_questions = nested.get("interview_questions")
                if not isinstance(nested_questions, list):
                    nested_questions = nested.get("questions")
                if isinstance(nested_questions, list):
                    normalized["interview_questions"] = nested_questions
    return normalized


def _safe_openai_generation_failure_reason(
    exc: BaseException,
    *,
    default: str,
) -> str:
    """Classify a provider failure without reflecting exception text to users."""

    normalized = str(exc or "").strip().casefold()
    if isinstance(exc, json.JSONDecodeError):
        return "model_response_invalid_json"
    if isinstance(exc, httpx.TimeoutException) or any(
        marker in normalized
        for marker in ("timed out", "timeout", "readtimeout", "connecttimeout")
    ):
        return (
            "openrouter_request_timeout"
            if "openrouter" in normalized
            else "openai_request_timeout"
        )
    for code in _OPENAI_MODEL_OUTPUT_FAILURE_CODES:
        if code in normalized:
            return code
    http_status = re.search(r"(openai|openrouter)_http_(\d{3})", normalized)
    if http_status:
        return f"{http_status.group(1)}_http_{http_status.group(2)}"
    for code in (
        "openrouter_request_timeout",
        "openrouter_network_unreachable",
        "openrouter_request_failed",
    ):
        if code in normalized:
            return code
    return default


def _openai_interview_response_format(
    *,
    expected_count: int,
    follow_up_count: int,
    interview_methods: list[str] | None = None,
) -> dict[str, Any]:
    """Return the strict Chat Completions schema for interview generation."""

    selected_methods = _selected_prompt_methods(interview_methods)
    question_schema: dict[str, Any] = {
        "type": "object",
        "properties": {
            "type": {"type": "string", "enum": selected_methods},
            "competency": {"type": "string"},
            "ncsClCd": {"type": "string"},
            "question": {"type": "string"},
            "follow_ups": {
                "type": "array",
                "items": {"type": "string"},
                "minItems": follow_up_count,
                "maxItems": follow_up_count,
            },
            "evaluation_points": {
                "type": "array",
                "items": {"type": "string"},
                "minItems": 4,
                "maxItems": 4,
            },
            "question_evidence_id": {"type": "string"},
            "question_focus_surface": {"type": "string"},
            "question_focus": {"type": "string"},
            "ksa_refs": {
                "type": "array",
                "items": {"type": "string"},
            },
        },
        "required": [
            "type",
            "competency",
            "ncsClCd",
            "question",
            "follow_ups",
            "evaluation_points",
            "question_evidence_id",
            "question_focus_surface",
            "question_focus",
            "ksa_refs",
        ],
        "additionalProperties": False,
    }
    return {
        "type": "json_schema",
        "json_schema": {
            "name": "ncs_interview_questions",
            "strict": True,
            "schema": {
                "type": "object",
                "properties": {
                    "interview_questions": {
                        "type": "array",
                        "items": question_schema,
                        "minItems": expected_count,
                        "maxItems": expected_count,
                    },
                    "ncs_link": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "ncsClCd": {"type": "string"},
                                "compeUnitName": {"type": "string"},
                                "why": {"type": "string"},
                            },
                            "required": ["ncsClCd", "compeUnitName", "why"],
                            "additionalProperties": False,
                        },
                    },
                },
                "required": ["interview_questions", "ncs_link"],
                "additionalProperties": False,
            },
        },
    }


def _openai_interview_completion_budget(
    target_count: int,
    reasoning_effort: str = "",
) -> int:
    """Allow enough output for rich questions while retaining a hard ceiling."""

    return quality_completion_budget(
        target_count,
        reasoning_effort=reasoning_effort,
    )


def build_strategy_with_openai(
    jd_text: str,
    notice_text: str,
    strengths: str,
    region: str,
    ncs_matches: list[dict[str, Any]],
    ncs_ksa: list[dict[str, Any]] | None = None,
    ncs_context: dict[str, Any] | None = None,
    duty_text: str = "",
    evaluation_text: str = "",
    desired_job: str = "",
    api_key_override: str = "",
    target_count_override: int | None = None,
    follow_up_count: int = 3,
    question_plan: dict[str, Any] | None = None,
    interview_methods: list[str] | None = None,
    extra_context: str = "",
    generation_provider: str = "",
    generation_model: str = "",
    max_model_requests: int = 2,
    transport_max_attempts: int = 1,
    allow_partial_model_output: bool | None = None,
) -> dict[str, Any]:
    inferred_provider = (
        OPENROUTER_PROVIDER
        if str(api_key_override or "").strip().casefold().startswith("sk-or-")
        else "openai_api"
    )
    generation_provider = normalize_generation_provider(
        generation_provider
        or os.getenv("INTERVIEW_GENERATION_PROVIDER")
        or os.getenv("JD_STRATEGY_PROVIDER")
        or inferred_provider
    )
    if generation_provider not in {"openai_api", OPENROUTER_PROVIDER}:
        raise ValueError(
            "INTERVIEW_GENERATION_PROVIDER/generation_provider must be "
            "'openai_api' or 'openrouter_api'; personal subscription CLI providers are disabled"
        )

    key_is_openrouter = str(api_key_override or "").strip().casefold().startswith("sk-or-")
    if key_is_openrouter != (generation_provider == OPENROUTER_PROVIDER):
        raise ValueError("generation_provider_key_mismatch")

    api_key = (
        settings.resolve_openrouter_key(api_key_override)
        if generation_provider == OPENROUTER_PROVIDER
        else settings.resolve_openai_key(api_key_override)
    )
    default_target = max(5, min(50, int(os.getenv("INTERVIEW_TARGET_COUNT", "10") or "10")))
    target_count = int(target_count_override or default_target)
    target_count = max(1, min(50, target_count))
    max_model_requests = max(
        1,
        min(3 if generation_provider == OPENROUTER_PROVIDER else 2, int(max_model_requests or 1)),
    )
    transport_max_attempts = max(1, min(3, int(transport_max_attempts or 1)))
    if allow_partial_model_output is None:
        allow_partial_model_output = (
            os.getenv("OPENAI_ALLOW_PARTIAL_MODEL_OUTPUT", "false").strip().lower()
            in {"1", "true", "yes", "y"}
        )
    strict_count = not bool(allow_partial_model_output)
    retry_target_count = target_count
    follow_up_count = max(0, min(5, int(follow_up_count if follow_up_count is not None else 3)))
    openai_primary_model = (
        os.getenv("OPENAI_STRATEGY_MODEL", DEFAULT_QUALITY_MODEL)
        or DEFAULT_QUALITY_MODEL
    ).strip()
    primary_model = provider_model(
        generation_provider,
        str(generation_model or openai_primary_model).strip(),
    )
    retry_model = provider_model(
        generation_provider,
        str(
            generation_model
            or os.getenv("OPENAI_STRATEGY_RETRY_MODEL", primary_model)
            or primary_model
        ).strip(),
    )
    force_fallback = (os.getenv("OPENAI_FORCE_FALLBACK", "false").strip().lower() in {"1", "true", "yes", "y"})
    if not api_key:
        return build_strategy_with_rule_fallback(
            ncs_matches=ncs_matches,
            ncs_ksa=ncs_ksa,
            error_message=f"model_generation_failed: request {generation_provider} API key is not set",
            target_count=target_count,
        )
    if force_fallback:
        return build_strategy_with_rule_fallback(
            ncs_matches=ncs_matches,
            ncs_ksa=ncs_ksa,
            error_message="model_generation_skipped: OPENAI_FORCE_FALLBACK is enabled (template fallback disabled)",
            target_count=target_count,
        )

    strict_net_check = os.getenv("OPENAI_NET_CHECK_STRICT", "false").strip().lower() in {"1", "true", "yes", "y"}
    if generation_provider == OPENROUTER_PROVIDER:
        # A separate models probe adds latency but cannot prove that this
        # specific model request will succeed. Let the bounded completion
        # requests perform authentication and capability validation directly.
        net_ok, net_msg = True, ""
    else:
        net_ok, net_msg = _check_openai_connectivity(
            api_key=api_key,
            ttl_sec=60,
            provider=generation_provider,
        )
    precheck_warning = ""
    if not net_ok:
        detail = (
            "openrouter_network_unreachable"
            if generation_provider == OPENROUTER_PROVIDER
            else "openai_network_unreachable"
        )
        if strict_net_check:
            return build_strategy_with_rule_fallback(
                ncs_matches=ncs_matches,
                ncs_ksa=ncs_ksa,
                error_message=f"model_generation_skipped: {detail}",
                target_count=target_count,
            )
        precheck_warning = detail

    strengths = (strengths or "").strip()
    duty_text = (duty_text or "").strip()
    evaluation_text = (evaluation_text or "").strip()
    extra_context = str(extra_context or "").strip()
    has_priority_context = bool(duty_text or evaluation_text)
    priority_rules = ""
    if has_priority_context:
        priority_rules = (
            "[우선 반영 규칙]\n"
            "- 담당업무/면접평가항목 텍스트를 JD 일반문맥보다 우선 반영하세요.\n"
            "- 담당업무/평가항목의 뜻은 실제 사건·산출물·이해관계자·판단으로 번역하고, 입력의 제목이나 핵심 표현을 질문 문장에 그대로 복사하지 마세요.\n"
            "- 위 두 입력과 무관한 일반론 질문은 생성하지 마세요.\n"
            f"[담당업무-최우선]{duty_text[:2200]}\n"
            f"[면접평가항목-최우선]{evaluation_text[:1600]}\n\n"
        )
    profile_mode = (
        "개인특성 모드: 개인 강점/경험을 반드시 질문에 반영하세요."
        if strengths
        else "개인특성 모드: 개인 강점 입력이 없으므로 JD/NCS 기준으로만 생성하세요."
    )
    plan_items: list[dict[str, Any]] = []
    if isinstance(question_plan, dict):
        plan_items = [
            x for x in (question_plan.get("selected_items") or [])
            if isinstance(x, dict)
        ]
    method_names = [str(x).strip() for x in (interview_methods or []) if str(x).strip()]
    if not method_names:
        method_names = list(_PROMPT_INTERVIEW_METHODS)
    method_schema_hint = "|".join(_selected_prompt_methods(method_names))
    planned_sequence = _planned_question_sequence_for_prompt(
        question_plan,
        method_names,
        target_count,
        ncs_matches=ncs_matches,
        ncs_ksa=ncs_ksa,
    )
    custom_plan_rules = ""
    if plan_items:
        sequence_rules = ""
        if planned_sequence:
            sequence_rules = (
                f"[질문별 생성 순서]{json.dumps(planned_sequence, ensure_ascii=False)}\n"
                "- interview_questions 배열 순서는 [질문별 생성 순서]의 index와 정확히 같아야 합니다.\n"
                "- 각 index의 detail, compeUnitName, required_job_context, required_factorName, required_ksa_type은 내부 분류·근거 힌트입니다. 지원자용 question/follow_ups/evaluation_points에 공식 라벨 원문을 복사하지 마세요.\n"
                "- required_surface_focus는 공개 가능한 업무 의미 힌트입니다. 그 대상·행동이 주질문에서 식별되게 구체 문서·자료·산출물로 자연스럽게 표현하되, 힌트 문구 자체를 여러 번 복사하지 마세요.\n"
                "- required_ksa_type에 따라 지식은 고유 적용 논리, 기술은 고유 수행 흔적과 도메인 산출물, 태도는 중립적인 딜레마에서 지원자가 고른 대응·권한 내 행동·상충효과·검증 또는 수정 방식을 관찰하게 하세요.\n"
                "- 태도이면 question은 핵심 선택 1개와 판단 산출물 1개만 요구하고, 지원자의 역할·승인 권한 안에서 가능한 행동을 열어 두세요. 비용·반발 감수나 개인의 결과 책임을 전제하지 말고, 빠진 상충효과와 검증·수정·담당 역할은 답변 연동 follow_ups로 확인하세요.\n"
                "- 진행·조건부 진행·보류·권한자 이송은 선택지의 예일 뿐 특정 행동을 정답으로 강제하지 마세요. 보고서 작성 요령이면 question 자체에 확정/잠정 구분·본문/주석 배치·증빙 연결 중 최소 2개를 같은 보고서 한 장에 적용하게 하세요.\n"
                "- required_task_statement와 required_observable_behavior의 뜻을 관찰 가능한 판단·행동·산출물로 번역하되 그 문구도 완성문 골격처럼 복사하지 마세요.\n"
                "- question_evidence_id, question_focus_surface, question_focus, ksa_refs에는 같은 index의 내부 값을 정확히 보존해 문항과 근거를 연결하세요.\n"
                "- required_scenario_frame은 해당 index의 required_task_statement에 맞춰 서버가 만든 KSA 정렬 상황 축입니다. 다른 일반 사건으로 교체하지 말고, 같은 표현을 복사하는 대신 구체 문서·데이터·이해관계자·제약을 정하세요. frame이 다르면 사건도 달라야 합니다.\n"
                "- required_difficulty는 문항의 추론 난이도, required_question_angle은 주된 관찰 초점, required_constraint_axis는 사건의 핵심 제약입니다. 세 축을 해당 index의 사건·판단·최소 산출물에 실제로 반영하되 지원자에게 축 이름을 읽어 주지는 마세요.\n"
                "- 기본/심화/고난도, 적용 근거/직접 수행/오류·예외/결과 검증, 자료 불일치/마감/이해관계자/권한/자원 제약이 전체 세트에서 고르게 분산되어야 합니다.\n"
                "- required_question_example은 완성 질문이 아니라 면접기법별 설계 자산입니다. 필요한 자산만 골라 자연스러운 하나의 사건으로 구성하고 문구를 복사하지 마세요.\n"
                "- required_followup_focus_example은 완성 꼬리질문이 아니라 답변 연동 방식입니다. 지정 slot을 포함해 꼬리질문 3개 중 최소 2개가 지원자의 직전 답변 내용·누락·선택·결과를 명시적으로 받아 묻게 하세요. 가능하면 꼬리1은 '방금 …', 꼬리2는 '앞서 …'로 시작해 참조 대상을 드러내세요.\n"
                "- type=토론면접이면 '[토론과제]'로 시작하고 현장 사건, 구체적인 두 입장 충돌, 근거 검토와 공동안 또는 미합의 쟁점·결정권자 이송 기준을 포함하세요. 합의를 강제하거나 시간·입장발표 조건을 넣지 마세요.\n"
                "- type=인바스켓면접이면 구체 문서·요청의 마감과 권한 충돌을 제시하고 지정 follow_up slot에서 지원자가 고른 우선순위와 처리 주체를 받아 묻습니다.\n"
                "- 각 index의 type은 반드시 [질문별 생성 순서]의 type과 같아야 합니다. 이전 index의 직무 표현을 다음 index로 재사용하지 마세요.\n"
                "- index별 detail/type이 맞지 않으면 해당 모델 질문은 템플릿으로 교체됩니다.\n"
            )
        custom_plan_rules = (
            "[사용자 지정 질문 계획]\n"
            "- 아래 세분류만 질문 생성 대상으로 사용하세요.\n"
            "- 각 세분류별 main_count만큼 주질문을 생성하세요.\n"
            "- 각 주질문에는 follow_ups를 지정 개수만큼 생성하세요.\n"
            f"{json.dumps(plan_items, ensure_ascii=False)}\n"
            f"{sequence_rules}"
            f"[선택 면접기법]{', '.join(method_names)}\n"
            "- 각 질문의 type과 method에는 선택 면접기법 중 하나를 넣으세요.\n"
            "[배정 KSA 중심 설계]\n"
            "- evidence_id가 배정된 index를 지원동기·일반 협업·강점 같은 일반 질문으로 바꾸지 마세요.\n"
            "- required_scenario_frame보다 익숙한 일반 사건을 우선하지 말고, required_task_statement의 판단·행동이 사건 해결의 핵심이 되게 하세요.\n"
            "- 직무/NCS 질문은 같은 면접기법을 반복하더라도 required_surface_focus와 required_scenario_frame이 가리키는 사건·대상·판단이 서로 달라야 합니다. surface의 업무 의미는 주질문에 보이게 하되 기계적인 원문 반복은 피하세요.\n"
            "- 질문마다 배정 KSA의 적용 근거·예외, 수행 행동, 결과 증거 중 주된 검증 초점을 다르게 설계하세요.\n\n"
        )

    _guide_summary = (
        _load_structured_interview_guide_summary()
        if len(_selected_prompt_methods(method_names)) == len(_PROMPT_INTERVIEW_METHODS)
        else ""
    )
    _gate_contract = _model_question_gate_contract(method_names)

    prompt = (
        "JSON만 출력하세요.\n"
        "목표: NCS 능력단위 기반 구조화 면접 질문 생성\n"
        "언어: 모든 문자열은 한국어\n"
        "출력 스키마: {"
        f'"interview_questions":[{{"type":"{method_schema_hint}","competency":"능력단위명","ncsClCd":"코드","question":"경험형은 실제 사건·역할·KSA 고유 행동 하나·관찰 결과, 과제형은 핵심 판단 1개와 최소 산출물 1개를 묻는 주질문","follow_ups":["조건부 답변 연동 질문","조건부 답변 연동 질문","표준화 가능 질문"],"evaluation_points":["관찰 가능한 핵심1","관찰 가능한 핵심2","관찰 가능한 핵심3","관찰 가능한 핵심4"],"question_evidence_id":"배정된 evidence_id","question_focus_surface":"required_surface_focus 원문","question_focus":"평가위원용 required_factorName","ksa_refs":["평가위원용 required_factorName"]}}],'
        '"ncs_link":[{"ncsClCd":"...","compeUnitName":"...","why":"..."}]'
        "}\n\n"
        "[구조화 면접 원칙]\n"
        f"{_untrusted_context_prompt_contract()}"
        f"{_guide_summary}\n\n"
        f"{_gate_contract}\n"
        f"{custom_plan_rules}"
        f"{priority_rules}"
        "생성 규칙:\n"
        f"- interview_questions {target_count}개 생성\n"
        "- 전체 세트를 직무/능력단위, 상황, 난이도, KSA(지식·기술·태도), 질문 유형의 5개 축으로 먼저 배정한 뒤 작성\n"
        "- 난이도는 기본(핵심 근거 적용), 심화(상충 근거 비교), 고난도(불완전 정보·권한·시간 제약 아래 예외 판단)를 가능한 한 고르게 분산\n"
        "- 같은 문장을 바꿔 쓰는 것은 다른 후보가 아님. 사건 사실, 판단 갈등, 요구 산출물 중 최소 2개가 달라야 함\n"
        f"- 각 항목: 주질문 1개 + follow_ups 꼬리질문 정확히 {follow_up_count}개\n"
        f"- type/method는 선택 면접기법({', '.join(method_names)}) 중 하나만 사용\n"
        "- evidence_id가 배정된 문항은 일반 질문으로 대체하지 말고 해당 KSA의 판단·행동·결과를 직접 검증\n"
        "- 직무/NCS 질문은 서로 다른 내부 의미 힌트와 scenario frame을 실제 사건·대상·제약으로 번역해 질문 의도가 겹치지 않게 생성\n"
        "- question_evidence_id에는 같은 index에 배정된 evidence_id를 그대로 넣고, question_focus_surface에는 required_surface_focus를 넣으세요.\n"
        "- question_focus_surface, question_focus, ksa_refs는 추적 필드입니다. 각각 required_surface_focus와 required_factorName을 정확히 저장하세요. 지원자용 문장에는 공식 required_factorName은 숨기되 required_surface_focus의 구체 업무 의미는 주질문에서 관찰 가능해야 합니다.\n"
        "\n"
        "[주질문 작성 필수 기준]\n"
        "- 경험면접을 제외한 과제형 주질문은 핵심 판단 1개와 그 판단을 기록하는 최소 산출물 1개에 집중하고, 경험면접은 실제 사건·역할·행동 하나·관찰 결과만 질문. 다른 판단 family와 협의·기록·사후점검은 꼬리질문으로 이동\n"
        "- 정밀한 수치·계산·조항 요구는 [서버 검증 자료가 없는 정밀 요구 경계]를 따르며, 업로드 원문에 표나 조항이 보인다는 이유로 예외를 두지 않음\n"
        "- 준비·발표·토론·질의응답 시간과 제출 방식은 question 본문이 아니라 task_conditions로 분리\n"
        "- 면접기법별 자산은 선택적으로 조합하고, 모든 문항에 같은 체크리스트를 기계적으로 나열하지 말 것\n"
        "\n"
        "[꼬리질문 작성 기준]\n"
        "- 꼬리물기 구조: 주질문 → 꼬리질문, 앞 답변에서 실제로 언급한 내용·누락·선택·결과를 받아 더 깊이 파고드는 질문\n"
        "- follow_ups가 3개이면 최소 2개는 '방금 말씀하신 선택', '앞서 언급한 결과', '답변에 근거가 없다면'처럼 답변 참조 또는 조건을 명시하고, 나머지 1개만 표준화 가능\n"
        "- 적응형 꼬리질문 2개는 가능하면 각각 '방금 …', '앞서 …'로 시작하고, 앞 답변에서 언급한 자료·선택·누락·결과 중 하나를 다시 집어야 함. 수정·승인·변화를 사실로 전제하지 말고 '수정했다면/하지 않았다면', '변화를 만들었다면/없었다면'처럼 조건을 두 갈래로 열 것. 답변 참조 없이 일반론으로 반복하면 실패\n"
        "- 꼬리1·2·3은 각각 evaluation_points의 서로 다른 항목을 검증\n"
        "- 같은 내용을 반복하거나 독립적인 질문 나열 금지\n"
        "- 주질문은 개방형 단일 의도, '네/아니오'로 답할 수 없는 문장\n"
        "- 각 질문은 compeUnitDef의 의미를 실제 업무 사건으로 반영하되 능력단위명이나 정의 문구를 복사하지 말 것\n"
        "- evaluation_points는 정확히 4개이며, 모두 질문과 꼬리질문에서 직접 관찰 가능해야 하고 질문하지 않은 숨은 기준은 금지\n"
        "- 동일 패턴('~경험을 말씀해 주세요' 반복) 금지 — 질문마다 다른 도입부와 다른 검증 초점 사용\n"
        f"[생성시드]{int(time.time())}\n"
        f"{profile_mode}\n"
        f"[희망직무]{desired_job}\n"
        f"[희망지역]{region}\n"
        f"[개인강점]{strengths}\n"
        f"[공고문]{notice_text[:1500]}\n"
        f"[직무기술서]{jd_text[:1500]}\n"
        f"[매칭NCS]{json.dumps((ncs_matches or [])[:5], ensure_ascii=False)}\n"
        f"[NCS평가요소]{json.dumps((ncs_ksa or [])[:15], ensure_ascii=False)}\n"
        + (f"[추가 반복회피 컨텍스트]\n{extra_context[:2000]}\n" if extra_context else "")
    )
    experience_only = method_names == ["경험면접"] and bool(planned_sequence)
    if experience_only:
        prompt = _experience_only_generation_prompt(
            planned_sequence=planned_sequence,
            target_count=target_count,
            follow_up_count=follow_up_count,
            notice_text=notice_text,
            jd_text=jd_text,
            duty_text=duty_text,
            evaluation_text=evaluation_text,
            extra_context=extra_context,
        )
    candidate_variants = quality_candidate_variants(
        "OPENAI_STRATEGY_CANDIDATE_MULTIPLIER",
        default=3.0,
    )
    if generation_provider == OPENROUTER_PROVIDER:
        # Ox Alpha does not advertise multi-choice ``n``. Preserve the 2–3x
        # candidate pool with independent bounded requests instead.
        max_model_requests = max(max_model_requests, candidate_variants)
    prompt += (
        "\n\n[후보 풀 운영]\n"
        f"- 서버는 같은 슬롯 계획에 대해 독립 후보 세트 {candidate_variants}개를 받아 "
        "의미 중복을 제거하고 품질·다양성 점수로 최종 선별합니다.\n"
        "- 각 응답 선택지는 모든 슬롯을 완결해야 하며, 익숙한 사건 골격을 반복하지 말고 "
        "직무·상황·난이도·KSA·면접기법 축을 함께 점검하세요.\n"
    )

    payload = {
        "model": primary_model,
        "messages": [
            {"role": "system", "content": "너는 공공기관 면접 코치다. 반드시 한국어 JSON만 출력한다."},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.25,
        "n": candidate_variants,
        "response_format": _openai_interview_response_format(
            expected_count=target_count,
            follow_up_count=follow_up_count,
            interview_methods=method_names,
        ),
    }
    if generation_provider == OPENROUTER_PROVIDER:
        configured_openrouter_effort = str(
            os.getenv("OPENROUTER_PRIMARY_REASONING_EFFORT") or "max"
        ).strip().casefold()
        if configured_openrouter_effort not in {"low", "medium", "high", "xhigh", "max"}:
            configured_openrouter_effort = "max"
        primary_reasoning_effort = configured_openrouter_effort
        payload["reasoning_effort"] = configured_openrouter_effort
        payload.pop("temperature", None)
    else:
        primary_reasoning_effort = apply_quality_reasoning(
            payload,
            model=primary_model,
            specific_env_name="OPENAI_STRATEGY_REASONING_EFFORT",
        )
    payload["max_completion_tokens"] = _openai_interview_completion_budget(
        target_count,
        primary_reasoning_effort,
    )
    retry_reasoning_effort = ""
    timeout_sec = provider_timeout_sec(
        generation_provider,
        float(os.getenv("OPENAI_STRATEGY_TIMEOUT_SEC", "120") or "120"),
    )
    model_error = ""
    recovered_with_slim_retry = False
    model_request_count = 0
    obj: dict[str, Any] = {}

    def _request_json(
        local_payload: dict[str, Any],
        local_timeout: float,
        *,
        expected_count: int,
    ) -> dict[str, Any]:
        nonlocal model_request_count
        remaining_request_budget = max_model_requests - model_request_count
        if remaining_request_budget <= 0:
            raise RuntimeError("provider_generation_request_budget_exhausted")
        request_count = (
            min(candidate_variants, remaining_request_budget)
            if generation_provider == OPENROUTER_PROVIDER
            else 1
        )

        def _post_variant(variant_index: int) -> tuple[int, dict[str, Any], bool]:
            variant_payload = prepare_chat_payload(local_payload, generation_provider)
            if generation_provider == OPENROUTER_PROVIDER:
                messages = variant_payload.get("messages")
                if isinstance(messages, list) and messages:
                    last_message = messages[-1]
                    if isinstance(last_message, dict):
                        last_message["content"] = (
                            str(last_message.get("content") or "")
                            + f"\n\n[독립 후보 세트 {variant_index}/{request_count}] "
                            "다른 후보와 사건·판단 갈등·산출물 중 최소 2개가 다르게 작성하세요."
                        )
            data = post_chat_completions_with_retries(
                payload=variant_payload,
                api_key=api_key,
                timeout_sec=local_timeout,
                max_attempts=transport_max_attempts,
                provider=generation_provider,
            )
            timeout_recovery_used = bool(
                isinstance(data, dict)
                and data.get("_ncscope_openrouter_timeout_recovery_used") is True
            )
            if isinstance(data, dict):
                data = dict(data)
                data.pop("_ncscope_openrouter_timeout_recovery_used", None)
            return variant_index, data, timeout_recovery_used

        model_request_count += request_count
        response_sets: list[tuple[int, dict[str, Any], bool]] = []
        request_errors: list[BaseException] = []
        concurrency = provider_candidate_concurrency(generation_provider, request_count)
        if generation_provider == OPENROUTER_PROVIDER and concurrency > 1:
            with ThreadPoolExecutor(
                max_workers=concurrency,
                thread_name_prefix="openrouter-strategy-candidate",
            ) as executor:
                futures = {
                    executor.submit(_post_variant, index): index
                    for index in range(1, request_count + 1)
                }
                for future in as_completed(futures):
                    try:
                        response_sets.append(future.result())
                    except BaseException as exc:
                        request_errors.append(exc)
        else:
            for index in range(1, request_count + 1):
                try:
                    response_sets.append(_post_variant(index))
                except BaseException as exc:
                    request_errors.append(exc)

        if not response_sets:
            if request_errors:
                raise request_errors[0]
            raise ValueError("model_response_invalid_shape")

        indexed_choices: list[tuple[int, int, dict[str, Any]]] = []
        timeout_recovery_used = any(
            recovered for _index, _data, recovered in response_sets
        )
        for request_variant_index, data, _recovered in sorted(
            response_sets,
            key=lambda item: item[0],
        ):
            choices = data.get("choices") if isinstance(data, dict) else None
            if not isinstance(choices, list):
                continue
            for choice_index, choice in enumerate(choices, start=1):
                if isinstance(choice, dict):
                    indexed_choices.append(
                        (request_variant_index, choice_index, choice)
                    )
        if not indexed_choices:
            raise ValueError("model_response_invalid_shape")
        valid_responses: list[dict[str, Any]] = []
        failure_codes: list[str] = []
        for request_variant_index, choice_index, choice in indexed_choices:
            variant_index = (
                request_variant_index
                if generation_provider == OPENROUTER_PROVIDER
                else choice_index
            )
            if not isinstance(choice, dict):
                failure_codes.append("model_response_invalid_shape")
                continue
            finish_reason = str(choice.get("finish_reason") or "").strip().casefold()
            if finish_reason == "length":
                failure_codes.append("model_response_truncated")
                continue
            if finish_reason == "content_filter":
                failure_codes.append("model_response_content_filtered")
                continue
            message = choice.get("message")
            if not isinstance(message, dict):
                failure_codes.append("model_response_invalid_shape")
                continue
            if str(message.get("refusal") or "").strip():
                failure_codes.append("model_response_refused")
                continue
            content = message.get("content")
            if not content:
                failure_codes.append("model_response_invalid_shape")
                continue
            try:
                parsed = _decode_strategy_model_content(content)
            except ValueError:
                failure_codes.append("model_response_invalid_json")
                continue
            if not isinstance(parsed, dict):
                failure_codes.append("model_response_not_object")
                continue
            generated_questions = parsed.get("interview_questions")
            if not isinstance(generated_questions, list):
                failure_codes.append("model_question_count_mismatch")
                continue

            cleaned_questions: list[dict[str, Any]] = []
            for question in generated_questions:
                if not isinstance(question, dict):
                    continue
                if not str(question.get("question") or "").strip():
                    continue
                cleaned_questions.append(dict(question))
            if not cleaned_questions:
                failure_codes.append("model_question_content_missing")
                continue

            raw_question_count = len(cleaned_questions)
            # Some OpenRouter models honor the schema fields but still return
            # more array items than requested. Extra complete questions are
            # safe to discard deterministically; only an underfilled set needs
            # another model attempt (or a fail-closed response in strict mode).
            if raw_question_count < expected_count and strict_count:
                failure_codes.append("model_question_count_mismatch")
                continue
            if raw_question_count > expected_count:
                cleaned_questions = cleaned_questions[:expected_count]
            for slot_index, question in enumerate(cleaned_questions):
                question["_candidate_slot"] = slot_index
                question["_candidate_variant"] = variant_index
            parsed["interview_questions"] = cleaned_questions
            parsed["_model_question_raw_count"] = raw_question_count
            parsed["_model_question_count_mismatch"] = raw_question_count < expected_count
            valid_responses.append(parsed)

        if not valid_responses:
            failure_priority = (
                "model_response_truncated",
                "model_response_content_filtered",
                "model_response_refused",
                "model_question_count_mismatch",
                "model_question_content_missing",
                "model_response_invalid_json",
                "model_response_not_object",
                "model_response_invalid_shape",
            )
            failure_set = set(failure_codes)
            code = next(
                (candidate for candidate in failure_priority if candidate in failure_set),
                "model_response_invalid_shape",
            )
            if timeout_recovery_used:
                raise _OpenRouterTimeoutRecoveryOutputError(code)
            raise ValueError(code)

        selected_response = dict(valid_responses[0])
        first_questions = [
            dict(question)
            for question in valid_responses[0].get("interview_questions", [])
            if isinstance(question, dict)
        ]
        if len(valid_responses) > 1:
            candidate_pool = [
                dict(question)
                for response in valid_responses
                for question in response.get("interview_questions", [])
                if isinstance(question, dict)
            ]
            selected_questions, selection_metadata = select_question_candidates(
                candidate_pool,
                expected_count,
            )
            if len(selected_questions) != expected_count:
                raise ValueError("model_question_count_mismatch")
            selected_questions.sort(
                key=lambda question: int(question.get("_candidate_slot", expected_count))
            )
            selected_response["interview_questions"] = selected_questions
            selected_response["question_candidate_selection"] = {
                **selection_metadata,
                "requested_variant_count": (
                    request_count
                    if generation_provider == OPENROUTER_PROVIDER
                    else int(local_payload.get("n") or 1)
                ),
                "received_variant_count": len(valid_responses),
                "failed_variant_count": max(0, request_count - len(response_sets)),
                "candidate_pool_count": len(candidate_pool),
            }
            selected_response["_model_question_raw_count"] = expected_count
            selected_response["_model_question_count_mismatch"] = False
        else:
            selected_response["interview_questions"] = first_questions
            selected_response["question_candidate_selection"] = {
                "strategy": "single_valid_candidate_set",
                "requested_variant_count": (
                    request_count
                    if generation_provider == OPENROUTER_PROVIDER
                    else int(local_payload.get("n") or 1)
                ),
                "received_variant_count": 1,
                "failed_variant_count": max(0, request_count - len(response_sets)),
                "candidate_pool_count": len(first_questions),
            }

        for question in selected_response.get("interview_questions", []):
            if not isinstance(question, dict):
                continue
            question.pop("_candidate_slot", None)
            question.pop("_candidate_variant", None)
        return selected_response

    try:
        obj = _request_json(payload, timeout_sec, expected_count=target_count)
    except Exception as primary_exc:
        # First failure: retry with slimmer, compeUnitDef-focused prompt.
        primary_reason = _safe_openai_generation_failure_reason(
            primary_exc,
            default="primary_request_failed",
        )
        log_primary_failure = (
            logger.warning
            if primary_reason == "openrouter_request_timeout"
            else logger.error
        )
        log_primary_failure(
            "strategy_primary_failed provider=%s reason=%s",
            generation_provider,
            primary_reason,
        )
        allow_slim_retry = not (
            generation_provider == OPENROUTER_PROVIDER
            and (
                primary_reason == "openrouter_request_timeout"
                or isinstance(
                    primary_exc,
                    _OpenRouterTimeoutRecoveryOutputError,
                )
            )
        )
        model_error = primary_reason
        slim_priority = ""
        if model_request_count >= max_model_requests:
            obj = {}
            # The outer institution quality retry has already consumed the
            # second semantic generation budget.  Do not hide more upstream
            # calls behind this builder's slim retry.
            model_error = primary_reason
        if (
            allow_slim_retry
            and model_request_count < max_model_requests
            and has_priority_context
        ):
            slim_priority = (
                f"[priority_duty]{duty_text[:1500]}\n"
                f"[priority_eval]{evaluation_text[:1200]}\n"
            )
        slim_prompt = (
            "JSON만 출력하세요.\n"
            "목표: NCS 능력단위 기반 구조화 면접 질문 생성\n"
            "언어: 한국어\n"
            "스키마: {"
            f'"interview_questions":[{{"type":"{method_schema_hint}","competency":"...","ncsClCd":"...","question":"경험형은 실제 사건·역할·KSA 고유 행동 하나·관찰 결과, 과제형은 핵심 판단 1개와 최소 산출물 1개를 묻는 주질문","follow_ups":["조건부 답변 연동 질문","조건부 답변 연동 질문","표준화 가능 질문"],"evaluation_points":["직접 관찰 가능한 핵심1","핵심2","핵심3","핵심4"],"question_evidence_id":"배정된 evidence_id","question_focus_surface":"required_surface_focus 원문","question_focus":"평가위원용 required_factorName","ksa_refs":["평가위원용 required_factorName"]}}],'
            '"ncs_link":[{"ncsClCd":"...","compeUnitName":"...","why":"..."}]'
            "}\n"
            "규칙:\n"
            f"{_untrusted_context_prompt_contract()}"
            f"- interview_questions {retry_target_count}개 생성\n"
            f"- 각 항목: 주질문 1개 + follow_ups 꼬리질문 {follow_up_count}개 (3개이면 최소 2개는 앞 답변의 내용·누락·선택·결과를 명시적으로 받아 묻고, 1개만 표준화 가능)\n"
            "- 적응형 꼬리질문 2개는 가능하면 각각 '방금 …', '앞서 …'로 시작해 참조 대상을 드러내되, 수정·승인·변화를 사실로 전제하지 말고 '수정했다면/하지 않았다면', '변화를 만들었다면/없었다면'처럼 조건을 두 갈래로 열 것\n"
            f"- 선택 면접기법: {', '.join(method_names)}\n"
            f"{_gate_contract}"
            f"{custom_plan_rules}"
            "- 경험면접을 제외한 과제형 주질문은 핵심 판단 1개와 최소 산출물 1개만 묻고, 경험면접은 실제 사건·역할·행동 하나·관찰 결과만 질문. 다른 판단 family와 협의·기록·사후점검은 꼬리질문으로 이동\n"
            "- 태도는 정답이 없는 가치·결과 충돌을 중립적으로 제시하고, question에는 역할·승인 권한 안의 핵심 선택 1개와 판단 산출물 1개만 요구. 비용·반발 감수나 개인 결과 책임을 전제하지 않음\n"
            "- 태도의 행동은 지원자가 가능한 대응 중 고른 권한 내 조치로 관찰하고 특정한 보류·수정·거절을 정답으로 강제하지 않음. 상충효과와 검증·수정 조건·담당 역할은 산출물 핵심 필드 3개 이하 또는 답변 연동 follow_ups로 확인\n"
            "- 책임성은 개인의 결과 책임이 아니라 판단 기록·승인/이송 경계·확인 지표·오류 시 수정 조건과 담당 역할로 관찰. 보고서 작성 요령이면 확정/잠정 구분·본문/주석 배치·증빙 연결 중 최소 2개를 같은 산출물 한 장에 적용\n"
            "- 정밀한 수치·계산·조항 요구는 [서버 검증 자료가 없는 정밀 요구 경계]를 따르며, 업로드 원문에 표나 조항이 보인다는 이유로 예외를 두지 않음\n"
            "- 각 질문은 compeUnitDef의 의미를 실제 사건·문서·데이터·판단으로 번역하되 능력단위명과 정의 문구를 복사하지 말 것\n"
            "- evaluation_points는 정확히 4개이며 모두 질문과 꼬리질문에서 직접 관찰 가능해야 하고 숨은 기준은 금지\n"
            "- evidence_id가 배정된 문항은 지원동기·일반 협업 같은 질문으로 대체하지 말고 해당 KSA의 판단·행동·결과를 직접 검증\n"
            "- 직무/NCS 질문은 required_surface_focus와 required_scenario_frame이 가리키는 사건·대상·판단을 문항마다 다르게 구성하고, surface의 업무 의미가 주질문에서 식별되게 하되 문구를 기계적으로 반복하지 말 것\n"
            f"{slim_priority}"
            f"[ncs_matches]{json.dumps((ncs_matches or [])[:5], ensure_ascii=False)}\n"
            f"[ncs_factors]{json.dumps((ncs_ksa or [])[:20], ensure_ascii=False)}\n"
            f"[notice_core]{notice_text[:1200]}\n"
            f"[jd_core]{jd_text[:1200]}\n"
            + (f"[avoid_questions]\n{extra_context[:1200]}\n" if extra_context else "")
        )
        if experience_only:
            slim_prompt = _experience_only_generation_prompt(
                planned_sequence=planned_sequence,
                target_count=retry_target_count,
                follow_up_count=follow_up_count,
                notice_text=notice_text,
                jd_text=jd_text,
                duty_text=duty_text,
                evaluation_text=evaluation_text,
                extra_context="\n\n".join(
                    part
                    for part in (
                        "이전 응답은 JSON 형식·문항 수 또는 필수 필드 검사에 실패했습니다. "
                        "slot별 메타데이터와 정확한 배열 개수를 다시 확인하세요.",
                        str(extra_context or "").strip(),
                    )
                    if part
                ),
            )
        slim_prompt += (
            "\n\n[후보 풀 운영]\n"
            f"- 독립 후보 세트 {candidate_variants}개 각각에서 정확히 "
            f"{retry_target_count}개 슬롯을 완결하세요.\n"
            "- 직무·상황·난이도·KSA·면접기법 축과 사건·판단·산출물의 중복을 "
            "출력 전에 점검하세요.\n"
        )
        slim_payload = {
            "model": retry_model,
            "messages": [
                {"role": "system", "content": "너는 능력단위 정의 기반 면접 질문 생성기다. 한국어 JSON만 출력한다."},
                {"role": "user", "content": slim_prompt},
            ],
            "temperature": 0.2,
            "n": candidate_variants,
            "response_format": _openai_interview_response_format(
                expected_count=retry_target_count,
                follow_up_count=follow_up_count,
                interview_methods=method_names,
            ),
        }
        if generation_provider == OPENROUTER_PROVIDER:
            retry_reasoning_effort = str(
                os.getenv("OPENROUTER_INVALID_OUTPUT_RETRY_REASONING_EFFORT")
                or os.getenv("OPENROUTER_FALLBACK_REASONING_EFFORT")
                or "medium"
            ).strip().casefold()
            if retry_reasoning_effort not in {"low", "medium", "high", "xhigh"}:
                retry_reasoning_effort = "medium"
            slim_payload["reasoning_effort"] = retry_reasoning_effort
            slim_payload["_openrouter_internal_recovery_effort"] = (
                retry_reasoning_effort
            )
            if openrouter_recovery_model():
                slim_payload["_openrouter_internal_recovery_model"] = "configured"
            slim_payload.pop("temperature", None)
        else:
            retry_reasoning_effort = apply_quality_reasoning(
                slim_payload,
                model=retry_model,
                specific_env_name="OPENAI_STRATEGY_REASONING_EFFORT",
            )
        slim_payload["max_completion_tokens"] = _openai_interview_completion_budget(
            retry_target_count,
            retry_reasoning_effort,
        )
        slim_timeout_sec = min(timeout_sec, 90.0)
        if generation_provider == OPENROUTER_PROVIDER:
            try:
                slim_timeout_sec = float(
                    str(
                        os.getenv("OPENROUTER_INVALID_OUTPUT_RETRY_TIMEOUT_SEC")
                        or os.getenv("OPENROUTER_FALLBACK_TIMEOUT_SEC")
                        or "65"
                    ).strip()
                )
            except (TypeError, ValueError):
                slim_timeout_sec = 65.0
            slim_timeout_sec = max(15.0, min(110.0, slim_timeout_sec))
        if allow_slim_retry and model_request_count < max_model_requests:
            try:
                obj = _request_json(
                    slim_payload,
                    slim_timeout_sec,
                    expected_count=retry_target_count,
                )
                recovered_with_slim_retry = True
                # The primary request failed, but the bounded slim retry produced
                # a valid model response.  Do not carry the stale primary error
                # into the successful result: public fail-closed boundaries treat
                # any non-empty ``error`` field as a provider failure.
                model_error = ""
            except Exception as retry_exc:
                retry_reason = _safe_openai_generation_failure_reason(
                    retry_exc,
                    default="retry_request_failed",
                )
                logger.error(
                    "strategy_retry_failed provider=%s reason=%s",
                    generation_provider,
                    retry_reason,
                )
                model_error = (
                    primary_reason
                    if retry_reason == "retry_request_failed"
                    and primary_reason != "primary_request_failed"
                    else retry_reason
                )
                obj = {}

    obj["ncs_candidates_raw"] = ncs_matches
    obj["ncs_ksa_used"] = ncs_ksa or []
    obj["ncs_context_used"] = ncs_context or {}
    obj["provider_generation_request_count"] = model_request_count
    obj["provider_generation_request_limit"] = max_model_requests
    obj["transport_attempt_limit_per_generation_request"] = transport_max_attempts
    obj["provider_generation_model"] = retry_model if recovered_with_slim_retry else primary_model
    obj["generation_provider"] = generation_provider
    obj["provider_reasoning_effort"] = (
        retry_reasoning_effort if recovered_with_slim_retry else primary_reasoning_effort
    )
    obj["provider_candidate_variant_count"] = candidate_variants
    candidate_selection = obj.get("question_candidate_selection")
    if isinstance(candidate_selection, dict):
        obj["provider_candidate_variant_received_count"] = int(
            candidate_selection.get("received_variant_count") or 0
        )
    q_list = obj.get("interview_questions")
    if not isinstance(q_list, list):
        q_list = []
    q_list = [q for q in q_list if isinstance(q, dict)]
    for question in q_list:
        question["question_source"] = generation_provider

    generated_has_content = any(str((q or {}).get("question", "")).strip() for q in q_list)
    q_raw_count = int(obj.pop("_model_question_raw_count", 0) or 0)
    q_count_mismatch = bool(obj.pop("_model_question_count_mismatch", False))
    obj["interview_questions"] = list(q_list[:target_count])
    if generated_has_content and q_raw_count > target_count:
        obj.setdefault("provider_generation_notes", []).append(
            f"model_question_count_trimmed:{q_raw_count}->{target_count}"
        )
    if generated_has_content and q_count_mismatch and strict_count:
        generated_has_content = False
        model_error = "question_set_count_or_diversity_failed"
        obj["interview_questions"] = []
    elif generated_has_content and q_count_mismatch:
        obj.setdefault("provider_generation_notes", []).append(
            f"model_question_count_adjusted:{q_raw_count}->{target_count}"
        )
    if q_raw_count:
        obj["model_question_generation_counts"] = {
            "expected": target_count,
            "actual": q_raw_count,
        }
    obj["interview_questions"] = _apply_entry_level_policy_to_questions(obj["interview_questions"])
    obj["interview_by_competency"] = _build_interview_by_competency_from_questions(obj["interview_questions"])
    if "ncs_link" not in obj or not isinstance(obj.get("ncs_link"), list):
        obj["ncs_link"] = [
            {
                "ncsClCd": str(x.get("ncsClCd", "")).strip(),
                "compeUnitName": str(x.get("compeUnitName", "")).strip(),
                "why": "NCS 기반 자동 매핑",
            }
            for x in (ncs_matches or [])[:6]
        ]

    if generated_has_content:
        obj["question_generation_policy"] = (
            "model_autonomous_with_ncs_factor_context_slim_retry"
            if recovered_with_slim_retry
            else "model_autonomous_with_ncs_factor_context_and_competency_definition"
        )
    else:
        obj["question_generation_policy"] = "model_only_no_template_fallback"
    if model_error:
        obj["error"] = f"model_generation_failed: {model_error}"
    if precheck_warning and not model_error:
        obj["warning"] = f"{generation_provider}_precheck_warning: {precheck_warning}"
    return obj


# ---------------------------------------------------------------------------
# Recovered NCS pipeline helpers
# ---------------------------------------------------------------------------
_SCLASS_CSV_CACHE: dict[str, Any] = {"ts": 0.0, "items": [], "path": ""}
_SCLASS_SYNONYM_CACHE: dict[str, Any] = {
    "ts": 0.0,
    "items": {"by_code_no": {}, "by_name": {}},
    "path": "",
}
_KSA_FACTOR_CACHE_BY_CODE: dict[str, list[dict[str, str]]] = {}
_NCS_XLSX_CACHE: dict[str, Any] = {"ts": 0.0, "items": [], "path": "", "map": {}}
_NCS_LOCAL_DB_STATE: dict[str, Any] = {"ready": False, "db_path": "", "xlsx_path": "", "xlsx_mtime": 0.0}


def _default_sclass_csv_path() -> str:
    here = os.path.abspath(__file__)
    root = os.path.dirname(os.path.dirname(os.path.dirname(here)))
    return os.path.join(root, "ncs_sclass_codes_with_code_no.csv")


def _default_sclass_synonym_path() -> str:
    here = os.path.abspath(__file__)
    root = os.path.dirname(os.path.dirname(os.path.dirname(here)))
    return os.path.join(root, "app", "data", "ncs_sclass_synonyms.json")


def _default_ncs_xlsx_path() -> str:
    here = os.path.abspath(__file__)
    root = os.path.dirname(os.path.dirname(os.path.dirname(here)))
    return os.path.join(root, "NCS_DB.xlsx")


def _default_app_db_path() -> str:
    here = os.path.abspath(__file__)
    root = os.path.dirname(os.path.dirname(os.path.dirname(here)))
    return os.path.join(root, "ncscope.db")


def _connect_local_db(db_path: str) -> sqlite3.Connection:
    con = sqlite3.connect(db_path, timeout=30.0)
    con.row_factory = sqlite3.Row
    return con


def ensure_ncs_local_index(db_path: str | None = None, xlsx_path: str | None = None) -> bool:
    dbp = db_path or os.getenv("APP_DB_PATH", "").strip() or _default_app_db_path()
    xlsx = xlsx_path or os.getenv("NCS_DB_XLSX_PATH", "").strip() or _default_ncs_xlsx_path()
    if not os.path.exists(dbp) or not os.path.exists(xlsx):
        return False

    xlsx_mtime = os.path.getmtime(xlsx)
    if (
        _NCS_LOCAL_DB_STATE.get("ready")
        and _NCS_LOCAL_DB_STATE.get("db_path") == dbp
        and _NCS_LOCAL_DB_STATE.get("xlsx_path") == xlsx
        and float(_NCS_LOCAL_DB_STATE.get("xlsx_mtime", 0.0)) == float(xlsx_mtime)
    ):
        return True

    try:
        import openpyxl  # type: ignore
    except Exception:
        return False

    try:
        con = _connect_local_db(dbp)
        cur = con.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS ncs_local_rows (
                ncs_cl_cd TEXT NOT NULL,
                code6 TEXT NOT NULL,
                compe_unit_name TEXT,
                compe_unit_level TEXT,
                ncs_lclass_code TEXT,
                ncs_lclass_name TEXT,
                ncs_mclass_code TEXT,
                ncs_mclass_name TEXT,
                ncs_sclass_code TEXT,
                ncs_sclass_name TEXT,
                ncs_subd_code TEXT,
                ncs_subd_name TEXT,
                unit_elem_name TEXT,
                unit_criteria TEXT,
                ksa_type_name TEXT,
                ksa_text TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS ncs_local_meta (
                key TEXT PRIMARY KEY,
                value TEXT
            )
            """
        )
        cur.execute("SELECT value FROM ncs_local_meta WHERE key='xlsx_mtime'")
        row = cur.fetchone()
        current = float(row["value"]) if row and row["value"] else 0.0
        cur.execute("SELECT COUNT(*) AS cnt FROM ncs_local_rows")
        cnt = int((cur.fetchone() or {"cnt": 0})["cnt"])
        if cnt > 0 and current == float(xlsx_mtime):
            con.close()
            _NCS_LOCAL_DB_STATE.update(
                {"ready": True, "db_path": dbp, "xlsx_path": xlsx, "xlsx_mtime": float(xlsx_mtime)}
            )
            return True

        cur.execute("DELETE FROM ncs_local_rows")
        con.commit()

        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        insert_sql = (
            "INSERT INTO ncs_local_rows ("
            "ncs_cl_cd, code6, compe_unit_name, compe_unit_level, "
            "ncs_lclass_code, ncs_lclass_name, ncs_mclass_code, ncs_mclass_name, "
            "ncs_sclass_code, ncs_sclass_name, ncs_subd_code, ncs_subd_name, "
            "unit_elem_name, unit_criteria, ksa_type_name, ksa_text"
            ") VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
        )
        batch: list[tuple[str, ...]] = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
                if i == 1 or not r or len(r) < 19:
                    continue
                l_cd = _normalize_code(r[0], 2)
                m_cd = _normalize_code(r[2], 2)
                s_cd = _normalize_code(r[4], 2)
                ncs_cl_cd = _normalize_code(r[8]).strip()
                if not ncs_cl_cd:
                    continue
                code6 = (f"{l_cd}{m_cd}{s_cd}" if (l_cd and m_cd and s_cd) else ncs_cl_cd[:6]).strip()
                if not code6:
                    continue
                batch.append(
                    (
                        ncs_cl_cd,
                        code6,
                        _repair_mojibake(str(r[9] or "").strip()),
                        _normalize_code(r[10]).strip(),
                        l_cd,
                        _repair_mojibake(str(r[1] or "").strip()),
                        m_cd,
                        _repair_mojibake(str(r[3] or "").strip()),
                        s_cd,
                        _repair_mojibake(str(r[5] or "").strip()),
                        _normalize_code(r[6], 2).strip(),
                        _repair_mojibake(str(r[7] or "").strip()),
                        _repair_mojibake(str(r[12] or "").strip()),
                        _repair_mojibake(str(r[14] or "").strip()),
                        _repair_mojibake(str(r[16] or "").strip()),
                        _repair_mojibake(str(r[18] or "").strip()),
                    )
                )
                if len(batch) >= 2000:
                    cur.executemany(insert_sql, batch)
                    con.commit()
                    batch = []
        if batch:
            cur.executemany(insert_sql, batch)
            con.commit()
        wb.close()

        cur.execute("CREATE INDEX IF NOT EXISTS idx_ncs_local_code6 ON ncs_local_rows(code6)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_ncs_local_clcd ON ncs_local_rows(ncs_cl_cd)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_ncs_local_code6_clcd ON ncs_local_rows(code6, ncs_cl_cd)")
        cur.execute("INSERT OR REPLACE INTO ncs_local_meta(key, value) VALUES('xlsx_mtime', ?)", (str(float(xlsx_mtime)),))
        con.commit()
        con.close()
    except Exception:
        return False

    _NCS_LOCAL_DB_STATE.update({"ready": True, "db_path": dbp, "xlsx_path": xlsx, "xlsx_mtime": float(xlsx_mtime)})
    return True


def get_ncs_local_index_status(db_path: str | None = None, xlsx_path: str | None = None) -> dict[str, Any]:
    dbp = db_path or os.getenv("APP_DB_PATH", "").strip() or _default_app_db_path()
    xlsx = xlsx_path or os.getenv("NCS_DB_XLSX_PATH", "").strip() or _default_ncs_xlsx_path()
    xlsx_exists = os.path.exists(xlsx)
    db_exists = os.path.exists(dbp)
    xlsx_mtime = os.path.getmtime(xlsx) if xlsx_exists else 0.0
    row_count = 0
    meta_mtime = 0.0
    indexed = False

    if db_exists:
        try:
            con = _connect_local_db(dbp)
            cur = con.cursor()
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='ncs_local_rows'")
            has_rows_tbl = cur.fetchone() is not None
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='ncs_local_meta'")
            has_meta_tbl = cur.fetchone() is not None
            if has_rows_tbl:
                cur.execute("SELECT COUNT(*) AS cnt FROM ncs_local_rows")
                row_count = int((cur.fetchone() or {"cnt": 0})["cnt"])
            if has_meta_tbl:
                cur.execute("SELECT value FROM ncs_local_meta WHERE key='xlsx_mtime'")
                row = cur.fetchone()
                if row and row["value"]:
                    meta_mtime = float(row["value"])
            con.close()
            indexed = bool(row_count > 0 and xlsx_exists and float(meta_mtime) == float(xlsx_mtime))
        except Exception:
            indexed = False

    return {
        "indexed": indexed,
        "db_path": dbp,
        "xlsx_path": xlsx,
        "db_exists": db_exists,
        "xlsx_exists": xlsx_exists,
        "row_count": row_count,
        "xlsx_mtime": xlsx_mtime,
        "indexed_mtime": meta_mtime,
    }


def _normalize_code(v: Any, width: int = 0) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    if s.isdigit() and width > 0:
        return s.zfill(width)
    return s


def load_ncs_rows_from_xlsx(xlsx_path: str | None = None, cache_ttl_sec: int = 60 * 30) -> list[dict[str, str]]:
    path = xlsx_path or os.getenv("NCS_DB_XLSX_PATH", "").strip() or _default_ncs_xlsx_path()
    now = time.time()
    if _NCS_XLSX_CACHE.get("items") and _NCS_XLSX_CACHE.get("path") == path:
        if (now - float(_NCS_XLSX_CACHE.get("ts", 0.0))) < cache_ttl_sec:
            return list(_NCS_XLSX_CACHE["items"])
    if not os.path.exists(path):
        return []

    try:
        import openpyxl  # type: ignore
    except Exception:
        return []

    out: list[dict[str, str]] = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if idx == 1:
                    continue
                if not row or len(row) < 19:
                    continue
                l_cd = _normalize_code(row[0], 2)
                l_nm = _repair_mojibake(str(row[1] or "").strip())
                m_cd = _normalize_code(row[2], 2)
                m_nm = _repair_mojibake(str(row[3] or "").strip())
                s_cd = _normalize_code(row[4], 2)
                s_nm = _repair_mojibake(str(row[5] or "").strip())
                subd_cd = _normalize_code(row[6], 2)
                subd_nm = _repair_mojibake(str(row[7] or "").strip())
                ncs_cl_cd = _normalize_code(row[8])
                compe_unit_name = _repair_mojibake(str(row[9] or "").strip())
                compe_unit_level = _normalize_code(row[10])
                unit_elem_name = _repair_mojibake(str(row[12] or "").strip())   # M
                unit_criteria = _repair_mojibake(str(row[14] or "").strip())    # O
                ksa_type_name = _repair_mojibake(str(row[16] or "").strip())    # Q
                ksa_text = _repair_mojibake(str(row[18] or "").strip())          # S
                if not ncs_cl_cd:
                    continue
                code_no = f"{l_cd}{m_cd}{s_cd}" if (l_cd and m_cd and s_cd) else ncs_cl_cd[:6]
                out.append(
                    {
                        "ncs_code_no": code_no,
                        "ncs_cl_cd": ncs_cl_cd,
                        "compe_unit_name": compe_unit_name,
                        "compe_unit_level": compe_unit_level,
                        "ncs_lclass_code": l_cd,
                        "ncs_lclass_name": l_nm,
                        "ncs_mclass_code": m_cd,
                        "ncs_mclass_name": m_nm,
                        "ncs_sclass_code": s_cd,
                        "ncs_sclass_name": s_nm,
                        "ncs_subd_code": subd_cd,
                        "ncs_subd_name": subd_nm,
                        "unit_elem_name": unit_elem_name,
                        "unit_criteria": unit_criteria,
                        "ksa_type_name": ksa_type_name,
                        "ksa_text": ksa_text,
                    }
                )
        wb.close()
    except Exception:
        return []

    _NCS_XLSX_CACHE["ts"] = now
    _NCS_XLSX_CACHE["path"] = path
    _NCS_XLSX_CACHE["items"] = out
    return list(out)


def _units_from_local_xlsx_by_sclass(
    ncs_lclass_code: str,
    ncs_mclass_code: str,
    ncs_sclass_code: str,
    sclass_name: str = "",
    max_items: int = 300,
) -> list[dict[str, Any]]:
    l_cd = str(ncs_lclass_code or "").strip()
    m_cd = str(ncs_mclass_code or "").strip()
    s_cd = str(ncs_sclass_code or "").strip()
    code_no = f"{l_cd}{m_cd}{s_cd}"
    if not (l_cd and m_cd and s_cd):
        return []

    db_path = os.getenv("APP_DB_PATH", "").strip() or _default_app_db_path()
    if ensure_ncs_local_index(db_path=db_path):
        try:
            con = _connect_local_db(db_path)
            cur = con.cursor()
            cur.execute(
                """
                SELECT
                    ncs_cl_cd,
                    compe_unit_name,
                    compe_unit_level,
                    ncs_sclass_name,
                    ncs_subd_code,
                    ncs_subd_name,
                    unit_criteria
                FROM ncs_local_rows
                WHERE code6 = ?
                GROUP BY ncs_cl_cd
                ORDER BY ncs_cl_cd
                LIMIT ?
                """,
                (code_no, int(max_items or 300)),
            )
            rows = cur.fetchall()
            con.close()
            out_db: list[dict[str, Any]] = []
            for r in rows:
                out_db.append(
                    {
                        "ncsClCd": str(r["ncs_cl_cd"] or "").strip(),
                        "compeUnitName": str(r["compe_unit_name"] or "").strip(),
                        "compeUnitLevel": str(r["compe_unit_level"] or "").strip(),
                        "ncsLclasCd": l_cd,
                        "ncsMclasCd": m_cd,
                        "ncsSclasCd": s_cd,
                        "ncsSclasCdnm": str(r["ncs_sclass_name"] or "").strip() or str(sclass_name or "").strip(),
                        "ncsSubdCd": str(r["ncs_subd_code"] or "").strip(),
                        "ncsSubdCdnm": str(r["ncs_subd_name"] or "").strip(),
                        "compeUnitDef": str(r["unit_criteria"] or "").strip(),
                        "score": 1.0,
                        "matched_keywords": [str(r["ncs_sclass_name"] or "").strip() or str(sclass_name or "").strip() or code_no],
                    }
                )
            if out_db:
                return out_db
        except Exception:
            pass

    cache_key = f"units:{code_no}"
    cache_map = _NCS_XLSX_CACHE.setdefault("map", {})
    if cache_key in cache_map:
        return list(cache_map.get(cache_key, []))

    path = os.getenv("NCS_DB_XLSX_PATH", "").strip() or _default_ncs_xlsx_path()
    if not os.path.exists(path):
        return []
    try:
        import openpyxl  # type: ignore
    except Exception:
        return []

    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    s_nm_fallback = str(sclass_name or "").strip()
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if idx == 1 or not row or len(row) < 19:
                    continue
                ncs_cl_cd = _normalize_code(row[8]).strip()
                if not ncs_cl_cd or len(ncs_cl_cd) < 6 or ncs_cl_cd[:6] != code_no:
                    continue
                if ncs_cl_cd in seen:
                    continue
                seen.add(ncs_cl_cd)
                out.append(
                    {
                        "ncsClCd": ncs_cl_cd,
                        "compeUnitName": _repair_mojibake(str(row[9] or "").strip()),
                        "compeUnitLevel": _normalize_code(row[10]).strip(),
                        "ncsLclasCd": l_cd,
                        "ncsMclasCd": m_cd,
                        "ncsSclasCd": s_cd,
                        "ncsSclasCdnm": _repair_mojibake(str(row[5] or "").strip()) or s_nm_fallback,
                        "ncsSubdCd": _normalize_code(row[6], 2).strip(),
                        "ncsSubdCdnm": _repair_mojibake(str(row[7] or "").strip()),
                        "compeUnitDef": _repair_mojibake(str(row[14] or "").strip()),
                        "score": 1.0,
                        "matched_keywords": [_repair_mojibake(str(row[5] or "").strip()) or s_nm_fallback or code_no],
                    }
                )
                if len(out) >= max_items:
                    wb.close()
                    cache_map[cache_key] = list(out)
                    return out
        wb.close()
    except Exception:
        return []
    cache_map[cache_key] = list(out)
    return out


def _ksa_from_local_xlsx_by_code(ncs_cl_cd: str, limit: int = 20) -> list[dict[str, str]]:
    code = str(ncs_cl_cd or "").strip()
    if not code:
        return []

    db_path = os.getenv("APP_DB_PATH", "").strip() or _default_app_db_path()
    if ensure_ncs_local_index(db_path=db_path):
        try:
            con = _connect_local_db(db_path)
            cur = con.cursor()
            cur.execute(
                """
                SELECT
                    compe_unit_name,
                    unit_elem_name,
                    unit_criteria,
                    ksa_type_name,
                    ksa_text
                FROM ncs_local_rows
                WHERE ncs_cl_cd = ?
                """,
                (code,),
            )
            rows = cur.fetchall()
            con.close()
            out_db: list[dict[str, str]] = []
            seen_db: set[str] = set()
            for r in rows:
                ksa_text = str(r["ksa_text"] or "").strip()
                fallback = str(r["unit_elem_name"] or "").strip() or str(r["unit_criteria"] or "").strip()
                factor = ksa_text or fallback
                if not factor:
                    continue
                k = re.sub(r"\s+", "", factor)
                if k in seen_db:
                    continue
                seen_db.add(k)
                out_db.append(
                    {
                        "factorName": factor[:120],
                        "factorLevel": "",
                        "compeUnitName": str(r["compe_unit_name"] or "").strip(),
                        "factorSource": "xlsx-qs" if ksa_text else "xlsx-unit",
                    }
                )
                if len(out_db) >= max(1, int(limit or 20)):
                    break
            if out_db:
                return out_db
        except Exception:
            pass

    cache_key = f"ksa:{code}:{int(limit or 20)}"
    cache_map = _NCS_XLSX_CACHE.setdefault("map", {})
    if cache_key in cache_map:
        return list(cache_map.get(cache_key, []))

    path = os.getenv("NCS_DB_XLSX_PATH", "").strip() or _default_ncs_xlsx_path()
    if not os.path.exists(path):
        return []
    try:
        import openpyxl  # type: ignore
    except Exception:
        return []

    out: list[dict[str, str]] = []
    seen: set[str] = set()
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if idx == 1 or not row or len(row) < 19:
                    continue
                row_code = _normalize_code(row[8]).strip()
                if row_code != code:
                    continue
                ksa_text = _repair_mojibake(str(row[18] or "").strip())      # S
                fallback = _repair_mojibake(str(row[12] or "").strip()) or _repair_mojibake(str(row[14] or "").strip())
                factor = ksa_text or fallback
                if not factor:
                    continue
                key = re.sub(r"\s+", "", factor)
                if key in seen:
                    continue
                seen.add(key)
                out.append(
                    {
                        "factorName": factor[:120],
                        "factorLevel": "",
                        "compeUnitName": _repair_mojibake(str(row[9] or "").strip()),
                        "factorSource": "xlsx-qs" if ksa_text else "xlsx-unit",
                    }
                )
                if len(out) >= limit:
                    wb.close()
                    cache_map[cache_key] = list(out)
                    return out
        wb.close()
    except Exception:
        return []
    cache_map[cache_key] = list(out)
    return out


def _norm_text(v: str) -> str:
    return re.sub(r"\s+", "", _repair_mojibake(str(v or "")).strip()).lower()


def load_sclass_catalog_from_csv(csv_path: str | None = None, cache_ttl_sec: int = 60 * 30) -> list[dict[str, str]]:
    path = csv_path or os.getenv("NCS_SCLASS_CSV_PATH", "").strip() or _default_sclass_csv_path()
    now = time.time()
    if _SCLASS_CSV_CACHE.get("items") and _SCLASS_CSV_CACHE.get("path") == path:
        if (now - float(_SCLASS_CSV_CACHE.get("ts", 0.0))) < cache_ttl_sec:
            return list(_SCLASS_CSV_CACHE["items"])

    if not os.path.exists(path):
        return []

    out: list[dict[str, str]] = []
    seen = set()
    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                code_no = str((row or {}).get("NCS_CODE_NO", "")).strip()
                l_cd = str((row or {}).get("NCS_LCLAS_CD", "")).strip()
                m_cd = str((row or {}).get("NCS_MCLAS_CD", "")).strip()
                s_cd = str((row or {}).get("NCS_SCLAS_CD", "")).strip()
                s_nm = _repair_mojibake(str((row or {}).get("NCS_SCLAS_CDNM", "")).strip())
                l_nm = _repair_mojibake(str((row or {}).get("NCS_LCLAS_CDNM", "")).strip())
                m_nm = _repair_mojibake(str((row or {}).get("NCS_MCLAS_CDNM", "")).strip())
                if not (code_no and l_cd and m_cd and s_cd and s_nm):
                    continue
                key = (code_no, l_cd, m_cd, s_cd)
                if key in seen:
                    continue
                seen.add(key)
                out.append(
                    {
                        "ncs_code_no": code_no,
                        "ncs_lclass_code": l_cd,
                        "ncs_lclass_name": l_nm,
                        "ncs_mclass_code": m_cd,
                        "ncs_mclass_name": m_nm,
                        "ncs_sclass_code": s_cd,
                        "ncs_sclass_name": s_nm,
                    }
                )
    except Exception:
        return []

    _SCLASS_CSV_CACHE["ts"] = now
    _SCLASS_CSV_CACHE["path"] = path
    _SCLASS_CSV_CACHE["items"] = out
    return list(out)


def load_sclass_synonym_dictionary(
    dict_path: str | None = None,
    cache_ttl_sec: int = 60 * 30,
) -> dict[str, dict[str, list[str]]]:
    path = (
        dict_path
        or os.getenv("NCS_SCLASS_SYNONYM_PATH", "").strip()
        or _default_sclass_synonym_path()
    )
    now = time.time()

    if _SCLASS_SYNONYM_CACHE.get("path") == path and _SCLASS_SYNONYM_CACHE.get("items"):
        if (now - float(_SCLASS_SYNONYM_CACHE.get("ts", 0.0))) < cache_ttl_sec:
            cached = _SCLASS_SYNONYM_CACHE.get("items", {})
            return {
                "by_code_no": dict(cached.get("by_code_no", {})),
                "by_name": dict(cached.get("by_name", {})),
            }

    default_pack = {"by_code_no": {}, "by_name": {}}
    if not os.path.exists(path):
        return default_pack

    try:
        with open(path, "r", encoding="utf-8") as f:
            obj = json.load(f)
    except Exception:
        return default_pack

    by_code_raw = obj.get("by_code_no", {}) if isinstance(obj, dict) else {}
    by_name_raw = obj.get("by_name", {}) if isinstance(obj, dict) else {}

    by_code_no: dict[str, list[str]] = {}
    by_name: dict[str, list[str]] = {}

    if isinstance(by_code_raw, dict):
        for code_no, vals in by_code_raw.items():
            code = str(code_no or "").strip()
            if not code or not isinstance(vals, list):
                continue
            terms = [str(v).strip() for v in vals if str(v or "").strip()]
            if terms:
                by_code_no[code] = terms

    if isinstance(by_name_raw, dict):
        for name, vals in by_name_raw.items():
            nm = _norm_text(str(name or ""))
            if not nm or not isinstance(vals, list):
                continue
            terms = [str(v).strip() for v in vals if str(v or "").strip()]
            if terms:
                by_name[nm] = terms

    pack = {"by_code_no": by_code_no, "by_name": by_name}
    _SCLASS_SYNONYM_CACHE["ts"] = now
    _SCLASS_SYNONYM_CACHE["path"] = path
    _SCLASS_SYNONYM_CACHE["items"] = pack
    return {
        "by_code_no": dict(pack["by_code_no"]),
        "by_name": dict(pack["by_name"]),
    }


def ai_pick_sclass_from_csv(
    small_categories: list[str],
    subcategory_text: str,
    jd_text: str,
    max_items: int = 6,
    csv_path: str | None = None,
) -> list[dict[str, Any]]:
    catalog = load_sclass_catalog_from_csv(csv_path=csv_path)
    if not catalog:
        return []

    terms: list[str] = []
    for t in (small_categories or []):
        s = _repair_mojibake(str(t or "")).strip()
        if s and s not in terms:
            terms.append(s)
    for t in re.findall(r"[\uac00-\ud7a3]{2,12}", _repair_mojibake(subcategory_text or "")):
        if t and t not in terms:
            terms.append(t)
    for t in re.findall(r"[\uac00-\ud7a3]{2,12}", _repair_mojibake(jd_text or ""))[:50]:
        if t and t not in terms:
            terms.append(t)

    term_norm = {_norm_text(t) for t in terms if _norm_text(t)}
    if not term_norm:
        return []

    out: list[dict[str, Any]] = []
    seen = set()
    for row in catalog:
        s_nm = str(row.get("ncs_sclass_name", "")).strip()
        s_n = _norm_text(s_nm)
        if not s_n:
            continue
        exact = s_n in term_norm
        partial = any((s_n in t or t in s_n) for t in term_norm)
        if not (exact or partial):
            continue
        key = (
            str(row.get("ncs_code_no", "")).strip(),
            str(row.get("ncs_lclass_code", "")).strip(),
            str(row.get("ncs_mclass_code", "")).strip(),
            str(row.get("ncs_sclass_code", "")).strip(),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(
            {
                "sclass_name": s_nm,
                "ncs_sclass_code": str(row.get("ncs_sclass_code", "")).strip(),
                "ncs_lclass_code": str(row.get("ncs_lclass_code", "")).strip(),
                "ncs_mclass_code": str(row.get("ncs_mclass_code", "")).strip(),
                "ncs_code_no": str(row.get("ncs_code_no", "")).strip(),
                "confidence": 1.0 if exact else 0.8,
                "evidence": "csv-ncs_sclass_cdnm-match",
            }
        )
        if len(out) >= max_items:
            break
    return out


def _hrdk_base_url() -> str:
    base = os.getenv("NCS_HRDK_BASE_URL", "https://apis.data.go.kr/B490007/hrdkapi").strip()
    return base.rstrip("/")


def _try_get_json(url: str, params: dict[str, Any], timeout: float = 12.0) -> dict[str, Any] | None:
    try:
        with httpx.Client(timeout=timeout) as client:
            r = client.get(url, params=params)
        if r.status_code != 200:
            return None
        return r.json()
    except Exception:
        return None


def _extract_hrdk_items(obj: dict[str, Any]) -> list[dict[str, Any]]:
    body = obj.get("body") or ((obj.get("response") or {}).get("body") or {})
    items = (body.get("items") or {}).get("item")
    if isinstance(items, list):
        return items
    if isinstance(items, dict):
        return [items]
    return []


def _hrdk_call(op: str, extra: dict[str, Any]) -> list[dict[str, Any]]:
    base = _hrdk_base_url()
    key = settings.ncs_key()
    if not key:
        return []

    page_no = (os.getenv("NCS_PAGE_NO", "1") or "1").strip()
    num_of_rows = (os.getenv("NCS_NUM_OF_ROWS", "10") or "10").strip()
    usg_yn = (os.getenv("NCS_USG_YN", "N") or "N").strip().upper()
    ncs_degr = (os.getenv("NCS_DEGR", "22") or "22").strip()

    key_vars = [key]
    enc = quote(key, safe="")
    if enc != key:
        key_vars.append(enc)

    for key_name in ("serviceKey", "ServiceKey"):
        for kval in key_vars:
            params = {
                "pageNo": page_no,
                "numOfRows": num_of_rows,
                "USG_YN": usg_yn,
                "NCS_DEGR": ncs_degr,
                "returnType": "json",
                key_name: kval,
            }
            params.update(extra or {})
            obj = _try_get_json(f"{base}/{op}", params=params, timeout=8.0)
            if not obj:
                continue
            header = obj.get("header") or ((obj.get("response") or {}).get("header") or {})
            rc = str(header.get("resultCode", "")).strip()
            if rc and rc not in {"200", "00", "03"}:
                continue
            rows = _extract_hrdk_items(obj)
            if rows:
                return rows
    return []


def fetch_ncs_units_hrdk_by_sclass_code(
    ncs_lclass_code: str,
    ncs_mclass_code: str,
    ncs_sclass_code: str,
    sclass_name: str = "",
) -> list[dict[str, Any]]:
    l_cd = str(ncs_lclass_code or "").strip()
    m_cd = str(ncs_mclass_code or "").strip()
    s_cd = str(ncs_sclass_code or "").strip()
    s_nm = str(sclass_name or "").strip()
    if not (l_cd and m_cd and s_cd):
        return []

    # Local-first: NCS_DB.xlsx (I column prefix by code_no, e.g., 020201xxxx)
    local_rows = _units_from_local_xlsx_by_sclass(
        ncs_lclass_code=l_cd,
        ncs_mclass_code=m_cd,
        ncs_sclass_code=s_cd,
        sclass_name=s_nm,
    )
    if local_rows:
        return local_rows

    # NCS003 validation
    s_rows = _hrdk_call(
        "NCS003",
        {"NCS_LCLAS_CD": l_cd, "NCS_MCLAS_CD": m_cd, "NCS_SCLAS_CD": s_cd},
    )
    if not s_rows:
        return []
    if not s_nm:
        s_nm = str(s_rows[0].get("NCS_SCLAS_CDNM", "")).strip()

    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    subd_rows = _hrdk_call(
        "NCS004",
        {"NCS_LCLAS_CD": l_cd, "NCS_MCLAS_CD": m_cd, "NCS_SCLAS_CD": s_cd},
    )
    for subd in subd_rows:
        subd_cd = str(subd.get("NCS_SUBD_CD", "")).strip()
        subd_nm = str(subd.get("NCS_SUBD_CDNM", "")).strip()
        if not subd_cd:
            continue
        units = _hrdk_call(
            "NCS005",
            {
                "NCS_LCLAS_CD": l_cd,
                "NCS_MCLAS_CD": m_cd,
                "NCS_SCLAS_CD": s_cd,
                "NCS_SUBD_CD": subd_cd,
            },
        )
        for u in units:
            cl = str(u.get("NCS_CL_CD", "")).strip()
            if not cl or cl in seen:
                continue
            seen.add(cl)
            out.append(
                {
                    "ncsClCd": cl,
                    "compeUnitName": str(u.get("COMPE_UNIT_NAME", "")).strip(),
                    "compeUnitLevel": str(u.get("COMPE_UNIT_LEVEL", "")).strip(),
                    "ncsLclasCd": l_cd,
                    "ncsMclasCd": m_cd,
                    "ncsSclasCd": s_cd,
                    "ncsSclasCdnm": str(u.get("NCS_SCLAS_CDNM", "")).strip() or s_nm,
                    "ncsSubdCd": subd_cd,
                    "ncsSubdCdnm": str(u.get("NCS_SUBD_CDNM", "")).strip() or subd_nm,
                    "compeUnitDef": str(u.get("COMPE_UNIT_DEF", "")).strip(),
                    "score": 1.0,
                    "matched_keywords": [s_nm or s_cd],
                }
            )
    return out


def fetch_ncs_units_hrdk_by_verified_sclass(verified_sclass: list[dict[str, Any]], max_sclass: int = 6) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for v in (verified_sclass or [])[:max_sclass]:
        l_cd = str((v or {}).get("ncs_lclass_code", "")).strip()
        m_cd = str((v or {}).get("ncs_mclass_code", "")).strip()
        s_cd = str((v or {}).get("ncs_sclass_code", "")).strip()
        code_no = str((v or {}).get("ncs_code_no", "")).strip()
        if (not l_cd or not m_cd or not s_cd) and len(code_no) >= 6 and code_no.isdigit():
            l_cd, m_cd, s_cd = code_no[:2], code_no[2:4], code_no[4:6]
        s_nm = str((v or {}).get("sclass_name", "")).strip()
        items = fetch_ncs_units_hrdk_by_sclass_code(
            ncs_lclass_code=l_cd,
            ncs_mclass_code=m_cd,
            ncs_sclass_code=s_cd,
            sclass_name=s_nm,
        )
        for it in items:
            code = str(it.get("ncsClCd", "")).strip()
            if not code or code in seen:
                continue
            seen.add(code)
            out.append(it)
    return out


def fetch_ncs_units_hrdk_by_sclass_names(sclass_names: list[str], max_sclass: int = 6) -> list[dict[str, Any]]:
    csv_hits = ai_pick_sclass_from_csv(
        small_categories=list(sclass_names or []),
        subcategory_text=" ".join(sclass_names or []),
        jd_text="",
        max_items=max_sclass,
    )
    out = fetch_ncs_units_hrdk_by_verified_sclass(csv_hits, max_sclass=max_sclass)
    if out:
        return out

    # Fallback: keyword search from NCS007 and then resolve tuple
    hits: list[dict[str, Any]] = []
    seen = set()
    for term in (sclass_names or [])[:max_sclass]:
        rows = _hrdk_call("NCS007", {"LVL": "4", "SWRD": term, "SNUM": "1", "ENUM": "100"})
        for r in rows:
            l_cd = str(r.get("NCS_LCLAS_CD", "")).strip()
            m_cd = str(r.get("NCS_MCLAS_CD", "")).strip()
            s_cd = str(r.get("NCS_SCLAS_CD", "")).strip()
            if not (l_cd and m_cd and s_cd):
                continue
            key = (l_cd, m_cd, s_cd)
            if key in seen:
                continue
            seen.add(key)
            hits.append(
                {
                    "sclass_name": str(r.get("NCS_SCLAS_CDNM", "")).strip() or str(term).strip(),
                    "ncs_lclass_code": l_cd,
                    "ncs_mclass_code": m_cd,
                    "ncs_sclass_code": s_cd,
                    "confidence": 0.8,
                    "evidence": "ncs007-keyword",
                }
            )
    return fetch_ncs_units_hrdk_by_verified_sclass(hits, max_sclass=max_sclass)


def fetch_ncs_units_hrdk_by_keywords(keywords: list[str], max_items: int = 20) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen = set()
    for kw in (keywords or [])[:10]:
        rows = _hrdk_call("NCS007", {"LVL": "6", "SWRD": str(kw).strip(), "SNUM": "1", "ENUM": "80"})
        for r in rows:
            cl = str(r.get("NCS_CL_CD", "")).strip()
            if not cl or cl in seen:
                continue
            seen.add(cl)
            out.append(
                {
                    "ncsClCd": cl,
                    "compeUnitName": str(r.get("COMPE_UNIT_NAME", "")).strip(),
                    "compeUnitLevel": str(r.get("COMPE_UNIT_LEVEL", "")).strip(),
                    "ncsSclasCdnm": str(r.get("NCS_SCLAS_CDNM", "")).strip(),
                    "ncsSubdCdnm": str(r.get("NCS_SUBD_CDNM", "")).strip(),
                    "compeUnitDef": str(r.get("COMPE_UNIT_DEF", "")).strip(),
                    "score": 0.7,
                    "matched_keywords": [str(kw).strip()],
                }
            )
            if len(out) >= max_items:
                return out
    return out


def fetch_ncs_units_hrdk_by_cl_codes(code_rows: list[dict[str, Any]], max_items: int = 20) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen = set()
    for row in (code_rows or [])[: max_items * 2]:
        code = str(row.get("ncsClCd", row.get("ncs_cl_cd", ""))).strip()
        if not code or code in seen:
            continue
        rows = _hrdk_call("NCS007", {"LVL": "6", "SWRD": code, "SNUM": "1", "ENUM": "20"})
        picked = None
        for r in rows:
            if str(r.get("NCS_CL_CD", "")).strip() == code:
                picked = r
                break
        if not picked and rows:
            picked = rows[0]
        if not picked:
            continue
        seen.add(code)
        out.append(
            {
                "ncsClCd": code,
                "compeUnitName": str(picked.get("COMPE_UNIT_NAME", "")).strip(),
                "compeUnitLevel": str(picked.get("COMPE_UNIT_LEVEL", "")).strip(),
                "ncsSclasCdnm": str(picked.get("NCS_SCLAS_CDNM", "")).strip(),
                "ncsSubdCdnm": str(picked.get("NCS_SUBD_CDNM", "")).strip(),
                "compeUnitDef": str(picked.get("COMPE_UNIT_DEF", "")).strip(),
                "score": 0.9,
                "matched_keywords": [code],
            }
        )
        if len(out) >= max_items:
            break
    return out


def fetch_ncs_ksa_by_units(
    ncs_matches: list[dict[str, Any]],
    max_units: int = 5,
    max_factors_per_unit: int = 3,
    use_ncs007_fallback: bool | None = None,
) -> list[dict[str, Any]]:
    # MVP policy: KSA must come from the compact read-only NCS_MCP serving DB.
    # Do not silently fall back to NCS_DB.xlsx, public HRDK endpoints, or
    # definition-derived pseudo factors. If MCP is unavailable, generation must
    # fail or switch to the explicit manual-selection flow.
    if not settings.ncs_mcp_endpoint():
        raise NcsMcpError("NCS_MCP_URL is required for official KSA lookup")

    mcp_rows = get_ksa_by_units(
        list(ncs_matches or [])[: max(1, int(max_units or 5))],
        max_factors_per_unit=max(1, int(max_factors_per_unit or 3)),
    )
    if not mcp_rows:
        raise NcsMcpError("NCS MCP returned no official KSA rows")
    return mcp_rows


def _compact_text_for_tfidf(text: str) -> str:
    cleaned = _repair_mojibake(str(text or ""))
    cleaned = re.sub(r"\s+", "", cleaned).lower()
    return re.sub(r"[^0-9a-z\uac00-\ud7a3]", "", cleaned)


def _char_ngram_tf(text: str, ngram_min: int = 2, ngram_max: int = 4) -> Counter[str]:
    src = _compact_text_for_tfidf(text)
    if not src:
        return Counter()

    grams: list[str] = []
    lo = max(1, int(ngram_min or 2))
    hi = max(lo, int(ngram_max or 4))
    for n in range(lo, hi + 1):
        if len(src) < n:
            continue
        for i in range(0, len(src) - n + 1):
            grams.append(src[i : i + n])

    if not grams:
        grams = [src]
    return Counter(grams)


def rank_ksa_factors_by_query(
    ksa_rows: list[dict[str, Any]],
    query_text: str,
    unit_scores: dict[str, float] | None = None,
    target_count: int = 12,
    per_unit_limit: int = 3,
    similarity_weight: float = 0.75,
    unit_weight: float = 0.25,
    ngram_min: int = 2,
    ngram_max: int = 4,
) -> list[dict[str, Any]]:
    rows = [dict(x) for x in (ksa_rows or []) if isinstance(x, dict)]
    if not rows:
        return []

    keep_n = max(1, min(50, int(target_count or 12)))
    per_unit_cap = max(1, min(6, int(per_unit_limit or 2)))

    sim_w = max(0.0, float(similarity_weight or 0.0))
    unit_w = max(0.0, float(unit_weight or 0.0))
    if sim_w <= 0 and unit_w <= 0:
        sim_w = 1.0
    total_w = sim_w + unit_w
    sim_w = sim_w / total_w
    unit_w = unit_w / total_w

    unit_raw_scores: dict[str, float] = {}
    if unit_scores:
        for code, v in unit_scores.items():
            c = str(code or "").strip()
            if not c:
                continue
            try:
                unit_raw_scores[c] = float(v or 0.0)
            except Exception:
                unit_raw_scores[c] = 0.0

    deduped_rows: list[dict[str, Any]] = []
    seen_factors: set[tuple[str, str]] = set()
    for row in rows:
        code = str(row.get("ncsClCd", "")).strip()
        factor = str(row.get("factorName", "")).strip()
        if not code or not factor:
            continue
        dedup_key = (code, re.sub(r"\s+", "", factor).lower())
        if dedup_key in seen_factors:
            continue
        seen_factors.add(dedup_key)
        deduped_rows.append(row)
    if not deduped_rows:
        return []

    if not unit_raw_scores:
        for row in deduped_rows:
            code = str(row.get("ncsClCd", "")).strip()
            if not code:
                continue
            try:
                base_score = float(row.get("score", 1.0) or 1.0)
            except Exception:
                base_score = 1.0
            prev = unit_raw_scores.get(code)
            if prev is None or base_score > prev:
                unit_raw_scores[code] = base_score

    score_values = list(unit_raw_scores.values()) or [1.0]
    score_min = min(score_values)
    score_max = max(score_values)

    def _norm_unit_score(code: str) -> float:
        v = float(unit_raw_scores.get(code, 0.0))
        if score_max > score_min:
            return (v - score_min) / (score_max - score_min)
        if 0.0 <= v <= 1.0:
            return v
        return 1.0 if v > 0 else 0.0

    query_tf = _char_ngram_tf(query_text, ngram_min=ngram_min, ngram_max=ngram_max)
    doc_tfs: list[Counter[str]] = []
    for row in deduped_rows:
        text = f"{str(row.get('factorName', '')).strip()} {str(row.get('compeUnitName', '')).strip()}"
        doc_tfs.append(_char_ngram_tf(text, ngram_min=ngram_min, ngram_max=ngram_max))

    similarity_scores: list[float] = [0.0] * len(deduped_rows)
    if query_tf and any(doc_tfs):
        df: Counter[str] = Counter()
        for tf in doc_tfs:
            df.update(tf.keys())

        doc_count = max(1, len(doc_tfs))
        idf = {term: (math.log((doc_count + 1) / (freq + 1)) + 1.0) for term, freq in df.items()}

        query_w: dict[str, float] = {}
        for term, cnt in query_tf.items():
            if term not in idf:
                continue
            query_w[term] = (1.0 + math.log(max(1, cnt))) * idf[term]
        query_norm = math.sqrt(sum(v * v for v in query_w.values())) if query_w else 0.0

        if query_norm > 0:
            for i, tf in enumerate(doc_tfs):
                if not tf:
                    continue
                dot = 0.0
                doc_norm_sq = 0.0
                for term, cnt in tf.items():
                    weight = (1.0 + math.log(max(1, cnt))) * idf.get(term, 0.0)
                    if weight <= 0:
                        continue
                    doc_norm_sq += weight * weight
                    qv = query_w.get(term)
                    if qv:
                        dot += qv * weight
                doc_norm = math.sqrt(doc_norm_sq)
                if doc_norm > 0 and dot > 0:
                    similarity_scores[i] = dot / (query_norm * doc_norm)

    scored_rows: list[dict[str, Any]] = []
    for i, row in enumerate(deduped_rows):
        code = str(row.get("ncsClCd", "")).strip()
        sim = float(similarity_scores[i])
        unit_score_norm = _norm_unit_score(code)
        final_score = (sim_w * sim) + (unit_w * unit_score_norm)
        merged = dict(row)
        merged["__idx"] = i
        merged["similarityScore"] = round(sim, 6)
        merged["unitScore"] = round(unit_score_norm, 6)
        merged["finalScore"] = round(final_score, 6)
        scored_rows.append(merged)

    scored_rows.sort(
        key=lambda x: (
            float(x.get("finalScore", 0.0) or 0.0),
            float(x.get("similarityScore", 0.0) or 0.0),
            float(x.get("unitScore", 0.0) or 0.0),
            -int(x.get("__idx", 0) or 0),
        ),
        reverse=True,
    )

    def _ksa_kind(row: dict[str, Any]) -> str:
        raw = re.sub(
            r"\s+",
            "",
            str(
                row.get("ksaTypeName")
                or row.get("factorType")
                or row.get("ksa_type")
                or row.get("ksa_type_name")
                or ""
            ),
        ).lower()
        if raw in {"k", "knowledge", "지식"} or "지식" in raw:
            return "지식"
        if raw in {"s", "skill", "skills", "기술"} or any(token in raw for token in ("기술", "스킬")):
            return "기술"
        if raw in {"a", "attitude", "태도"} or "태도" in raw:
            return "태도"
        return ""

    available_types_by_unit: dict[str, set[str]] = {}
    for row in scored_rows:
        code = str(row.get("ncsClCd", "")).strip()
        kind = _ksa_kind(row)
        if code and kind:
            available_types_by_unit.setdefault(code, set()).add(kind)

    selected: list[dict[str, Any]] = []
    per_unit_count: dict[str, int] = {}
    selected_types_by_unit: dict[str, set[str]] = {}
    for row in scored_rows:
        code = str(row.get("ncsClCd", "")).strip()
        if not code:
            continue
        current_count = per_unit_count.get(code, 0)
        if current_count >= per_unit_cap:
            continue
        kind = _ksa_kind(row)
        selected_types = selected_types_by_unit.setdefault(code, set())
        unseen_types = available_types_by_unit.get(code, set()) - selected_types
        remaining_slots = per_unit_cap - current_count
        if unseen_types and remaining_slots <= len(unseen_types) and kind not in unseen_types:
            continue
        row.pop("__idx", None)
        selected.append(row)
        per_unit_count[code] = current_count + 1
        if kind:
            selected_types.add(kind)
        if len(selected) >= keep_n:
            break
    return selected


def fetch_ncs_ksa_by_sclass_code(
    ncs_lclass_code: str,
    ncs_mclass_code: str,
    ncs_sclass_code: str,
    sclass_name: str = "",
    max_units: int = 80,
) -> dict[str, Any]:
    units = fetch_ncs_units_hrdk_by_sclass_code(
        ncs_lclass_code=ncs_lclass_code,
        ncs_mclass_code=ncs_mclass_code,
        ncs_sclass_code=ncs_sclass_code,
        sclass_name=sclass_name,
    )
    limited = units[: max(1, int(max_units or 80))]
    ksa = fetch_ncs_ksa_by_units(ncs_matches=limited, max_units=len(limited))
    return {"units": limited, "ksa": ksa}


def resolve_sclass_candidates_with_catalog(
    candidates: list[dict[str, Any]],
    fallback_terms: list[str] | None = None,
    max_terms: int = 8,
) -> list[dict[str, Any]]:
    catalog = load_sclass_catalog_from_csv()
    if not catalog:
        return []

    terms: list[str] = []
    for c in (candidates or []):
        nm = str((c or {}).get("sclass_name", "")).strip()
        if nm and nm not in terms:
            terms.append(nm)
    for t in (fallback_terms or []):
        s = str(t).strip()
        if s and s not in terms:
            terms.append(s)

    out: list[dict[str, Any]] = []
    seen = set()
    for term in terms[:max_terms]:
        q = _norm_text(term)
        best = None
        best_score = 0.0
        for row in catalog:
            nm = str(row.get("ncs_sclass_name", "")).strip()
            n = _norm_text(nm)
            if not n:
                continue
            if q == n:
                score = 1.0
            elif q and (q in n or n in q):
                score = 0.86
            else:
                score = SequenceMatcher(None, q, n).ratio()
            if score > best_score:
                best_score = score
                best = row
        if not best or best_score < 0.62:
            continue
        key = (
            str(best.get("ncs_lclass_code", "")).strip(),
            str(best.get("ncs_mclass_code", "")).strip(),
            str(best.get("ncs_sclass_code", "")).strip(),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(
            {
                "sclass_name": str(best.get("ncs_sclass_name", "")).strip(),
                "ncs_sclass_code": key[2],
                "ncs_lclass_code": key[0],
                "ncs_mclass_code": key[1],
                "ncs_code_no": str(best.get("ncs_code_no", "")).strip(),
                "confidence": float(best_score),
                "evidence": f"catalog-fuzzy:{term}",
            }
        )
    return out


def verify_sclass_candidates_with_ncs_api(candidates: list[dict[str, Any]], max_terms: int = 6) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen = set()
    for c in (candidates or [])[:max_terms]:
        l_cd = str((c or {}).get("ncs_lclass_code", "")).strip()
        m_cd = str((c or {}).get("ncs_mclass_code", "")).strip()
        s_cd = str((c or {}).get("ncs_sclass_code", "")).strip()
        code_no = str((c or {}).get("ncs_code_no", "")).strip()
        if (not l_cd or not m_cd or not s_cd) and len(code_no) >= 6 and code_no.isdigit():
            l_cd, m_cd, s_cd = code_no[:2], code_no[2:4], code_no[4:6]
        if not (l_cd and m_cd and s_cd):
            continue
        rows = _hrdk_call(
            "NCS003",
            {"NCS_LCLAS_CD": l_cd, "NCS_MCLAS_CD": m_cd, "NCS_SCLAS_CD": s_cd},
        )
        if not rows:
            continue
        key = (l_cd, m_cd, s_cd)
        if key in seen:
            continue
        seen.add(key)
        s_nm = str(rows[0].get("NCS_SCLAS_CDNM", "")).strip() or str((c or {}).get("sclass_name", "")).strip()
        out.append(
            {
                "sclass_name": s_nm,
                "ncs_sclass_code": s_cd,
                "ncs_lclass_code": l_cd,
                "ncs_mclass_code": m_cd,
                "ncs_code_no": code_no,
                "confidence": float((c or {}).get("confidence", 1.0) or 1.0),
                "evidence": "ncs003-verified",
            }
        )
    return out


def ai_extract_sclass_candidates(
    subcategory_text: str,
    jd_text: str,
    seed_terms: list[str] | None = None,
    max_items: int = 8,
) -> list[dict[str, Any]]:
    return resolve_sclass_candidates_with_catalog(
        candidates=[{"sclass_name": t, "confidence": 0.7} for t in (seed_terms or []) if str(t).strip()],
        fallback_terms=re.findall(r"[\uac00-\ud7a3]{2,12}", _repair_mojibake(subcategory_text or "")),
        max_terms=max_items,
    )


def ai_extract_ncs_cl_codes(seed_terms: list[str], jd_text: str, max_items: int = 8) -> list[dict[str, Any]]:
    text = " ".join([str(x) for x in (seed_terms or [])]) + " " + str(jd_text or "")
    candidates = re.findall(r"\b\d{8,12}\b", text)
    out: list[dict[str, Any]] = []
    seen = set()
    for c in candidates:
        if c in seen:
            continue
        seen.add(c)
        out.append({"ncsClCd": c, "confidence": 0.8})
        if len(out) >= max_items:
            break
    return out


def review_ocr_terms_with_openai(terms: list[str], jd_text: str) -> list[str]:
    out: list[str] = []
    for t in (terms or []):
        s = _repair_mojibake(str(t or "")).strip()
        if not s or s in out:
            continue
        out.append(s)
        if len(out) >= 20:
            break
    return out


def build_notice_context_from_jd(jd_text: str, notice_text: str = "", max_chars: int = 5000) -> str:
    note = _repair_mojibake(str(notice_text or "")).strip()
    if not note:
        return ""
    return note[: max(200, int(max_chars or 5000))]


def build_ncs_context_pack(
    jd_text: str,
    notice_text: str,
    ncs_items: list[dict[str, Any]],
    ncs_matches: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "jd_preview": _repair_mojibake(jd_text or "")[:1200],
        "notice_preview": _repair_mojibake(notice_text or "")[:1200],
        "ncs_item_count": len(ncs_items or []),
        "ncs_match_count": len(ncs_matches or []),
        "top_ncs": [
            {
                "ncsClCd": str(x.get("ncsClCd", "")).strip(),
                "compeUnitName": str(x.get("compeUnitName", "")).strip(),
                "compeUnitDef": str(x.get("compeUnitDef", "")).strip()[:500],
            }
            for x in (ncs_matches or [])[:8]
        ],
    }


def diagnose_ncs_hrdk() -> dict[str, Any]:
    key = settings.ncs_key()
    if not key:
        return {"ok": False, "message": "NCS key is missing.", "endpoint": _hrdk_base_url(), "samples": []}

    base = _hrdk_base_url()
    samples: list[dict[str, Any]] = []
    for key_name in ("serviceKey", "ServiceKey"):
        for kval in (key, quote(key, safe="")):
            params = {"pageNo": "1", "numOfRows": "3", "returnType": "json", key_name: kval}
            try:
                with httpx.Client(timeout=10.0) as client:
                    r = client.get(f"{base}/NCS001", params=params)
                row: dict[str, Any] = {"key_name": key_name, "status": r.status_code, "preview": (r.text or "")[:300]}
                if r.status_code == 200:
                    try:
                        obj = r.json()
                    except Exception:
                        obj = {}
                    rows = _extract_hrdk_items(obj) if obj else []
                    row["count"] = len(rows)
                    header = obj.get("header") or ((obj.get("response") or {}).get("header") or {})
                    row["resultCode"] = str(header.get("resultCode", ""))
                    row["resultMsg"] = str(header.get("resultMsg", ""))
                    if rows:
                        return {"ok": True, "message": "HRDK NCS API is reachable.", "endpoint": base, "samples": samples + [row]}
                samples.append(row)
            except Exception as e:
                samples.append({"key_name": key_name, "status": None, "error": str(e)})
    return {"ok": False, "message": "HRDK NCS API call failed.", "endpoint": base, "samples": samples}


def diagnose_ncs_v18_flow(sample_job_cd: str = "02020101") -> dict[str, Any]:
    return {
        "ok": False,
        "message": "V1.8 flow is not used in this pipeline. Use HRDK /NCS003~/NCS005 and optional /NCS006.",
        "steps": [{"sample_job_cd": sample_job_cd}],
    }


def fetch_ncs_units_v18_by_sclass(ncs_sclass_code: str, sclass_name: str = "", max_items: int = 50) -> list[dict[str, Any]]:
    _ = (ncs_sclass_code, sclass_name, max_items)
    return []






